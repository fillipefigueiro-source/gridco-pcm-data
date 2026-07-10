"""
Atualização Semanal - Sync com BD + Sugestões PCM
==================================================

Companion ao `programacao_v7.py`. Roda VÁRIAS VEZES AO LONGO DO DIA com 2
responsabilidades independentes:

  STEP A - SINCRONIZAR a Programação Semanal com o BD_Relatório Semanal:
    - Atualiza coluna "Status Atual (BD)" das OS já programadas
    - Detecta OS novas no BD (semana corrente) que não estão na Programação:
        * Se a OS veio de uma Solicitação (campo "Número de Solicitação"),
          marca como "Nova OS (via Solic #X)" e atualiza o Sugestoes_PCM.xlsx
        * Se não veio de solicitação, marca como "Nova OS (BD)"
    - Usa SEMPRE a Data Programada e horários DEFINIDOS PELO PROGRAMADOR no BD
      (o script NÃO sobrescreve a data/hora que o programador colocou)

  STEP B - SUGERIR datas/horas para o PCM em Sugestoes_PCM.xlsx:
    - Lê as Solicitações em status "Aberto" sem Ordem de Serviço
    - Calcula prioridade (Criticidade + Urgência + RPN da Lista de Prioridades)
    - Estima duração via Planilha de Confiabilidade (default 2h)
    - Sugere equipe, dia e horário com base na agenda da semana
    - Grava arquivo Sugestoes_PCM.xlsx (sister-file da planilha de solicitações)
    - O programador olha esse arquivo, cria a OS no SISTEMA com a data desejada
    - Na próxima execução, STEP A detecta a OS criada e incorpora ao semanal

NÃO há mais "Aceite (S/N)" - o aceite é a OS aparecer no BD.

Modo de execução: SINGLE-SHOT (roda uma vez e termina). Use o Agendador de
Tarefas do Windows para chamá-lo a cada X minutos.

Uso:
    python atualizacao_semanal.py                    # roda tudo
    python atualizacao_semanal.py --no-sync          # pula STEP A
    python atualizacao_semanal.py --no-sugestoes     # pula STEP B
    python atualizacao_semanal.py --semana 22        # força semana específica
"""

import argparse
import os
import sys as _sys_init
# Forca stdout/stderr UTF-8 (corrige UnicodeEncodeError no Windows cp1252)
try:
    _sys_init.stdout.reconfigure(encoding='utf-8', errors='replace')
    _sys_init.stderr.reconfigure(encoding='utf-8', errors='replace')
except (AttributeError, Exception):
    pass
import os
import re
import shutil
import sys
import unicodedata
import fonte_bd_api  # fonte de dados via API Fracttal (substitui leitura do BD xlsx)
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ====================== Configuração ======================

import pathlib


def _detectar_base_dir():
    """Encontra a pasta '09. Programação Semanal' tentando em ordem:
       1. Variável de ambiente PCM_PROG_DIR
       2. Diretório do próprio .py (caso esteja COPIADO pra dentro da pasta)
       3. Vários layouts comuns de OneDrive/SharePoint no Windows
       4. Mount FUSE do OneDrive (Cowork/Linux)
       Retorna o caminho ou None.
    """
    # 1. Env var (sobrescreve tudo)
    env_val = os.environ.get("PCM_PROG_DIR", "").strip()
    if env_val and os.path.isdir(env_val):
        return env_val

    # 2. Diretório do próprio script — útil se .py estiver na pasta
    try:
        script_dir = pathlib.Path(__file__).resolve().parent
        if (script_dir / "BD_Relatório Semanal.xlsx").exists():
            return str(script_dir)
    except NameError:
        pass

    _subpath = pathlib.Path("4. O&M") / "11.Pré-Operação" / "6. PCM" / "09. Programação Semanal"
    _subpath_pcm = pathlib.Path("6. PCM") / "09. Programação Semanal"
    home = pathlib.Path.home()

    # 3. Windows/OneDrive — todos os layouts conhecidos
    candidatos = [
        # Layout principal (sync da raiz SharePoint Grid Co. - Gridco)
        home / "GRID CO" / "Grid Co. - Gridco" / _subpath,
        home / "Grid Co. - Gridco" / _subpath,
        # OneDrive tenant variants
        home / "OneDrive - Grid Co" / _subpath,
        home / "OneDrive - Gridco" / _subpath,
        home / "OneDrive - GRID CO" / _subpath,
        # Dentro de OneDrive pessoal
        home / "OneDrive" / "Grid Co. - Gridco" / _subpath,
        home / "OneDrive" / _subpath,
        # Documents
        home / "Documents" / "Grid Co. - Gridco" / _subpath,
        home / "Documentos" / "Grid Co. - Gridco" / _subpath,
        # Caso usuário sincronize só "6. PCM" (subnível do SharePoint)
        home / "GRID CO" / "Grid Co. - Gridco" / _subpath_pcm,
        home / "Grid Co. - Gridco" / _subpath_pcm,
        home / "OneDrive - Grid Co" / _subpath_pcm,
        home / "OneDrive - Gridco" / _subpath_pcm,
    ]
    for c in candidatos:
        try:
            if c.exists() and c.is_dir():
                return str(c)
        except (OSError, PermissionError):
            continue

    # 4. Cowork/Linux — mount FUSE
    try:
        sessions = pathlib.Path("/sessions")
        if sessions.exists():
            for mnt in sessions.glob("*/mnt/09. Programação Semanal"):
                return str(mnt)
    except (OSError, PermissionError):
        pass

    return None


_DETECTED = _detectar_base_dir()
if _DETECTED is None:
    print("=" * 72, flush=True)
    print("ERRO: Pasta '09. Programação Semanal' não encontrada nesta máquina.", flush=True)
    print("=" * 72, flush=True)
    print("Soluções (faça apenas UMA das opções abaixo):", flush=True)
    print("", flush=True)
    print("  A) Sincronize a pasta SharePoint via OneDrive (recomendado):", flush=True)
    print("     - Abra SharePoint > Grid Co. - Gridco > 4. O&M > 11.Pré-Operação", flush=True)
    print("       > 6. PCM > 09. Programação Semanal", flush=True)
    print("     - Clique em 'Sync' / 'Sincronizar'", flush=True)
    print("     - Aguarde aparecer em C:\\Users\\<seu_usuario>\\GRID CO\\...", flush=True)
    print("", flush=True)
    print("  B) Defina a variável de ambiente PCM_PROG_DIR antes de rodar:", flush=True)
    print("     PowerShell:  $env:PCM_PROG_DIR = 'C:\\caminho\\completo'", flush=True)
    print("     CMD:         set PCM_PROG_DIR=C:\\caminho\\completo", flush=True)
    print("", flush=True)
    print("  C) Copie atualizacao_semanal.py PARA DENTRO da pasta", flush=True)
    print("     '09. Programação Semanal' e rode de lá.", flush=True)
    print("=" * 72, flush=True)
    sys.exit(1)

BASE_DIR = _DETECTED
print(f"[INFO] Pasta detectada: {BASE_DIR}", flush=True)

INPUT_SOLICIT = os.path.join(BASE_DIR, "Exportacao_Solicitacoes_de_Servicos.xlsx")
INPUT_BD = os.path.join(BASE_DIR, "BD_Relatório Semanal.xlsx")
INPUT_PRIOR = os.path.join(BASE_DIR, "Lista_Prioridades_GridCo.xlsx")
INPUT_CONFIAB = os.path.join(BASE_DIR, "Planilha Confiabilidade R00.xlsx")
OUTPUT_SUGEST = os.path.join(BASE_DIR, "Sugestoes_PCM.xlsx")

DEFAULT_DURACAO_H = 2.0
_MAP_RESPONSAVEL = {}  # populado em main() via carregar_map_responsavel()
HORA_INICIO_PADRAO = "07:30"
HORA_FIM_TURNO_MIN = 17 * 60
HORA_INICIO_TURNO_MIN = 7 * 60
HORA_ALMOCO_INI_MIN = 12 * 60
HORA_ALMOCO_FIM_MIN = 13 * 60 + 12
GAP_ENTRE_OS_MIN = 15

SCORE_CRITICIDADE = {
    "Muito alto": 100, "Alto": 75, "Médio": 50, "Medio": 50,
    "Baixo": 25, None: 10, "": 10,
}
SCORE_URGENCIA = {"Sim": 50, "sim": 50, "Não": 0, "Nao": 0, None: 0, "": 0}

COLS_EQUIPE = [
    "Equipe", "Dia", "OSs ID", "Ativo (Usina)", "Cidade", "Código Equipamento",
    "Tipo", "Tarefa", "Estado Tarefa (antes)", "RPN/Prioridade", "Reprogramada",
    "Nº vezes programada", "Termografia", "Hora Início", "Hora Fim",
    "Duração (h)", "Desloc (h)", "Paralelo (terceirizada)",
]

COLS_EXTRA_EQUIPE = ["Responsável", "Status Atual (BD)", "Nova OS", "Nº Solicitação Origem",
                     "Deslocada (Original)", "Última Atualização"]

COLS_SUGESTOES = [
    "Nº Solicitação", "Equipe", "Ativo", "Código Equipamento", "Descrição",
    "Criticidade", "Urgente?", "RPN", "Categoria RPN", "Score",
    "Dia Sugerido", "Hora Início Sugerida", "Duração (h)",
    "Status no PCM", "Observação", "Criado em", "Última Avaliação",
    "OS Criada no Sistema",
]

DAY_NAMES_FULL = {
    0: "Segunda-feira", 1: "Terça-feira", 2: "Quarta-feira",
    3: "Quinta-feira", 4: "Sexta-feira", 5: "Sábado", 6: "Domingo",
}

# ====================== Utilitários ======================

def _strip_accents(s):
    if not isinstance(s, str):
        return s
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def _norm(s):
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    s = _strip_accents(str(s)).lower().strip()
    return re.sub(r"\s+", " ", s)


def _semana_atual(ref=None):
    ref = ref or date.today()
    mon = ref - timedelta(days=ref.weekday())
    _, iso_week, _ = mon.isocalendar()
    return mon, iso_week


def _hora_min_to_str(m):
    h, mi = divmod(int(m), 60)
    return f"{h:02d}:{mi:02d}"


def _hora_str_to_min(s):
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return None
    if isinstance(s, (int, float)):
        return int(round(float(s) * 60))
    if isinstance(s, str):
        m = re.match(r"(\d{1,2}):(\d{2})", s.strip())
        if m:
            return int(m.group(1)) * 60 + int(m.group(2))
    try:
        return s.hour * 60 + s.minute
    except Exception:
        return None


def log(msg, level="INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {level:5s} {msg}", flush=True)


# ====================== Carregamento ======================

def _read_excel_robusto(path, max_tentativas=5, **kwargs):
    """Tenta ler arquivo Excel com retry e backoff exponencial.
    Cobre falhas transitórias: OneDrive sync, arquivo travado, zip corrompido.
    Atrasos: 3, 6, 12, 24, 48 segundos (total ~93s)."""
    import time
    if not os.path.exists(path):
        log(f"  ! Arquivo não existe: {path}", "ERROR")
        return None
    last_err = None
    for tentativa in range(1, max_tentativas + 1):
        try:
            return pd.read_excel(path, **kwargs)
        except Exception as e:
            last_err = e
            log(f"  ! Tentativa {tentativa}/{max_tentativas} ler {os.path.basename(path)}: {type(e).__name__}: {e}", "WARN")
            if tentativa < max_tentativas:
                wait = 3 * (2 ** (tentativa - 1))
                log(f"    aguardando {wait}s antes de tentar de novo...", "WARN")
                time.sleep(wait)
    log(f"  ! Não foi possível ler {os.path.basename(path)} após {max_tentativas} tentativas: {last_err}", "ERROR")
    return None


def _save_workbook_robusto(wb, path, max_tentativas=5, etapa="arquivo"):
    """Tenta salvar workbook com retry exponencial. Cobre PermissionError
    (Excel aberto), OSError (rede). Atrasos: 5, 10, 20, 40, 80s (~155s)."""
    import time
    last_err = None
    for tentativa in range(1, max_tentativas + 1):
        try:
            wb.save(path)
            return True
        except PermissionError as e:
            last_err = e
            log(f"  ! Tentativa {tentativa}/{max_tentativas} salvar {etapa} bloqueada (Excel aberto?): {e}", "WARN")
        except Exception as e:
            last_err = e
            log(f"  ! Tentativa {tentativa}/{max_tentativas} salvar {etapa}: {type(e).__name__}: {e}", "WARN")
        if tentativa < max_tentativas:
            wait = 5 * (2 ** (tentativa - 1))
            log(f"    aguardando {wait}s antes de tentar de novo...", "WARN")
            time.sleep(wait)
    log(f"  ! NÃO foi possível salvar {etapa} ({os.path.basename(path)}) após {max_tentativas} tentativas.", "ERROR")
    log(f"    Último erro: {last_err}", "ERROR")
    log(f"    Causas comuns: arquivo aberto no Excel, OneDrive em sync, permissão negada.", "ERROR")
    log(f"    O script vai continuar — a próxima execução vai tentar de novo.", "WARN")
    return False


def _load_workbook_robusto(path, max_tentativas=5):
    """Tenta carregar workbook com retry. Cobre arquivos corrompidos
    temporariamente por OneDrive sync."""
    import time
    if not os.path.exists(path):
        log(f"  ! Arquivo não existe: {path}", "ERROR")
        return None
    last_err = None
    for tentativa in range(1, max_tentativas + 1):
        try:
            return load_workbook(path)
        except Exception as e:
            last_err = e
            log(f"  ! Tentativa {tentativa}/{max_tentativas} abrir {os.path.basename(path)}: {type(e).__name__}: {e}", "WARN")
            if tentativa < max_tentativas:
                wait = 3 * (2 ** (tentativa - 1))
                log(f"    aguardando {wait}s antes de tentar de novo...", "WARN")
                time.sleep(wait)
    log(f"  ! NÃO foi possível abrir {os.path.basename(path)} após {max_tentativas} tentativas: {last_err}", "ERROR")
    return None


def _aquire_lock(arquivo_prog):
    """Lockfile impede execuções paralelas. Lockfile > 30 min é considerado stale."""
    import time
    lock_path = arquivo_prog + ".lock"
    if os.path.exists(lock_path):
        try:
            age = time.time() - os.path.getmtime(lock_path)
            if age < 1800:
                log(f"! Lockfile existe há {age:.0f}s. Outra execução em andamento? Abortando.", "WARN")
                return None
            log(f"  Lockfile antigo ({age:.0f}s) — removendo.", "WARN")
            os.remove(lock_path)
        except OSError:
            pass
    try:
        with open(lock_path, "w", encoding="utf-8") as f:
            f.write(f"{os.getpid()} {datetime.now().isoformat()}\n")
        return lock_path
    except OSError as e:
        log(f"  ! Não conseguiu criar lockfile ({e}). Prosseguindo sem lock.", "WARN")
        return None


def _release_lock(lock_path):
    if lock_path and os.path.exists(lock_path):
        try:
            os.remove(lock_path)
        except OSError:
            pass


_CRIT_MAP = {"VERY_HIGH":"Muito alto","HIGH":"Alto","MEDIUM":"Médio","LOW":"Baixo","VERY_LOW":"Muito baixo"}


def _carregar_solic_via_api():
    try:
        from gerar_bd_via_api import FracttalClient, CLIENT_ID, CLIENT_SECRET, BASE_URL
    except Exception as e:
        log(f"  ! Import falhou: {e}")
        return None
    try:
        client = FracttalClient(CLIENT_ID, CLIENT_SECRET, BASE_URL)
        client.autenticar()
    except Exception as e:
        log(f"  ! Auth falhou: {e}")
        return None
    rows = []
    try:
        for item in client.paginar("requests", page_size=100):
            if item.get("id_status") not in (1, 2, 7, 8):
                continue
            wo = item.get("wo_folio")
            if wo and str(wo).strip() not in ("", "0"):
                continue
            rows.append({
                "Número de Solicitação de Serviço": item.get("id_code"),
                "Status": "Aberto",
                "Ordem de Serviço": "",
                "Descrição": item.get("description") or "",
                "Ativo": item.get("items_description") or "",
                "Código do ativo": item.get("code_item") or "",
                "Localização ou parte de": item.get("parent_description") or "",
                "Criticidade": _CRIT_MAP.get(item.get("priorities_description"), item.get("priorities_description") or ""),
                "É urgente?": "Sim" if item.get("is_urgent") else "Não",
                "Data": item.get("date") or "",
                "Solicitado por": item.get("requested_by") or "",
            })
    except Exception as e:
        log(f"  ! Paginação falhou: {e}")
        return None
    if not rows:
        return None
    log(f"  → {len(rows)} solicitações abertas sem OS (via API)")
    return pd.DataFrame(rows)


def carregar_solicitacoes():
    log("Carregando solicitações (API REST do Fracttal)...")
    df = _carregar_solic_via_api()
    if df is not None:
        return df
    log(f"  Fallback: lendo {os.path.basename(INPUT_SOLICIT)}...")
    df = _read_excel_robusto(INPUT_SOLICIT, sheet_name="Dados")
    if df is None:
        return None
    df = df[df["Status"] == "Aberto"].copy()
    os_col = df["Ordem de Serviço"]
    mask_sem_os = os_col.isna() | (os_col.astype(str).str.strip() == "") | (os_col == 0)
    df = df[mask_sem_os].copy()
    log(f"  → {len(df)} solicitações abertas sem OS (via Excel)")
    return df


def carregar_bd_semanal(week_start, week_end):
    log("Lendo OS/tarefas via API Fracttal...")
    df = fonte_bd_api.df_semanal()   # era _read_excel_robusto(INPUT_BD, sheet_name="Semanal")
    if df is None:
        return None, None
    df["OSs ID"] = pd.to_numeric(df["OSs ID"], errors="coerce")
    df = df.dropna(subset=["OSs ID"]).copy()
    df["OSs ID"] = df["OSs ID"].astype(int)
    df["Data Programada"] = pd.to_datetime(df["Data Programada"], errors="coerce")
    log(f"  → {len(df)} registros no BD")
    df_semana = df[(df["Data Programada"] >= pd.Timestamp(week_start)) &
                   (df["Data Programada"] < pd.Timestamp(week_end))].copy()
    log(f"  → {len(df_semana)} OS com Data Programada na semana ({week_start} a {week_end})")
    return df, df_semana


def carregar_map_responsavel(df_aux):
    """Map UFV-normalizada → RESPONSAVEL O&M. Recebe df ja lido pra evitar releitura."""
    mp = {}
    if df_aux is None:
        return mp
    col_ufv = col_resp = None
    for c in df_aux.columns:
        cu = str(c).strip().upper()
        if cu == 'UFV':
            col_ufv = c
        if ('RESPONS' in cu and ('O&M' in cu or 'OM' in cu)
                and 'URL' not in cu and 'IMAGEM' not in cu):
            col_resp = c
    if col_ufv and col_resp:
        for _, r in df_aux.iterrows():
            u = r.get(col_ufv)
            v = r.get(col_resp)
            if isinstance(u, str) and isinstance(v, str) and v.strip():
                mp[_norm(u)] = v.strip()
    log(f"  → Map Responsavel: {len(mp)} UFVs")
    return mp


def resolver_responsavel(ativo, mapa_resp):
    if not isinstance(ativo, str) or not mapa_resp:
        return ''
    nrm = _norm(ativo)
    if nrm in mapa_resp:
        return mapa_resp[nrm]
    for k, v in mapa_resp.items():
        if nrm.startswith(k) or k in nrm:
            return v
    return ''


def carregar_auxiliar():
    log("Lendo cadastro AUXILIAR (AUXILIAR - FABRICIO.xlsx) p/ mapeamento UFV → Equipe...")
    df = fonte_bd_api.df_auxiliar()   # era _read_excel_robusto(INPUT_BD, sheet_name="AUXILIAR")
    if df is None:
        return {}
    mp = {}
    for _, row in df.iterrows():
        ufv = row.get("UFV")
        eq = row.get("Equipe Cluster")
        if isinstance(ufv, str) and isinstance(eq, str) and eq.strip():
            mp[_norm(ufv)] = eq.strip()
    log(f"  → {len(mp)} UFVs mapeadas")
    return mp


def carregar_lista_prioridades():
    log(f"Lendo {os.path.basename(INPUT_PRIOR)}...")
    df = _read_excel_robusto(INPUT_PRIOR, sheet_name="📋 Lista de Prioridades", header=7)
    if df is None:
        return []
    df = df.rename(columns=lambda c: str(c).replace("\n", " ").strip())
    items = []
    for _, row in df.iterrows():
        try:
            rpn = int(row["# Prioridade"])
        except (ValueError, TypeError, KeyError):
            continue
        atividade = row.get("Atividade (Agrupada)") or ""
        if not isinstance(atividade, str):
            continue
        items.append({
            "rpn": rpn, "categoria": atividade.strip(),
            "norm": _norm(atividade),
        })
    log(f"  → {len(items)} categorias de prioridade carregadas")
    return items


def carregar_confiabilidade():
    log(f"Lendo {os.path.basename(INPUT_CONFIAB)}...")
    df = _read_excel_robusto(INPUT_CONFIAB, sheet_name="Resumo por Categoria", header=3)
    if df is None:
        return []
    df = df.rename(columns=lambda c: str(c).strip())
    items = []
    for _, row in df.iterrows():
        cat = row.get("Categoria")
        med = row.get("Média (h)")
        if not isinstance(cat, str) or cat.strip().upper() == "TOTAL":
            continue
        try:
            med_f = float(med)
        except (ValueError, TypeError):
            continue
        items.append({"categoria": cat.strip(), "norm": _norm(cat), "media_h": med_f})
    log(f"  → {len(items)} categorias de confiabilidade")
    return items




# ====================== Feriados ======================

INPUT_FERIADOS = os.path.join(BASE_DIR, "Feriados",
                              "FERIADOS ESTADUAIS, MUNICIPAIS E NACIONAIS 2026.xlsx")


def carregar_feriados():
    """Carrega feriados da pasta Feriados.
    Retorna dict {data: set(equipes_bloqueadas_ou_'TODOS')}.
    NACIONAL bloqueia TODOS; ESTADUAL bloqueia equipes da UF.
    """
    log(f"Lendo {os.path.basename(INPUT_FERIADOS)}...")
    if not os.path.exists(INPUT_FERIADOS):
        log(f"  ! Arquivo de feriados não encontrado: {INPUT_FERIADOS}", "WARN")
        log(f"  ! Prosseguindo sem bloqueio de feriados.", "WARN")
        return {}
    df = _read_excel_robusto(INPUT_FERIADOS, sheet_name="Feriados 2026", header=1)
    if df is None:
        return {}
    feriados = {}
    for _, row in df.iterrows():
        tipo = str(row.get("Tipo") or "").strip().upper()
        uf = str(row.get("Estado") or "").strip().upper()
        data = row.get("Data")
        if data is None or (hasattr(data, "year") is False):
            continue
        try:
            d = data.date() if hasattr(data, "date") else data
        except Exception:
            continue
        if tipo == "NACIONAL":
            feriados.setdefault(d, set()).add("TODOS")
        elif tipo == "ESTADUAL" and uf and uf != "TODOS":
            feriados.setdefault(d, set()).add(uf)
        # MUNICIPAL ignorado (precisaria mapear cidade->UFV)
    log(f"  → {len(feriados)} datas de feriado carregadas")
    return feriados


def _equipe_bloqueada_em(equipe, data, feriados):
    """Retorna True se a equipe não pode trabalhar nessa data por feriado."""
    if not feriados or data not in feriados:
        return False
    bloqueios = feriados[data]
    if "TODOS" in bloqueios:
        return True
    if not isinstance(equipe, str):
        return False
    # Nome de equipe começa com UF (ex.: "ES Norte 01", "BA Sul 01")
    partes = equipe.strip().split()
    if partes and partes[0].upper() in bloqueios:
        return True
    return False

# ====================== Match UFV → Equipe ======================

UFV_LOCALIZACAO_RE = re.compile(r"//\s*[^/]+/\s*([^/]+?)(?:/|$)")


def extrair_ufv_da_localizacao(loc):
    if not isinstance(loc, str):
        return None
    m = UFV_LOCALIZACAO_RE.search(loc)
    return m.group(1).strip() if m else None


def _resolver_nome_aba(equipe, sheetnames):
    if not equipe:
        return None
    eq_norm = _norm(equipe)
    for sn in sheetnames:
        if _norm(sn) == eq_norm:
            return sn
    return None


def mapear_equipe(ufv_raw, mapa_aux, df_bd):
    if not ufv_raw:
        return None
    nrm = _norm(ufv_raw)
    if nrm in mapa_aux:
        return mapa_aux[nrm]
    if df_bd is not None and "Ativo Classificação 1" in df_bd.columns:
        ac1_norm = df_bd["Ativo Classificação 1"].fillna("").astype(str).apply(_norm)
        idx = df_bd.index[ac1_norm == nrm]
        if len(idx):
            eq = df_bd.loc[idx[0], "Ativo Classificação 2"]
            if isinstance(eq, str) and eq.strip():
                return eq.strip()
    # Fallback por prefixo
    for k_norm, eq in mapa_aux.items():
        if len(k_norm) >= 5 and nrm.startswith(k_norm + " "):
            return eq
    return None


# ====================== Priorização ======================

def calcular_rpn(descricao, lista_prior):
    if not isinstance(descricao, str) or not descricao.strip():
        return (None, None)
    desc_norm = _norm(descricao)
    for item in lista_prior:
        cat_tokens = [t for t in re.split(r"[^a-z0-9]+", item["norm"]) if len(t) > 3]
        if not cat_tokens:
            continue
        hits = sum(1 for t in cat_tokens if t in desc_norm)
        n = len(cat_tokens)
        threshold = 1.0 if n <= 2 else (0.66 if n == 3 else 0.6)
        if hits / n >= threshold and hits >= 1:
            return (item["rpn"], item["categoria"])
    p_emerg = ["religamento", "trip", "desligamento", "parada total"]
    p_corr = ["inversor", "relé", "rele", "cabine", "tracker", "string", "módulo", "modulo"]
    p_prev = ["preventiva", "inspeção", "inspecao", "limpeza", "mpm", "mps", "mpa", "mpt", "mpq"]
    if any(p in desc_norm for p in p_emerg):
        return (5, "Emergencial (fallback keyword)")
    if any(p in desc_norm for p in p_corr):
        return (15, "Corretiva (fallback keyword)")
    if any(p in desc_norm for p in p_prev):
        return (25, "Preventiva (fallback keyword)")
    return (None, None)


def calcular_score(sol_row, rpn):
    s_crit = SCORE_CRITICIDADE.get(sol_row.get("Criticidade"), 10)
    s_urg = SCORE_URGENCIA.get(sol_row.get("É urgente?"), 0)
    s_rpn = (40 - min(int(rpn), 40)) if rpn is not None else 5
    return s_crit + s_urg + s_rpn


def estimar_duracao(descricao, conf_items):
    if not isinstance(descricao, str) or not conf_items:
        return DEFAULT_DURACAO_H
    desc_norm = _norm(descricao)
    melhor, melhor_score = None, 0
    for item in conf_items:
        tokens = [t for t in re.split(r"[^a-z0-9]+", item["norm"]) if len(t) > 3]
        if not tokens:
            continue
        hits = sum(1 for t in tokens if t in desc_norm)
        if hits > melhor_score and hits / len(tokens) >= 0.5:
            melhor = item
            melhor_score = hits
    return round(float(melhor["media_h"]), 2) if melhor else DEFAULT_DURACAO_H


# ====================== Alocação de slots ======================

def carregar_ocupacao_atual(wb_prog):
    ocup = {}
    for sheet_name in wb_prog.sheetnames:
        if sheet_name.startswith("_"):
            continue
        ws = wb_prog[sheet_name]
        if ws.max_row < 2:
            ocup[sheet_name] = {}
            continue
        header = [c.value for c in ws[1]]
        try:
            idx_dia = header.index("Dia")
            idx_ini = header.index("Hora Início")
            idx_fim = header.index("Hora Fim")
        except ValueError:
            continue
        slots = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            dia = row[idx_dia]
            hi = _hora_str_to_min(row[idx_ini])
            hf = _hora_str_to_min(row[idx_fim])
            if dia and hi is not None and hf is not None:
                slots.setdefault(dia, []).append((hi, hf))
        for d in slots:
            slots[d].sort()
        ocup[sheet_name] = slots
    return ocup


def _encaixar_evitando_almoco(inicio, dur_min):
    fim = inicio + dur_min
    if inicio < HORA_ALMOCO_INI_MIN < fim:
        inicio = HORA_ALMOCO_FIM_MIN
        fim = inicio + dur_min
    if fim > HORA_FIM_TURNO_MIN:
        return None
    return (inicio, fim)


def proximo_slot_livre(slots_dia, dur_min, start_min_override=None):
    cursor = start_min_override if start_min_override is not None else (HORA_INICIO_TURNO_MIN + 30)
    # Se cursor está no almoço, empurra
    if HORA_ALMOCO_INI_MIN <= cursor < HORA_ALMOCO_FIM_MIN:
        cursor = HORA_ALMOCO_FIM_MIN
    for ini, fim in sorted(slots_dia):
        if fim <= cursor:
            continue
        if ini - cursor >= dur_min + GAP_ENTRE_OS_MIN:
            slot = _encaixar_evitando_almoco(cursor, dur_min)
            if slot:
                return slot
        cursor = max(cursor, fim + GAP_ENTRE_OS_MIN)
        if HORA_ALMOCO_INI_MIN <= cursor < HORA_ALMOCO_FIM_MIN:
            cursor = HORA_ALMOCO_FIM_MIN
    if cursor + dur_min <= HORA_FIM_TURNO_MIN:
        return _encaixar_evitando_almoco(cursor, dur_min)
    return None


def encontrar_slot(equipe, ocupacao, dias_semana, dur_h, feriados=None,
                   min_ini_por_dia=None):
    """Procura próximo slot livre. min_ini_por_dia: dict {date: minutos_min}
    para forçar não sugerir no passado quando a data é hoje."""
    dur_min = int(round(dur_h * 60))
    slots_equipe = ocupacao.get(equipe, {})
    for dt, dia_str in dias_semana:
        if feriados and _equipe_bloqueada_em(equipe, dt, feriados):
            continue
        ini_minimo = None
        if min_ini_por_dia and dt in min_ini_por_dia:
            ini_minimo = min_ini_por_dia[dt]
            if ini_minimo >= HORA_FIM_TURNO_MIN:
                continue
        slot = proximo_slot_livre(slots_equipe.get(dia_str, []), dur_min,
                                   start_min_override=ini_minimo)
        if slot:
            ini, fim = slot
            slots_equipe.setdefault(dia_str, []).append((ini, fim))
            slots_equipe[dia_str].sort()
            ocupacao[equipe] = slots_equipe
            return (dia_str, ini, dur_h)
    return None


# ====================== Sugestões PCM (arquivo separado) ======================

def carregar_sugestoes_pcm_anteriores():
    """Lê Sugestoes_PCM.xlsx (se existir). Preserva histórico e status."""
    if not os.path.exists(OUTPUT_SUGEST):
        return {}
    try:
        wb = load_workbook(OUTPUT_SUGEST)
    except Exception as e:
        log(f"  ! Não foi possível ler {OUTPUT_SUGEST}: {e}", "WARN")
        return {}
    if "Sugestões" not in wb.sheetnames:
        return {}
    ws = wb["Sugestões"]
    if ws.max_row < 2:
        return {}
    header = [c.value for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(header, row))
        num = d.get("Nº Solicitação")
        if num is not None:
            out[num] = d
    return out


def gerar_sugestoes_pcm(df_solic, mapa_aux, df_bd, lista_prior, conf_items,
                        wb_prog, dias_semana, feriados=None):
    """Calcula sugestões para cada solicitação aberta."""
    log("STEP B: Gerando sugestões PCM (para o programador criar as OS)...")

    # Não sugerir slots no passado: hoje, mínimo = agora + 1h
    agora = datetime.now()
    min_ini_por_dia = {}
    for dt, dia_str in dias_semana:
        if dt == agora.date():
            min_minutos = (agora.hour * 60) + agora.minute + 60  # now + 1h
            # Arredonda pra próximo múltiplo de 15min
            min_minutos = ((min_minutos + 14) // 15) * 15
            min_ini_por_dia[dt] = min_minutos
            log(f"  → Hoje ({dt}): só sugerir a partir de {_hora_min_to_str(min_minutos)}")

    if feriados:
        n_blq = sum(1 for dt, _ in dias_semana if dt in feriados)
        if n_blq:
            log(f"  → {n_blq} feriado(s) na semana — bloqueando alocação")

    ocupacao = carregar_ocupacao_atual(wb_prog)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    anteriores = carregar_sugestoes_pcm_anteriores()
    sugestoes = {}

    enriquecidas = []
    for _, row in df_solic.iterrows():
        ufv = extrair_ufv_da_localizacao(row.get("Localização ou parte de"))
        equipe = mapear_equipe(ufv, mapa_aux, df_bd)
        desc = row.get("Descrição") or ""
        rpn, categoria = calcular_rpn(desc, lista_prior)
        score = calcular_score(row, rpn)
        dur_h = estimar_duracao(desc, conf_items)
        enriquecidas.append({"row": row, "ufv": ufv, "equipe": equipe,
                             "rpn": rpn, "categoria": categoria,
                             "score": score, "dur_h": dur_h})
    enriquecidas.sort(key=lambda x: -x["score"])

    for e in enriquecidas:
        row = e["row"]
        equipe = e["equipe"]
        num = row.get("Número de Solicitação de Serviço")
        status = "Aguardando criação no sistema"
        observacao = ""

        equipe_aba = _resolver_nome_aba(equipe, wb_prog.sheetnames) if equipe else None
        if not equipe:
            dia_sug, hora_str, dur_used = ("", "", e["dur_h"])
            status = "Sem equipe mapeada"
            observacao = f"UFV \"{e['ufv']}\" não casou com nenhuma equipe"
        elif not equipe_aba:
            dia_sug, hora_str, dur_used = ("", "", e["dur_h"])
            status = "Equipe sem aba na Programação"
            observacao = f"Equipe \"{equipe}\" não tem aba"
        else:
            equipe = equipe_aba
            slot = encontrar_slot(equipe, ocupacao, dias_semana, e["dur_h"],
                                   feriados=feriados, min_ini_por_dia=min_ini_por_dia)
            if slot:
                dia_sug, ini_min, dur_used = slot
                hora_str = _hora_min_to_str(ini_min)
            else:
                dia_sug, hora_str, dur_used = ("", "", e["dur_h"])
                status = "Sem slot na semana - avaliar próxima"
                observacao = "Semana sem capacidade disponível"

        ant = anteriores.get(num, {})
        criado_em = ant.get("Criado em") or timestamp

        sugestoes[num] = {
            "Nº Solicitação": num, "Equipe": equipe or "",
            "Ativo": row.get("Ativo"),
            "Código Equipamento": row.get("Código do ativo"),
            "Descrição": (row.get("Descrição") or "")[:500],
            "Criticidade": row.get("Criticidade"),
            "Urgente?": row.get("É urgente?"),
            "RPN": e["rpn"], "Categoria RPN": e["categoria"],
            "Score": e["score"],
            "Dia Sugerido": dia_sug, "Hora Início Sugerida": hora_str,
            "Duração (h)": dur_used,
            "Status no PCM": status, "Observação": observacao,
            "Criado em": criado_em, "Última Avaliação": timestamp,
            "OS Criada no Sistema": ant.get("OS Criada no Sistema") or "",
        }

    # Mantém sugestões antigas cuja solicitação saiu de Aberto (OS criada),
    # com o status correspondente
    for num, ant in anteriores.items():
        if num not in sugestoes:
            ant = dict(ant)
            ant["Última Avaliação"] = timestamp
            if not ant.get("OS Criada no Sistema"):
                ant["Status no PCM"] = "Solicitação não mais Aberta (verificar)"
            sugestoes[num] = ant

    log(f"  → {len(sugestoes)} sugestões geradas")
    return sugestoes


def escrever_arquivo_sugestoes_pcm(sugestoes):
    """Grava Sugestoes_PCM.xlsx (sister-file da planilha de solicitações).
    Ordenado por Score desc. PCM olha esse arquivo e cria as OS no sistema."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sugestões"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for j, col in enumerate(COLS_SUGESTOES, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    fill_alta = PatternFill("solid", fgColor="FCE4D6")
    fill_criada = PatternFill("solid", fgColor="C6EFCE")
    fill_sem_slot = PatternFill("solid", fgColor="FFEB9C")
    lista = sorted(sugestoes.values(), key=lambda r: -(r.get("Score") or 0))
    for i, sug in enumerate(lista, start=2):
        for j, col in enumerate(COLS_SUGESTOES, start=1):
            ws.cell(row=i, column=j, value=sug.get(col))
        score = sug.get("Score") or 0
        status = sug.get("Status no PCM") or ""
        if sug.get("OS Criada no Sistema"):
            fill = fill_criada
        elif "Sem slot" in status or "Sem equipe" in status or "sem aba" in status:
            fill = fill_sem_slot
        elif score >= 150:
            fill = fill_alta
        else:
            fill = None
        if fill:
            for j in range(1, len(COLS_SUGESTOES) + 1):
                ws.cell(row=i, column=j).fill = fill

    widths = {
        "Nº Solicitação": 13, "Equipe": 13, "Ativo": 30,
        "Código Equipamento": 22, "Descrição": 50, "Criticidade": 11,
        "Urgente?": 9, "RPN": 6, "Categoria RPN": 30, "Score": 8,
        "Dia Sugerido": 24, "Hora Início Sugerida": 12, "Duração (h)": 11,
        "Status no PCM": 30, "Observação": 35, "Criado em": 18,
        "Última Avaliação": 18, "OS Criada no Sistema": 22,
    }
    for j, col in enumerate(COLS_SUGESTOES, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(col, 14)

    # Aba "Como Usar"
    ws2 = wb.create_sheet("Como Usar")
    instrucoes = [
        ("COMO USAR ESTE ARQUIVO", True),
        ("", False),
        ("Este arquivo é gerado automaticamente pelo atualizacao_semanal.py.", False),
        ("Ele lista as Solicitações de Serviço em status 'Aberto' que ainda não têm OS,", False),
        ("e SUGERE a equipe, dia e horário para o programador do PCM criar a OS no sistema.", False),
        ("", False),
        ("FLUXO:", True),
        ("  1. PCM abre este arquivo e olha a coluna 'Status no PCM'.", False),
        ("  2. PCM analisa as sugestões (Score = prioridade; maior = mais urgente).", False),
        ("  3. PCM cria a OS no sistema com a data/hora desejada (pode ser diferente da sugerida).", False),
        ("  4. O script roda novamente, detecta a OS no BD, e a insere automaticamente no", False),
        ("     arquivo Programação Semana XX.xlsx com a data/hora QUE O PROGRAMADOR DEFINIU.", False),
        ("", False),
        ("CORES:", True),
        ("  Verde   = OS já criada no sistema (status atualizado).", False),
        ("  Laranja = Alta prioridade (Score >= 150).", False),
        ("  Amarelo = Sem slot disponível na semana / sem equipe mapeada.", False),
        ("  Branco  = Sugestão normal aguardando criação.", False),
        ("", False),
        ("NÃO EDITE ESTE ARQUIVO MANUALMENTE - ele é sobrescrito a cada execução.", True),
    ]
    bold = Font(bold=True, size=12)
    for i, (text, is_bold) in enumerate(instrucoes, start=1):
        c = ws2.cell(row=i, column=1, value=text)
        if is_bold:
            c.font = bold
    ws2.column_dimensions["A"].width = 100

    # Salva com retry
    if _save_workbook_robusto(wb, OUTPUT_SUGEST, etapa="Sugestões PCM"):
        log(f"  → Arquivo gravado: {os.path.basename(OUTPUT_SUGEST)}")


# ====================== STEP A: Sync de status + Incorporar novas OS ======================

def garantir_colunas_extras(ws):
    header_row = list(ws[1])
    header_map = {}
    for j, c in enumerate(header_row, start=1):
        if c.value:
            header_map[c.value] = j
    next_col = len(header_row) + 1
    header_fill = PatternFill("solid", fgColor="70AD47")
    header_font = Font(bold=True, color="FFFFFF")
    for col in COLS_EXTRA_EQUIPE:
        if col not in header_map:
            c = ws.cell(row=1, column=next_col, value=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            header_map[col] = next_col
            ws.column_dimensions[get_column_letter(next_col)].width = 18
            next_col += 1
    return header_map


def sync_status_bd(wb_prog, df_bd):
    log("STEP A1: Sincronizando status com BD Relatório Semanal...")
    bd_map = {int(r["OSs ID"]): r for _, r in df_bd.iterrows() if pd.notna(r["OSs ID"])}
    total_atu, total_inalt, total_removidas = 0, 0, 0
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    for sheet_name in wb_prog.sheetnames:
        if sheet_name.startswith("_"):
            continue
        ws = wb_prog[sheet_name]
        if ws.max_row < 2:
            continue
        cols = garantir_colunas_extras(ws)
        idx_osid = cols.get("OSs ID")
        idx_status = cols.get("Status Atual (BD)")
        idx_update = cols.get("Última Atualização")
        if not (idx_osid and idx_status):
            continue
        linhas_remover = []
        for r in range(2, ws.max_row + 1):
            os_id_val = ws.cell(row=r, column=idx_osid).value
            try:
                os_id = int(os_id_val)
            except (ValueError, TypeError):
                continue
            bd_row = bd_map.get(os_id)
            if bd_row is None:
                continue
            estado_bd = bd_row.get("Estado da Tarefa")
            # Remove linhas que viraram Cancelada/Finalizada
            if isinstance(estado_bd, str):
                est_low = estado_bd.lower()
                if "cancelad" in est_low:
                    linhas_remover.append(r)
                    continue
            atual = ws.cell(row=r, column=idx_status).value
            if estado_bd != atual:
                ws.cell(row=r, column=idx_status).value = estado_bd
                if idx_update:
                    ws.cell(row=r, column=idx_update).value = timestamp
                if isinstance(estado_bd, str) and "progresso" in estado_bd.lower():
                    ws.cell(row=r, column=idx_status).fill = PatternFill("solid", fgColor="FFEB9C")
                total_atu += 1
            else:
                total_inalt += 1
        # Remove de baixo pra cima pra não invalidar índices
        for r in sorted(linhas_remover, reverse=True):
            ws.delete_rows(r, 1)
            total_removidas += 1
    log(f"  → {total_atu} OS com status atualizado; {total_inalt} inalteradas; {total_removidas} canceladas/finalizadas removidas")


def _formatar_dia(dt):
    if pd.isna(dt) or dt is None:
        return ""
    try:
        d = dt.date() if hasattr(dt, "date") else dt
        return f"{DAY_NAMES_FULL[d.weekday()]} ({d.strftime('%d/%m')})"
    except Exception:
        return ""


def _formatar_hora_bd(dt):
    if pd.isna(dt) or dt is None:
        return ""
    try:
        return dt.strftime("%H:%M")
    except Exception:
        return ""


def _converter_duracao(td):
    if td is None or pd.isna(td):
        return DEFAULT_DURACAO_H
    if hasattr(td, "total_seconds"):
        return round(td.total_seconds() / 3600, 2)
    try:
        return float(td)
    except (ValueError, TypeError):
        return DEFAULT_DURACAO_H


def incorporar_os_novas_do_bd(wb_prog, df_bd_semana, sugestoes_pcm, dias_semana=None, hoje=None):
    """OS na semana corrente do BD que NÃO estão no arquivo de programação
    são incorporadas USANDO A DATA/HORA QUE O PROGRAMADOR DEFINIU NO BD.
    Se a OS veio de uma Solicitação, atualiza Sugestoes_PCM.xlsx."""
    log("STEP A2: Buscando OS novas no BD (incorporar com data/hora do programador)...")
    presentes = set()
    for sheet_name in wb_prog.sheetnames:
        if sheet_name.startswith("_"):
            continue
        ws = wb_prog[sheet_name]
        if ws.max_row < 2:
            continue
        header = [c.value for c in ws[1]]
        try:
            idx = header.index("OSs ID")
        except ValueError:
            continue
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=idx + 1).value
            try:
                presentes.add(int(v))
            except (ValueError, TypeError):
                pass

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    incorporadas = 0
    via_solicitacao = 0
    for _, row in df_bd_semana.iterrows():
        os_id = row.get("OSs ID")
        if pd.isna(os_id):
            continue
        os_id = int(os_id)
        if os_id in presentes:
            continue
        # FIX dedupe in-loop: marca como presente AGORA, antes de incorporar.
        # Sem isso, próxima linha do df_bd_semana com mesma OSs ID (BD tem
        # múltiplas tarefas/subtasks por OS) reincorporaria, gerando duplicação
        # massiva (ex: 180 cópias da mesma OS em uma única rodada).
        presentes.add(os_id)
        # Ignora OSs Cancelada/Finalizada/Concluída — não fazem sentido na programação.
        # Crítico: sem "finaliz"/"conclu" aqui, o sync_status_bd remove a linha
        # (porque virou Finalizada) e o loop seguinte re-adiciona, gerando duplicação
        # massiva (ex: 180 cópias da mesma OS Finalizada acumuladas ao longo do dia).
        estado_bd_inc = row.get("Estado da Tarefa")
        if isinstance(estado_bd_inc, str):
            est_l = estado_bd_inc.lower()
            if "cancelad" in est_l or "finaliz" in est_l or "conclu" in est_l:
                continue
        equipe = row.get("Ativo Classificação 2")
        equipe_aba = _resolver_nome_aba(equipe, wb_prog.sheetnames) if isinstance(equipe, str) else None
        if not equipe_aba:
            continue
        ws = wb_prog[equipe_aba]
        cols = garantir_colunas_extras(ws)

        num_solic_bd = row.get("Número de Solicitação")
        try:
            num_solic_bd = int(num_solic_bd) if pd.notna(num_solic_bd) else None
        except (ValueError, TypeError):
            num_solic_bd = None

        # Determina se é preventiva (MPM/MPS/MPA/MPM-Inversor/MPM-Mod-Tracker etc)
        tipo_tarefa = str(row.get("Tipo de tarefa") or "").upper()
        # Também tenta inferir pelo tipo de OS via Tarefa text
        tarefa_str = str(row.get("Tarefa") or "").upper()
        eh_preventiva = (
            tipo_tarefa.startswith("PREVENT") or
            any(p in tarefa_str for p in ["MPM", "MPS", "MPA", "MPT", "MPQ"]) or
            any(p in tipo_tarefa for p in ["MPM", "MPS", "MPA", "MPT", "MPQ"])
        )
        if eh_preventiva:
            # Preventiva nunca é "Nova OS" — sempre Programada/Reprogramada
            label_origem = ""
            label_reprog = "Sim"
        else:
            label_origem = "Sim (via Solic #" + str(num_solic_bd) + ")" if num_solic_bd else "Sim (BD)"
            label_reprog = "Não"

        # Data/hora EXATAS do BD (definidas pelo programador no sistema)
        data_prog = row.get("Data Programada")
        data_ini = row.get("Data inicial")
        data_fim = row.get("Data final")
        hora_ini = _formatar_hora_bd(data_ini) or _formatar_hora_bd(data_prog)
        hora_fim = _formatar_hora_bd(data_fim)
        if not hora_fim and hora_ini:
            # Estima fim com base na duração
            dur = _converter_duracao(row.get("Tarefa -> Duração estimada"))
            ini_min = _hora_str_to_min(hora_ini)
            if ini_min is not None:
                hora_fim = _hora_min_to_str(ini_min + int(round(dur * 60)))

        nova = {
            "Equipe": equipe_aba,
            "Dia": _formatar_dia(data_prog),
            "OSs ID": os_id,
            "Ativo (Usina)": row.get("Ativo Classificação 1") or row.get("Ativo"),
            "Cidade": row.get("Ativo Classificação 1"),
            "Código Equipamento": row.get("Código"),
            "Tipo": row.get("Tipo de tarefa") or "",
            "Tarefa": row.get("Tarefa") or "",
            "Estado Tarefa (antes)": row.get("Estado da Tarefa"),
            "RPN/Prioridade": row.get("Tarefa -> Criticidade") or "",
            "Reprogramada": label_reprog,
            "Nº vezes programada": 1,
            "Termografia": "Não",
            "Hora Início": hora_ini,
            "Hora Fim": hora_fim,
            "Duração (h)": _converter_duracao(row.get("Tarefa -> Duração estimada")),
            "Desloc (h)": 0,
            "Paralelo (terceirizada)": "Não",
            "Status Atual (BD)": row.get("Estado da Tarefa"),
            "Responsável": _MAP_RESPONSAVEL.get(_norm(str(row.get("Ativo Classificação 1") or row.get("Ativo") or "")), "") if _MAP_RESPONSAVEL else "",
            "Nova OS": label_origem,
            "Nº Solicitação Origem": num_solic_bd or "",
            "Última Atualização": timestamp,
        }
        nova_linha = ws.max_row + 1
        for col_nome, idx in cols.items():
            ws.cell(row=nova_linha, column=idx, value=nova.get(col_nome))
        # Cor: verde se via solicitação, azul claro se OS direta no BD
        cor = "C6EFCE" if num_solic_bd else "DDEBF7"
        fill = PatternFill("solid", fgColor=cor)
        for j in range(1, max(cols.values()) + 1):
            ws.cell(row=nova_linha, column=j).fill = fill
        incorporadas += 1

        # DESLOCAR CONFLITANTES: se a nova OS tem hora definida e ocupa um slot,
        # móveis que conflitam são empurradas (cascata pela semana, overflow→pendentes)
        ini_min_nova = _hora_str_to_min(hora_ini)
        fim_min_nova = _hora_str_to_min(hora_fim)
        if (ini_min_nova is not None and fim_min_nova is not None
                and ini_min_nova < fim_min_nova and dias_semana):
            try:
                d, p = deslocar_conflitantes(
                    wb_prog, equipe_aba, nova["Dia"], ini_min_nova, fim_min_nova,
                    dias_semana, hoje or date.today(), os_id, ws_equipe=ws)
                if d or p:
                    log(f"    OS #{os_id}: {d} deslocada(s), {p} pendente(s)")
            except Exception as e:
                log(f"    ! Erro ao deslocar conflitantes de OS #{os_id}: {e}", "WARN")

        # Atualiza Sugestoes_PCM se houver
        if num_solic_bd and num_solic_bd in sugestoes_pcm:
            sug = sugestoes_pcm[num_solic_bd]
            sug["OS Criada no Sistema"] = os_id
            sug["Status no PCM"] = "OS criada e incorporada à Programação"
            sug["Observação"] = (
                f"OS #{os_id} criada em {_formatar_dia(data_prog)} {hora_ini} "
                f"(aba {equipe_aba})"
            )
            sug["Última Avaliação"] = timestamp
            via_solicitacao += 1

    log(f"  → {incorporadas} OS novas incorporadas ({via_solicitacao} vieram de solicitações)")


# ====================== Deslocamento em cascata ======================

DIA_DATE_RE = re.compile(r"\((\d{1,2})/(\d{1,2})\)")


def _parse_dia_str(dia_str, ano):
    """De 'Segunda-feira (18/05)' extrai date(ano, 5, 18)."""
    if not isinstance(dia_str, str):
        return None
    m = DIA_DATE_RE.search(dia_str)
    if not m:
        return None
    try:
        return date(ano, int(m.group(2)), int(m.group(1)))
    except (ValueError, TypeError):
        return None


def _is_os_travada(row_vals, header_map, hoje, data_dia):
    """Critérios escolhidos pelo usuário:
       - Status BD/orig contém 'finaliz'
       - Paralelo (terceirizada) == Sim
       - Dia anterior a hoje
    """
    if data_dia and hoje and data_dia < hoje:
        return True
    st_bd = row_vals[header_map.get("Status Atual (BD)", -1)] if "Status Atual (BD)" in header_map else None
    st_orig = row_vals[header_map.get("Estado Tarefa (antes)", -1)] if "Estado Tarefa (antes)" in header_map else None
    combo = str(st_bd or st_orig or "").lower()
    if "finaliz" in combo:
        return True
    if "Paralelo (terceirizada)" in header_map:
        p = str(row_vals[header_map["Paralelo (terceirizada)"]] or "").lower()
        if p == "sim":
            return True
    return False


def _coletar_slots_dia(ws, dia_str, header_map, hoje):
    """Retorna lista [(row, ini_min, fim_min, dur_min, travada)] do dia."""
    ano = date.today().year
    data_dia = _parse_dia_str(dia_str, ano)
    idx_dia = header_map.get("Dia")
    idx_ini = header_map.get("Hora Início")
    idx_fim = header_map.get("Hora Fim")
    if idx_dia is None or idx_ini is None or idx_fim is None:
        return []
    out = []
    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=j + 1).value for j in range(len(header_map))]
        # header_map maps name->idx (0-based), but cells indexed from len-of-header
        # melhor: pegar até max(idx)+1
        max_idx = max(header_map.values()) if header_map else 0
        row_vals = [ws.cell(row=r, column=j + 1).value for j in range(max_idx + 1)]
        if row_vals[idx_dia] != dia_str:
            continue
        ini_min = _hora_str_to_min(row_vals[idx_ini])
        fim_min = _hora_str_to_min(row_vals[idx_fim])
        if ini_min is None or fim_min is None:
            continue
        travada = _is_os_travada(row_vals, header_map, hoje, data_dia)
        out.append({
            "row": r, "ini": ini_min, "fim": fim_min,
            "dur": fim_min - ini_min, "travada": travada,
            "row_vals": row_vals,
        })
    return out


def _achar_slot_livre(ocupados_sorted, dur_min, start_min):
    """Dado lista de (ini,fim) ordenada e duração desejada, encontra
    o primeiro slot livre a partir de start_min, respeitando GAP e almoço.
    Retorna (ini, fim) ou None."""
    cursor = max(start_min, HORA_INICIO_TURNO_MIN + 30)
    # evita almoço se cursor cai dentro
    if HORA_ALMOCO_INI_MIN <= cursor < HORA_ALMOCO_FIM_MIN:
        cursor = HORA_ALMOCO_FIM_MIN
    for ini, fim in ocupados_sorted:
        if fim <= cursor:
            continue
        # se cabe entre cursor e ini
        if ini - cursor >= dur_min:
            slot = _encaixar_evitando_almoco(cursor, dur_min)
            if slot:
                return slot
        cursor = max(cursor, fim + GAP_ENTRE_OS_MIN)
        if HORA_ALMOCO_INI_MIN <= cursor < HORA_ALMOCO_FIM_MIN:
            cursor = HORA_ALMOCO_FIM_MIN
    # depois do último
    if cursor + dur_min <= HORA_FIM_TURNO_MIN:
        return _encaixar_evitando_almoco(cursor, dur_min)
    return None


def _registrar_pendente(wb_prog, equipe, row_vals, header_map, motivo):
    """Adiciona uma linha à aba '_Pendentes' descrevendo a OS que não coube."""
    sheet = wb_prog["_Pendentes"] if "_Pendentes" in wb_prog.sheetnames else None
    if sheet is None:
        sheet = wb_prog.create_sheet("_Pendentes", 1)
        for j, h in enumerate(["Equipe", "OSs ID", "Ativo", "Tarefa",
                                "Código", "Tipo", "Duração (h)", "Motivo"], start=1):
            sheet.cell(row=1, column=j, value=h)
    nova_linha = sheet.max_row + 1
    sheet.cell(row=nova_linha, column=1, value=equipe)
    sheet.cell(row=nova_linha, column=2, value=row_vals[header_map.get("OSs ID", -1)] if "OSs ID" in header_map else None)
    sheet.cell(row=nova_linha, column=3, value=row_vals[header_map.get("Ativo (Usina)", -1)] if "Ativo (Usina)" in header_map else None)
    sheet.cell(row=nova_linha, column=4, value=row_vals[header_map.get("Tarefa", -1)] if "Tarefa" in header_map else None)
    sheet.cell(row=nova_linha, column=5, value=row_vals[header_map.get("Código Equipamento", -1)] if "Código Equipamento" in header_map else None)
    sheet.cell(row=nova_linha, column=6, value=row_vals[header_map.get("Tipo", -1)] if "Tipo" in header_map else None)
    sheet.cell(row=nova_linha, column=7, value=row_vals[header_map.get("Duração (h)", -1)] if "Duração (h)" in header_map else None)
    sheet.cell(row=nova_linha, column=8, value=motivo)
    fill = PatternFill("solid", fgColor="F4B084")  # laranja
    for j in range(1, 9):
        sheet.cell(row=nova_linha, column=j).fill = fill


def deslocar_conflitantes(wb_prog, equipe_aba, dia_str_nova, ini_nova_min,
                          fim_nova_min, dias_semana, hoje, os_id_nova,
                          ws_equipe=None):
    """Após inserir uma nova OS no slot fixo (dia_str_nova,
    [ini_nova_min, fim_nova_min)), desloca as OS móveis (não travadas) que
    conflitam para próximos slots livres. Cascata: se não couber no dia,
    push pro próximo dia da semana. Overflow → _Pendentes.

    Retorna: nº de OS deslocadas (incluindo cascata) e nº de pendentes.
    """
    if ws_equipe is None:
        ws_equipe = wb_prog[equipe_aba]
    header = [c.value for c in ws_equipe[1]]
    header_map = {h: i for i, h in enumerate(header) if h is not None}

    # Garante que tem coluna "Deslocada (Original)"
    cols_dict = garantir_colunas_extras(ws_equipe)
    # cols_dict é 1-based; rebuild header_map after extras
    header = [c.value for c in ws_equipe[1]]
    header_map = {h: i for i, h in enumerate(header) if h is not None}

    idx_dia_1b = cols_dict.get("Dia")
    idx_ini_1b = cols_dict.get("Hora Início")
    idx_fim_1b = cols_dict.get("Hora Fim")
    idx_desloc_1b = cols_dict.get("Deslocada (Original)")
    idx_update_1b = cols_dict.get("Última Atualização")

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Ordena dias da semana a partir do dia_str_nova
    dias_a_partir = []
    inicio_encontrado = False
    for d_date, d_str in dias_semana:
        if d_str == dia_str_nova:
            inicio_encontrado = True
        if inicio_encontrado:
            if hoje and d_date < hoje:
                continue
            dias_a_partir.append((d_date, d_str))
    if not dias_a_partir:
        return 0, 0

    # Função: lista os slots ocupados de um dia (incluindo cascata pendente)
    def _obter_ocupados(d_str):
        slots = _coletar_slots_dia(ws_equipe, d_str, header_map, hoje)
        return sorted([(s["ini"], s["fim"]) for s in slots])

    # Acha conflitantes NO DIA DA NOVA OS
    slots_dia = _coletar_slots_dia(ws_equipe, dia_str_nova, header_map, hoje)
    conflitantes = []
    for s in slots_dia:
        if s["travada"]:
            continue
        # ignora a própria nova OS (slot recém inserido)
        if s["ini"] == ini_nova_min and s["fim"] == fim_nova_min:
            continue
        # overlap?
        if s["ini"] < fim_nova_min and s["fim"] > ini_nova_min:
            conflitantes.append(s)

    deslocadas = 0
    pendentes = 0

    for conf in conflitantes:
        # Salva original
        dia_orig = ws_equipe.cell(row=conf["row"], column=idx_dia_1b).value
        ini_orig = ws_equipe.cell(row=conf["row"], column=idx_ini_1b).value
        fim_orig = ws_equipe.cell(row=conf["row"], column=idx_fim_1b).value
        original_marker = f"{dia_orig} {ini_orig}-{fim_orig}"

        dur = conf["dur"]
        novo_slot = None
        dia_destino = None
        for d_date, d_str in dias_a_partir:
            ocupados = _obter_ocupados(d_str)
            # se é o mesmo dia da nova OS, busca depois do fim da nova
            start_min = fim_nova_min + GAP_ENTRE_OS_MIN if d_str == dia_str_nova else HORA_INICIO_TURNO_MIN + 30
            slot = _achar_slot_livre(ocupados, dur, start_min)
            if slot:
                novo_slot = slot
                dia_destino = d_str
                break

        if novo_slot is None:
            # Overflow → vira pendente
            row_vals = [ws_equipe.cell(row=conf["row"], column=j + 1).value for j in range(max(header_map.values()) + 1)]
            _registrar_pendente(wb_prog, equipe_aba, row_vals, header_map,
                                f"Deslocada por OS #{os_id_nova} e sem capacidade na semana")
            # Marca a linha original como "Não Realizada"
            if idx_desloc_1b:
                ws_equipe.cell(row=conf["row"], column=idx_desloc_1b).value = (
                    f"NÃO REALIZADA - deslocada de {original_marker} por OS #{os_id_nova}"
                )
            if idx_update_1b:
                ws_equipe.cell(row=conf["row"], column=idx_update_1b).value = timestamp
            fill = PatternFill("solid", fgColor="F4B084")  # laranja escuro
            for j in range(1, max(cols_dict.values()) + 1):
                ws_equipe.cell(row=conf["row"], column=j).fill = fill
            pendentes += 1
        else:
            ini_novo, fim_novo = novo_slot
            ws_equipe.cell(row=conf["row"], column=idx_dia_1b).value = dia_destino
            ws_equipe.cell(row=conf["row"], column=idx_ini_1b).value = _hora_min_to_str(ini_novo)
            ws_equipe.cell(row=conf["row"], column=idx_fim_1b).value = _hora_min_to_str(fim_novo)
            if idx_desloc_1b:
                ws_equipe.cell(row=conf["row"], column=idx_desloc_1b).value = (
                    f"Deslocada de {original_marker} por OS #{os_id_nova}"
                )
            if idx_update_1b:
                ws_equipe.cell(row=conf["row"], column=idx_update_1b).value = timestamp
            fill = PatternFill("solid", fgColor="FFD966")  # amarelo
            for j in range(1, max(cols_dict.values()) + 1):
                ws_equipe.cell(row=conf["row"], column=j).fill = fill
            deslocadas += 1

    return deslocadas, pendentes



# ====================== Backup ======================

def fazer_backup(arquivo):
    """Backup tolerante a falha."""
    if not os.path.exists(arquivo):
        return None
    try:
        bkp_dir = os.path.join(os.path.dirname(arquivo), "_backup_atualizacao")
        os.makedirs(bkp_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome, ext = os.path.splitext(os.path.basename(arquivo))
        dest = os.path.join(bkp_dir, f"{nome}_{ts}{ext}")
        shutil.copy2(arquivo, dest)
        bkps = sorted(Path(bkp_dir).glob(f"{nome}_*{ext}"),
                      key=lambda p: p.stat().st_mtime, reverse=True)
        for b in bkps[10:]:
            try:
                b.unlink()
            except OSError:
                pass
        return dest
    except Exception as e:
        log(f"  ! Falha no backup (não-crítico): {type(e).__name__}: {e}", "WARN")
        return None


# ====================== Observações da semana ATUAL (override manual — STEP A3) ======================
# Mesma gramática de Observacoes_Semana.txt (usada pelo programacao_v7 na sexta),
# porém aplicada à SEMANA ATUAL, no lugar:
#     OS_ID; DIA; TAREFAS; TURNO   -> reposiciona a OS (Dia/Hora) e empurra conflitantes
#     OS_ID; nao;                  -> remove a OS da semana (nem como pendente)
_DIA_MAP_OBS = {'SEG': 0, 'TER': 1, 'QUA': 2, 'QUI': 3, 'SEX': 4}
_TURNO_MAP_OBS = {'MANHA': 7 * 60 + 30, 'MANHÃ': 7 * 60 + 30, 'TARDE': 13 * 60, 'NOITE': 18 * 60}
OBS_ATUAL_FILE = "Observacoes_Semana_Atual.txt"


def _parse_observacoes_atual(linhas):
    """Devolve (pins, excluidas). pins: {os_id: {dia:int|None, tarefas:[str]|None, start_min:int|None}}."""
    pins, excluidas = {}, set()
    for ln in linhas:
        ln = (ln or "").strip()
        if not ln or ln.startswith("#"):
            continue
        parts = [p.strip() for p in ln.split(';')]
        if not parts or not parts[0].isdigit():
            continue  # linha de texto livre — ignora
        os_id = int(parts[0])
        dia_str = (parts[1].upper() if len(parts) > 1 else '')
        if dia_str in ('NAO', 'NÃO', 'NO'):
            excluidas.add(os_id)
            continue
        tarefas_s = parts[2] if len(parts) > 2 else ''
        turno_str = (parts[3] if len(parts) > 3 else '').replace('Ã', 'A').replace('ã', 'a').upper().strip()
        pins[os_id] = {
            'dia': _DIA_MAP_OBS.get(dia_str),
            'tarefas': [t.strip().lower() for t in tarefas_s.split(',') if t.strip()] or None,
            'start_min': _TURNO_MAP_OBS.get(turno_str),
        }
    return pins, excluidas


def aplicar_observacoes_semana_atual(wb_prog, dias_semana, hoje):
    """STEP A3: aplica Observacoes_Semana_Atual.txt à semana atual (override do usuário).
    Atualiza no lugar. Robusto: uma observação ruim nunca corrompe o arquivo (só loga WARN)."""
    path = os.path.join(BASE_DIR, OBS_ATUAL_FILE)
    if not os.path.exists(path):
        return
    try:
        linhas = open(path, encoding="utf-8").read().splitlines()
    except Exception as e:
        log(f"! Não consegui ler {OBS_ATUAL_FILE}: {e}", "WARN")
        return
    pins, excluidas = _parse_observacoes_atual(linhas)
    if not pins and not excluidas:
        return
    log(f"STEP A3: Observações da semana atual — {len(pins)} pin(s), {len(excluidas)} exclusão(ões)...")

    def _match_tarefa(cell_val, alvos):
        if not alvos:
            return True
        t = str(cell_val or "").lower()
        return any(a in t for a in alvos)

    # 1) EXCLUSÕES: apaga as linhas da OS de qualquer aba que tenha "OSs ID" (inclui _Pendentes)
    n_excl = 0
    if excluidas:
        for sheet_name in list(wb_prog.sheetnames):
            ws = wb_prog[sheet_name]
            if ws.max_row < 2:
                continue
            header = [c.value for c in ws[1]]
            if "OSs ID" not in header:
                continue
            idx_os0 = header.index("OSs ID")
            for r in range(ws.max_row, 1, -1):   # de baixo p/ cima (delete seguro)
                try:
                    v = int(ws.cell(row=r, column=idx_os0 + 1).value)
                except (ValueError, TypeError):
                    continue
                if v in excluidas:
                    ws.delete_rows(r, 1)
                    n_excl += 1
        log(f"    exclusões: {n_excl} linha(s) removida(s) {sorted(excluidas)}")

    # 2) PINS: reposiciona (Dia/Hora) e empurra conflitantes
    dias_por_idx = {i: (d_date, d_str) for i, (d_date, d_str) in enumerate(dias_semana)}
    n_pin = 0
    for os_id, pin in pins.items():
        alvo_idx = pin['dia']
        if alvo_idx is None and pin['start_min'] is None:
            continue  # sem dia nem turno: nada a reposicionar
        for sheet_name in list(wb_prog.sheetnames):
            if sheet_name.startswith("_"):
                continue
            ws = wb_prog[sheet_name]
            if ws.max_row < 2:
                continue
            try:
                cols = garantir_colunas_extras(ws)
            except Exception:
                continue
            idx_os = cols.get("OSs ID")
            if not idx_os:
                continue
            idx_dia, idx_ini, idx_fim = cols.get("Dia"), cols.get("Hora Início"), cols.get("Hora Fim")
            idx_dur, idx_tar = cols.get("Duração (h)"), cols.get("Tarefa")
            for r in range(2, ws.max_row + 1):
                try:
                    if int(ws.cell(row=r, column=idx_os).value) != os_id:
                        continue
                except (ValueError, TypeError):
                    continue
                if not _match_tarefa(ws.cell(row=r, column=idx_tar).value if idx_tar else None, pin['tarefas']):
                    continue
                try:
                    # novo dia
                    if alvo_idx is not None and alvo_idx in dias_por_idx and idx_dia:
                        d_date, d_str = dias_por_idx[alvo_idx]
                        ws.cell(row=r, column=idx_dia, value=d_str)
                    else:
                        d_str = ws.cell(row=r, column=idx_dia).value if idx_dia else None
                    # novo horário
                    ini_min = pin['start_min']
                    if ini_min is None and idx_ini:
                        ini_min = _hora_str_to_min(ws.cell(row=r, column=idx_ini).value)
                    dur = 0.0
                    if idx_dur:
                        try:
                            dur = float(ws.cell(row=r, column=idx_dur).value or 0)
                        except (ValueError, TypeError):
                            dur = 0.0
                    fim_min = (ini_min + int(round(dur * 60))) if ini_min is not None else None
                    if ini_min is not None and idx_ini:
                        ws.cell(row=r, column=idx_ini, value=_hora_min_to_str(ini_min))
                    if fim_min is not None and idx_fim:
                        ws.cell(row=r, column=idx_fim, value=_hora_min_to_str(fim_min))
                    n_pin += 1
                    # cascata: empurra conflitantes no novo slot
                    if d_str and ini_min is not None and fim_min is not None and ini_min < fim_min:
                        deslocar_conflitantes(wb_prog, sheet_name, d_str, ini_min, fim_min,
                                              dias_semana, hoje, os_id, ws_equipe=ws)
                except Exception as e:
                    log(f"    ! pin OS #{os_id} (linha {r}, aba {sheet_name}): {e}", "WARN")
    if pins:
        log(f"    pins: {n_pin} linha(s) reposicionada(s)")


# ====================== Main ======================

def main():
    parser = argparse.ArgumentParser(
        description="Atualização semanal: sync com BD + sugestões PCM"
    )
    parser.add_argument("--no-sync", action="store_true", help="Pula STEP A")
    parser.add_argument("--no-sugestoes", action="store_true", help="Pula STEP B")
    parser.add_argument("--semana", type=int, default=None, help="Força semana")
    parser.add_argument("--arquivo", type=str, default=None,
                        help="Caminho do arquivo de Programação Semanal")
    args = parser.parse_args()

    log("=" * 70)
    log("ATUALIZAÇÃO SEMANAL - Sync com BD + Sugestões PCM")
    log("=" * 70)

    if args.semana:
        ano = date.today().year
        mon = date.fromisocalendar(ano, args.semana, 1)
    else:
        mon, _ = _semana_atual()
    _, iso_week, _ = mon.isocalendar()
    semana_str = f"{iso_week:02d}"
    week_end = mon + timedelta(days=7)
    log(f"Semana corrente: {semana_str} ({mon} a {week_end - timedelta(days=1)})")

    arquivo_prog = args.arquivo or os.path.join(BASE_DIR, f"Programação Semana {semana_str}.xlsx")
    if not os.path.exists(arquivo_prog):
        log(f"! Arquivo não encontrado: {arquivo_prog}", "ERROR")
        log("Rode o programacao_v7.py primeiro.", "ERROR")
        sys.exit(1)
    log(f"Arquivo de Programação: {os.path.basename(arquivo_prog)}")

    bkp = fazer_backup(arquivo_prog)
    if bkp:
        log(f"Backup: {os.path.basename(bkp)}")

    dias_semana = []
    for i in range(5):
        d = mon + timedelta(days=i)
        dia_str = f"{DAY_NAMES_FULL[d.weekday()]} ({d.strftime('%d/%m')})"
        dias_semana.append((d, dia_str))

    log("Abrindo arquivo de Programação Semanal...")
    wb_prog = _load_workbook_robusto(arquivo_prog)
    if wb_prog is None:
        log("! Não foi possível abrir o arquivo. Abortando esta execução.", "ERROR")
        log("  Próxima execução do agendador vai tentar de novo.", "INFO")
        sys.exit(0)

    df_bd_completo, df_bd_semana = carregar_bd_semanal(mon, week_end)
    bd_acessivel = df_bd_completo is not None
    if not bd_acessivel:
        log("! BD não acessível agora. Pulando STEP A.", "WARN")
        log("  Sugestões PCM continuam (com info parcial). Próxima execução tenta de novo.", "INFO")
        df_bd_completo = pd.DataFrame()
        df_bd_semana = pd.DataFrame()
    mapa_aux = carregar_auxiliar()
    lista_prior = carregar_lista_prioridades()
    conf_items = carregar_confiabilidade()
    feriados = carregar_feriados()
    # Carrega map Responsável O&M (UFV → Responsável) reaproveitando AUXILIAR
    global _MAP_RESPONSAVEL
    try:
        _df_aux_full = fonte_bd_api.df_auxiliar()   # era _read_excel_robusto(INPUT_BD, sheet_name="AUXILIAR")
        _MAP_RESPONSAVEL = carregar_map_responsavel(_df_aux_full)
    except Exception as e:
        log(f"  ! Falha ao carregar map Responsavel: {e}", "WARN")
        _MAP_RESPONSAVEL = {}

    # STEP B primeiro para ter as sugestões prontas (necessárias no STEP A2)
    sugestoes_pcm = {}
    if not args.no_sugestoes:
        df_solic = carregar_solicitacoes()
        if df_solic is None:
            log("STEP B PULADO (falha ao ler solicitações)", "WARN")
            sugestoes_pcm = carregar_sugestoes_pcm_anteriores()
        elif len(df_solic) > 0:
            # Filtra dias passados: não faz sentido sugerir PCM pra dias que já acabaram.
            # dias_semana original é mantido pras outras rotinas (sync, incorporação de OS novas).
            _hoje = date.today()
            dias_semana_futuro = [(d, ds) for d, ds in dias_semana if d >= _hoje]
            if not dias_semana_futuro:
                log("  → Semana toda já passou; sugestões ficarão sem slot nesta rodada.", "WARN")
            sugestoes_pcm = gerar_sugestoes_pcm(
                df_solic, mapa_aux, df_bd_completo, lista_prior,
                conf_items, wb_prog, dias_semana_futuro, feriados=feriados
            )
        else:
            sugestoes_pcm = carregar_sugestoes_pcm_anteriores()
            log("  → Nenhuma solicitação aberta nova")
    else:
        log("STEP B: PULADO (--no-sugestoes)")
        sugestoes_pcm = carregar_sugestoes_pcm_anteriores()

    # STEP A: sync de status + incorporar OS novas (com link às sugestões)
    if not args.no_sync and bd_acessivel:
        sync_status_bd(wb_prog, df_bd_completo)
        incorporar_os_novas_do_bd(wb_prog, df_bd_semana, sugestoes_pcm, dias_semana, date.today())
        # STEP A3: aplica as observações da semana ATUAL (override manual do usuário)
        aplicar_observacoes_semana_atual(wb_prog, dias_semana, date.today())
    elif args.no_sync:
        log("STEP A: PULADO (--no-sync)")
    elif args.no_sync:
        log("STEP A: PULADO (--no-sync)")
    else:
        log("STEP A: PULADO (BD inacessivel)", "WARN")

    if sugestoes_pcm:
        escrever_arquivo_sugestoes_pcm(sugestoes_pcm)

    log(f"Salvando {os.path.basename(arquivo_prog)}...")
    if not _save_workbook_robusto(wb_prog, arquivo_prog, etapa="Programação Semanal"):
        log("! Mudanças desta execução NÃO foram salvas. Próxima execução refaz.", "WARN")

    log("=" * 70)
    log("CONCLUIDO")
    log("FLUXO:")
    log(f"  1. PCM abre {os.path.basename(OUTPUT_SUGEST)} e olha as sugestões")
    log("  2. PCM cria a OS no sistema com a data/hora desejada")
    log("  3. Script roda de novo, detecta a OS no BD, e incorpora à Programação")
    log("     usando a data/hora QUE O PROGRAMADOR DEFINIU")
    log("=" * 70)


def _executar_com_protecao():
    import traceback
    lock = None
    try:
        parser = argparse.ArgumentParser(add_help=False)
        parser.add_argument("--semana", type=int, default=None)
        parser.add_argument("--arquivo", type=str, default=None)
        args, _ = parser.parse_known_args()
        if args.arquivo:
            arquivo_prog = args.arquivo
        else:
            if args.semana:
                ano = date.today().year
                mon = date.fromisocalendar(ano, args.semana, 1)
            else:
                mon, _ = _semana_atual()
            _, iso_week, _ = mon.isocalendar()
            arquivo_prog = os.path.join(BASE_DIR, f"Programação Semana {iso_week:02d}.xlsx")
        lock = _aquire_lock(arquivo_prog)
        if lock is None:
            sys.exit(0)
        main()
    except SystemExit:
        raise
    except Exception:
        log("ERRO NAO TRATADO no executor principal:", "ERROR")
        log(traceback.format_exc(), "ERROR")
        sys.exit(1)
    finally:
        if lock:
            _release_lock(lock)


if __name__ == "__main__":
    _executar_com_protecao()
