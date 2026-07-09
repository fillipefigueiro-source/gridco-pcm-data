"""
Gera BD_Relatório Semanal.xlsx a partir da API REST do Fracttal One.

Substitui a exportação manual da plataforma. Lê todas as Ordens de Trabalho
(OTs) com suas tarefas, transforma para o formato do BD atual, e grava em
BD_Relatório Semanal.xlsx. A aba AUXILIAR vem de um arquivo separado
(configurado em AUXILIAR_SOURCE_PATH).

Uso:
    python gerar_bd_via_api.py            # gera o BD completo (~2 min, 17.900 OTs)
    python gerar_bd_via_api.py --test     # só testa autenticação
    python gerar_bd_via_api.py --limit 50 # gera só com 50 OTs (debug)
    python gerar_bd_via_api.py --com-subtarefas # tenta descobrir endpoint subtasks (experimental)
    python gerar_bd_via_api.py --no-auxiliar    # pula aba AUXILIAR

Credenciais: ler arquivo .env (NUNCA commitar):
    FRACTTAL_CLIENT_ID=<seu_id>
    FRACTTAL_CLIENT_SECRET=<seu_secret>
    FRACTTAL_BASE_URL=https://app.fracttal.com/api/
    AUXILIAR_SOURCE_PATH=<caminho do arquivo SharePoint com a aba AUXILIAR>
"""

import argparse
import json
import os
import pathlib
import sys
import time
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

import pandas as pd
import requests

# ============================================================
# Configuração (carrega de .env, NUNCA hardcoded)
# ============================================================

_sys_init = sys
try:
    _sys_init.stdout.reconfigure(encoding="utf-8", errors="replace")
    _sys_init.stderr.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, Exception):
    pass


def carregar_env():
    """Lê o arquivo .env na pasta do script (se existir) e popula env vars."""
    try:
        env_path = pathlib.Path(__file__).resolve().parent / ".env"
    except NameError:
        env_path = pathlib.Path(".env")
    if not env_path.exists():
        return False
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
    return True


carregar_env()

CLIENT_ID = os.environ.get("FRACTTAL_CLIENT_ID", "")
CLIENT_SECRET = os.environ.get("FRACTTAL_CLIENT_SECRET", "")
BASE_URL = os.environ.get("FRACTTAL_BASE_URL", "https://app.fracttal.com/api/").rstrip("/") + "/"
AUXILIAR_SOURCE_PATH = os.environ.get("AUXILIAR_SOURCE_PATH", "")
# JWT Bearer Token do app.fracttal.com (capturado via DevTools).
# Necessário SÓ pra chamar /rpc/proxy (companies.transactions_log_wo_list).
# Expira em ~12 HORAS! Quando expirar, capturar novo token e atualizar .env.
# Captura: DevTools → Application → Local Storage → app.fracttal.com → id_token
FRACTTAL_JWT_TOKEN = os.environ.get("FRACTTAL_JWT_TOKEN", "")

# Detecção da pasta base (mesmo padrão dos outros scripts)
def detectar_base_dir():
    if os.environ.get("PCM_PROG_DIR") and os.path.isdir(os.environ["PCM_PROG_DIR"]):
        return os.environ["PCM_PROG_DIR"]
    try:
        sd = pathlib.Path(__file__).resolve().parent
        if (sd / "BD_Relatório Semanal.xlsx").exists() or (sd / ".env").exists():
            return str(sd)
    except NameError:
        pass
    sub = pathlib.Path("4. O&M") / "11.Pré-Operação" / "6. PCM" / "09. Programação Semanal"
    home = pathlib.Path.home()
    for c in [home / "GRID CO" / "Grid Co. - Gridco" / sub,
              home / "OneDrive - Grid Co" / sub,
              home / "OneDrive - Gridco" / sub,
              home / "OneDrive" / "Grid Co. - Gridco" / sub]:
        try:
            if c.exists() and c.is_dir():
                return str(c)
        except OSError:
            continue
    return None


BASE_DIR = detectar_base_dir() or "."
OUTPUT_BD = os.path.join(BASE_DIR, "BD_Relatório Semanal.xlsx")
OUTPUT_BD_BACKUP_DIR = os.path.join(BASE_DIR, "_backup_BD")


# ============================================================
# Logging
# ============================================================

def log(msg, level="INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {level:5s} {msg}", flush=True)


# ============================================================
# Cliente Fracttal (OAuth2 client credentials + retry)
# ============================================================

class FracttalClient:
    def __init__(self, client_id, client_secret, base_url):
        if not client_id or not client_secret:
            raise RuntimeError(
                "Credenciais ausentes. Crie um arquivo .env na pasta com:\n"
                "  FRACTTAL_CLIENT_ID=...\n"
                "  FRACTTAL_CLIENT_SECRET=...\n"
                "Veja .env.example para o template."
            )
        self.client_id = client_id
        self.client_secret = client_secret
        self.base_url = base_url
        self.token = None
        self.token_expires_at = 0
        self.session = requests.Session()

    # Variantes de path conhecidas pra cada recurso da v1
    PATH_VARIANTS = {
        "work_orders": ["work_orders", "workorders", "work-orders", "tasks/wo",
                         "tasks_work_orders", "v1/work_orders"],
        # FIX 2026-06-15: endpoint usado pra obter duração POR subtarefa
        # (a view 'work_orders' retorna duration da OS pai em todas as linhas,
        # inflando HH na Programação Semana — ex: OS 7255 com 62×95.5h).
        "work_order_tasks": ["work_order_tasks", "work_orders_tasks",
                              "work-order-tasks", "tasks/wo_tasks"],
        "work_orders_task_subtasks": [
            "work_orders_task_subtasks", "work-orders-task-subtasks",
            "work_orders/tasks/subtasks", "tasks/subtasks", "subtasks",
            "wo_task_subtasks",
        ],
        "tasks_todo": ["tasks_todo", "tasks/todo", "tasks_pending"],
        "assets": ["items", "assets", "equipments"],
        "personnel": ["personnel", "human_resources", "accounts"],
        "requests": ["requests", "work_requests"],
    }

    def _resolver_path(self, path):
        """Descobre base+recurso corretos. Tenta combinacoes de base x resource ate
        achar uma que devolva 200. Cacheia em _path_cache."""
        if not hasattr(self, "_path_cache"):
            self._path_cache = {}
        if path in self._path_cache:
            return self._path_cache[path]
        variantes = self.PATH_VARIANTS.get(path, [path])
        bases = [self.base_url] + getattr(self, "_data_bases_alt", [])
        # Remove duplicatas mantendo ordem
        seen = set(); bases_uniq = []
        for b in bases:
            if b not in seen:
                seen.add(b); bases_uniq.append(b)

        log(f"  Descobrindo base+path correto para '{path}'...")
        for base in bases_uniq:
            for v in variantes:
                url = base + v.lstrip("/")
                try:
                    time.sleep(0.2)
                    r = self.session.get(url, headers=self._headers(),
                                         params={"limit": 1, "page": 1}, timeout=15)
                    if r.status_code == 200:
                        log(f"  → Resolvido: {url}")
                        self.base_url = base
                        self._path_cache[path] = v
                        return v
                    elif r.status_code == 404:
                        pass  # nao polui log
                    elif r.status_code == 400:
                        # Diagnostico mais util
                        snippet = r.text[:90]
                        log(f"    {url} → 400: {snippet}", "WARN")
                    else:
                        log(f"    {url} → HTTP {r.status_code}: {r.text[:90]}", "WARN")
                except Exception as e:
                    log(f"    {url} → erro: {e}", "WARN")
        log(f"  ! Nenhuma combinacao base+'{path}' respondeu 200; usando padrao", "WARN")
        self._path_cache[path] = path
        return path

    # Endpoints de OAuth conhecidos do Fracttal (relativo ao host)
    OAUTH_ENDPOINTS = [
        "oauth/token",
        "api/oauth/token",
        "api/v1/oauth/token",
        "api/v2/oauth/token",
        "v1/oauth/token",
        "v2/oauth/token",
        "auth/token",
        "oauth2/token",
        "api/access_tokens",
        "access_tokens",
    ]
    # Hosts conhecidos do Fracttal
    HOSTS_CANDIDATOS = [
        "https://app.fracttal.com",
        "https://api.fracttal.com",
    ]

    def autenticar(self):
        """OAuth2 client_credentials grant.
        Enumera várias combinações de host + endpoint até achar uma que retorne 200."""
        log("Autenticando no Fracttal (tentando combinações conhecidas)...")
        # Lista de URLs completas pra tentar, na ordem de prioridade
        urls_candidatas = []

        # 1. Primeiro tenta o que estava em base_url (com endpoint padrão)
        if self.base_url:
            base_lim = self.base_url.rstrip("/")
            for ep in self.OAUTH_ENDPOINTS:
                urls_candidatas.append(base_lim + "/" + ep)

        # 2. Depois enumera todas as combinações host × endpoint
        for host in self.HOSTS_CANDIDATOS:
            for ep in self.OAUTH_ENDPOINTS:
                u = host + "/" + ep
                if u not in urls_candidatas:
                    urls_candidatas.append(u)

        ultimo_erro = None
        for token_url in urls_candidatas:
            # Forma 1: Basic auth + form-urlencoded
            r = self._tentar_auth(token_url, "basic-form")
            if r == "ok":
                self._set_data_base(token_url)
                return
            ultimo_erro = r
            # Forma 2: tudo no body form-urlencoded
            r = self._tentar_auth(token_url, "body-form")
            if r == "ok":
                self._set_data_base(token_url)
                return
            ultimo_erro = r
            # Pausa curta entre tentativas pra evitar rate limit
            time.sleep(0.6)

        raise RuntimeError(
            f"Não consegui autenticar em NENHUM endpoint testado ({len(urls_candidatas)} tentativas). "
            f"Último erro: {ultimo_erro}. Possíveis causas: "
            f"(1) credenciais erradas/expiradas — verifique no Fracttal; "
            f"(2) tipo do OAuth client errado — deve ser 'Credenciais do cliente'; "
            f"(3) endpoint REST diferente — consulte docs.fracttal.com ou suporte."
        )

    def _tentar_auth(self, token_url, modo):
        """Tenta autenticar em 1 URL+modo. Retorna 'ok' se sucesso, ou string de erro."""
        try:
            if modo == "basic-form":
                r = self.session.post(
                    token_url,
                    auth=(self.client_id, self.client_secret),
                    data={"grant_type": "client_credentials"},
                    timeout=15,
                )
                rotulo = "basic+form"
            else:  # body-form
                r = self.session.post(
                    token_url,
                    data={
                        "grant_type": "client_credentials",
                        "client_id": self.client_id,
                        "client_secret": self.client_secret,
                    },
                    timeout=15,
                )
                rotulo = "body+form"

            if r.status_code == 200:
                try:
                    self._processar_token(r.json())
                    return "ok"
                except Exception as e:
                    return f"200 mas sem JSON válido: {e}"

            # Log compacto pra não inundar
            txt = r.text[:120].replace("\n", " ")
            log(f"  [{rotulo}] {token_url} → HTTP {r.status_code}: {txt}", "WARN")
            return f"HTTP {r.status_code}"
        except Exception as e:
            log(f"  [{modo}] {token_url} → erro de rede: {e}", "WARN")
            return str(e)

    def _processar_token(self, payload):
        self.token = payload.get("access_token") or payload.get("token")
        ttl = int(payload.get("expires_in") or 3600)
        self.token_expires_at = time.time() + ttl - 60
        log(f"  → Token OK (válido por {ttl}s)")

    def _set_data_base(self, token_url):
        """Deriva a base URL pra DADOS. Fracttal v1 (app.fracttal.com) usa /api/ sem versao."""
        from urllib.parse import urlparse
        parsed = urlparse(token_url)
        host = f"{parsed.scheme}://{parsed.netloc}"
        path = parsed.path
        if "/api/v2/" in path:
            api_prefix = "/api/v2/"
        elif "/api/v1/" in path:
            api_prefix = "/api/v1/"
        else:
            # Default Fracttal v1: /api/ (sem versao no path)
            api_prefix = "/api/"
        self.base_url = host + api_prefix
        log(f"  → Endpoint OAuth: {token_url}")
        log(f"  → Base URL p/ dados: {self.base_url}")
        # Bases alternativas pra tentar se o path padrao 404
        self._data_bases_alt = [
            host + "/api/",
            host + "/api/v1/",
            host + "/api/v2/",
            host + "/api/web/v1/",
            "https://api.fracttal.com/",
            "https://api.fracttal.com/v1/",
            "https://api.fracttal.com/v2/",
        ]

    def _headers(self):
        if not self.token or time.time() > self.token_expires_at:
            self.autenticar()
        return {"Authorization": f"Bearer {self.token}",
                "Accept": "application/json"}

    def get(self, path, params=None, max_tentativas=3):
        url = self.base_url + path.lstrip("/")
        time.sleep(0.15)
        last_err = None
        for tentativa in range(1, max_tentativas + 1):
            try:
                r = self.session.get(url, headers=self._headers(),
                                     params=params, timeout=60)
                if r.status_code == 401:
                    log("  401 → reautenticando...", "WARN")
                    self.token = None
                    continue
                if 400 <= r.status_code < 500:
                    # Mostra o corpo do erro pra diagnostico
                    log(f"  HTTP {r.status_code} em {url}", "WARN")
                    log(f"  Params: {params}", "WARN")
                    log(f"  Resposta: {r.text[:500]}", "WARN")
                    raise RuntimeError(f"HTTP {r.status_code}: {r.text[:300]}")
                if r.status_code >= 500:
                    raise RuntimeError(f"HTTP {r.status_code}: {r.text[:200]}")
                r.raise_for_status()
                return r.json()
            except Exception as e:
                last_err = e
                log(f"  Tentativa {tentativa}/{max_tentativas} GET {path} falhou: {e}", "WARN")
                if tentativa < max_tentativas:
                    time.sleep(3 * tentativa)
        raise RuntimeError(f"GET {path} falhou após {max_tentativas} tentativas: {last_err}")

    def paginar(self, path, params=None, page_param="page", size_param="limit",
                page_size=100, max_items=None):
        """Generator que itera por todas as paginas usando OFFSET (start).

        IMPORTANTE: A API REST Fracttal IGNORA o parametro 'page' e sempre retorna
        a primeira pagina! O parametro funcional eh 'start' (offset 0, 100, 200, ...).
        Descoberto via debug_paginacao.py em 14/06/2026.
        """
        path = self._resolver_path(path)
        params = dict(params or {})
        params[size_param] = page_size
        start = 0
        total = 0
        while True:
            params["start"] = start
            payload = self.get(path, params=params)
            # Fracttal: { "success": true, "data": [...], "total": N, "start": ..., "limit": ... }
            if isinstance(payload, dict):
                items = payload.get("data") or payload.get("items") or payload.get("results") or []
                total_remoto = payload.get("total")
            elif isinstance(payload, list):
                items = payload
                total_remoto = None
            else:
                break
            if not items:
                break
            for it in items:
                yield it
                total += 1
                if max_items and total >= max_items:
                    return
            if len(items) < page_size:
                break
            # NB (2026-07-07): a API passou a devolver 'total' = tamanho da PÁGINA
            # (=page_size), não o total geral -> confiar nele fazia o paginador parar
            # na 1ª página. Só usa 'total' como corte quando ele é claramente o total
            # geral (> page_size); caso contrário termina pela regra len(items)<page_size.
            if total_remoto is not None and int(total_remoto) > page_size and total >= int(total_remoto):
                break
            start += page_size


# Mapa: type 4 (VERIFICATION) valor numerico -> rotulo textual
_VERIF_LABELS = {1: "Sim", "1": "Sim", 2: "Alerta", "2": "Alerta", 3: "Nao", "3": "Nao"}


def _decodificar_valor(v, tipo_id):
    """Converte o valor cru da subtask em texto legivel.
    tipo_id: id_task_form_item_type (1=TEXT, 2=BOOLEAN, 3=NUMBER, 4=VERIFICATION,
    5=READING_METER, 7=DROPDOWN, 8=DATE_TIME, 9=AI)."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Sim" if v else "Nao"
    if tipo_id == 4:
        # Verification: 1=Sim, 2=Alerta, 3=Nao
        return _VERIF_LABELS.get(v, str(v))
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    s = str(v).strip()
    if not s or s.lower() in ("null", "none"):
        return ""
    return s


def enriquecer_com_subtarefas(client, rows):
    """Adiciona Descricao subtarefa e Valor preenchido para cada OT-task.

    Estrategia: usa endpoint /api/work_orders_task_subtasks com filtro JSON,
    que e o padrao da API Fracttal (1 chamada por OT-task).
    Concatena subtarefas com '|', so as que tem valor preenchido."""
    log("Coletando subtarefas (1 chamada por OT-task)...")
    enriquecidas = 0
    sem_subtarefa = 0
    falhas = 0

    # ID conhecido pra teste: OS 5547 / Tarefa Ar Condicionado (sabemos que tem 6 subtasks)
    ID_TESTE_CONHECIDO = 51518601

    # Coleta amostras: primeiro o ID conhecido, depois 300 rows da lista atual
    amostras = [ID_TESTE_CONHECIDO]
    amostras.extend([r["_id_work_orders_tasks"] for r in rows[:300]
                     if r.get("_id_work_orders_tasks")
                     and r["_id_work_orders_tasks"] != ID_TESTE_CONHECIDO])
    if not amostras:
        log("  ! Nenhuma row tem id_work_orders_tasks - pulando subtarefas", "WARN")
        return rows
    log(f"  Total {len(amostras)} amostras (1a = {ID_TESTE_CONHECIDO} OS 5547 conhecida)")

    def fazer_filter_obj(id_wot):
        return [{"property": "id_work_order_task",
                 "operator": "eq",
                 "value": int(id_wot),
                 "condition": "and"}]

    def fazer_filter_str(id_wot):
        return json.dumps(fazer_filter_obj(id_wot))

    # cada candidato = (descricao, builder_url, metodo, params_ou_body)
    # builder retorna (url, params_dict_se_GET_ou_None, json_body_se_POST_ou_None)
    candidatos = [
        # POSTs com JSON body — padrao tipico de endpoints _list
        ("POST work_orders_task_subtasks_list (json body)",
         lambda b, i_wot: (b + "work_orders_task_subtasks_list",
                           None,
                           {"filter": fazer_filter_obj(i_wot), "limit": 100})),
        ("POST work_orders_task_subtasks (json body)",
         lambda b, i_wot: (b + "work_orders_task_subtasks",
                           None,
                           {"filter": fazer_filter_obj(i_wot), "limit": 100})),
        # GETs com query params
        ("GET work_orders_task_subtasks?filter=JSON",
         lambda b, i_wot: (b + "work_orders_task_subtasks",
                           {"filter": fazer_filter_str(i_wot), "limit": 100},
                           None)),
        ("GET work_orders_task_subtasks?id_work_order_task=<id>",
         lambda b, i_wot: (b + "work_orders_task_subtasks",
                           {"id_work_order_task": i_wot, "limit": 100},
                           None)),
        # Path nested
        ("GET work_orders_tasks/<id>/subtasks",
         lambda b, i_wot: (b + f"work_orders_tasks/{i_wot}/subtasks", None, None)),
        ("GET work_orders_task_subtasks/<id>",
         lambda b, i_wot: (b + f"work_orders_task_subtasks/{i_wot}", None, None)),
        # Outras variacoes
        ("GET subtasks?filter=JSON",
         lambda b, i_wot: (b + "subtasks",
                           {"filter": fazer_filter_str(i_wot), "limit": 100},
                           None)),
        # === Hipotese: nome diferente do recurso ===
        ("GET work_orders_task_form_items_log",
         lambda b, i_wot: (b + "work_orders_task_form_items_log",
                           {"filter": fazer_filter_str(i_wot), "limit": 100},
                           None)),
        ("GET tasks_form_items_log",
         lambda b, i_wot: (b + "tasks_form_items_log",
                           {"filter": fazer_filter_str(i_wot), "limit": 100},
                           None)),
        ("GET form_items_log",
         lambda b, i_wot: (b + "form_items_log",
                           {"filter": fazer_filter_str(i_wot), "limit": 100},
                           None)),
        ("GET work_orders_tasks_form_items",
         lambda b, i_wot: (b + "work_orders_tasks_form_items",
                           {"filter": fazer_filter_str(i_wot), "limit": 100},
                           None)),
        ("GET task_form_items",
         lambda b, i_wot: (b + "task_form_items",
                           {"filter": fazer_filter_str(i_wot), "limit": 100},
                           None)),
        # === Hipotese: detalhe completo da WO traz subtasks inline ===
        # (precisa do id_work_order, nao id_work_order_task — vamos hardcodar OS 5547)
        ("GET work_orders/33241718 (OS 5547 full)",
         lambda b, i_wot: (b + "work_orders/33241718", None, None)),
        ("GET work_orders/33241718?include=subtasks",
         lambda b, i_wot: (b + "work_orders/33241718",
                           {"include": "subtasks"}, None)),
        ("GET work_orders/33241718?include=tasks.subtasks",
         lambda b, i_wot: (b + "work_orders/33241718",
                           {"include": "tasks.subtasks"}, None)),
        ("GET work_orders/33241718?include=form_items_log",
         lambda b, i_wot: (b + "work_orders/33241718",
                           {"include": "form_items_log"}, None)),
        ("GET work_orders/33241718/subtasks",
         lambda b, i_wot: (b + "work_orders/33241718/subtasks", None, None)),
        ("GET work_orders/33241718/form_items_log",
         lambda b, i_wot: (b + "work_orders/33241718/form_items_log", None, None)),
    ]

    # FASE 1: descobre o formato (qual URL responde 200 - usando ID conhecido OS 5547)
    fn_url = None
    formato_desc = ""
    log(f"  FASE 1: descobrindo URL valida (testando com id_wot={ID_TESTE_CONHECIDO} - OS 5547)...")
    # Arquivo de log detalhado pra diagnostico
    debug_log_path = os.path.join(BASE_DIR, "_subtasks_debug.log")
    debug_f = None
    try:
        debug_f = open(debug_log_path, "w", encoding="utf-8")
        debug_f.write(f"=== Debug subtasks discovery {datetime.now().isoformat()} ===\n")
        debug_f.write(f"Base URL: {client.base_url}\n")
        debug_f.write(f"Testando com ID conhecido OS 5547: {ID_TESTE_CONHECIDO}\n\n")
    except Exception as e:
        log(f"  ! Nao consegui criar debug log: {e}", "WARN")

    def _debug(msg):
        if debug_f:
            try:
                debug_f.write(msg + "\n")
                debug_f.flush()
            except Exception:
                pass

    for desc, builder in candidatos:
        url, params, body = builder(client.base_url, ID_TESTE_CONHECIDO)
        metodo = "POST" if body is not None else "GET"
        try:
            time.sleep(0.2)
            if metodo == "POST":
                hdrs = dict(client._headers())
                hdrs["Content-Type"] = "application/json"
                r = client.session.post(url, headers=hdrs, json=body, timeout=15)
            else:
                r = client.session.get(url, headers=client._headers(),
                                       params=params, timeout=15)
            _debug(f"\n--- {desc} ---")
            _debug(f"  Metodo: {metodo}")
            _debug(f"  URL:    {url}")
            _debug(f"  Params: {params}")
            _debug(f"  Body:   {body}")
            _debug(f"  Status: {r.status_code}")
            _debug(f"  Resp:   {r.text[:1000]}")
            if r.status_code == 200:
                payload = r.json() if r.text else {}
                items = payload.get("data") or payload.get("items") or []
                if items:
                    fn_url = builder
                    formato_desc = desc
                    log(f"  -> URL valida: {desc} ({len(items)} subtasks na OS 5547)")
                    _debug(f"  *** ESCOLHIDO ESTE FORMATO ***")
                    break
                else:
                    log(f"    {desc} -> 200 mas vazio", "WARN")
            elif r.status_code in (400, 404):
                log(f"    {desc} -> HTTP {r.status_code}: {r.text[:120]}", "WARN")
            else:
                log(f"    {desc} -> HTTP {r.status_code}: {r.text[:120]}", "WARN")
        except Exception as e:
            log(f"    {desc} -> erro: {e}", "WARN")
            _debug(f"\n--- {desc} ---")
            _debug(f"  EXCEPTION: {type(e).__name__}: {e}")
    if debug_f:
        try:
            debug_f.close()
            log(f"  -> Debug salvo em: {debug_log_path}")
        except Exception:
            pass

    if fn_url is None:
        log("  ! Nenhuma URL respondeu 200. Pulando enrichment.", "WARN")
        return rows

    # FASE 2 removida (FASE 1 ja valida com ID conhecido OS 5547)

    total = len(rows)
    for idx, row in enumerate(rows):
        i_wot = row.get("_id_work_orders_tasks")
        if not i_wot:
            sem_subtarefa += 1
            continue
        try:
            url, params, body = fn_url(client.base_url, i_wot)
            time.sleep(0.04)
            if body is not None:
                hdrs = dict(client._headers())
                hdrs["Content-Type"] = "application/json"
                r = client.session.post(url, headers=hdrs, json=body, timeout=20)
            else:
                r = client.session.get(url, headers=client._headers(),
                                       params=params, timeout=20)
            if r.status_code != 200:
                falhas += 1
                continue
            payload = r.json()
            items = payload.get("data") or payload.get("items") or []
            # Ordena por order_number
            items = sorted(items, key=lambda x: x.get("order_number", 999))
            descs, valores, relatorio_blocos = [], [], []
            for it in items:
                tipo_id = it.get("id_task_form_item_type")
                v = _decodificar_valor(it.get("value"), tipo_id)
                if not v:
                    continue
                d = (it.get("description") or it.get("name") or
                     it.get("question") or "").strip()
                if not d:
                    d = "(sem descricao)"
                descs.append(d)
                valores.append(v)
                # Relatório Execução: texto formatado pra display rico no frontend.
                # Para TEXT longos (>80 chars), só o valor (relatório do executor).
                # Para itens curtos (verificação, número, etc.), "campo: valor".
                if tipo_id == 1 and len(v) > 80:
                    relatorio_blocos.append(v)
                else:
                    relatorio_blocos.append(f"**{d}:** {v}")
            if descs:
                row["Descrição subtarefa"] = " | ".join(descs)
                row["Valor preenchido"] = " | ".join(valores)
                row["Relatório Execução"] = "\n\n".join(relatorio_blocos)
                enriquecidas += 1
            else:
                sem_subtarefa += 1
        except Exception as e:
            falhas += 1
            if falhas < 3:
                log(f"    Erro buscando subtarefas de wot={i_wot}: {e}", "WARN")
        if (idx + 1) % 1000 == 0:
            log(f"  -> {idx+1}/{total} processadas (enriquecidas: {enriquecidas})")

    log(f"  -> Enriquecimento: {enriquecidas} com subtarefa, "
        f"{sem_subtarefa} sem dados, {falhas} falhas")
    return rows


def enriquecer_com_duracoes_subtarefas(client, rows):
    """Substitui 'Tarefa -> Duração estimada' pela duração POR subtarefa,
    obtida via JSON-RPC tasks.work_orders_tasks_list.

    BUG FIX 2026-06-15: A view 'work_orders' (usada na coleta principal)
    retorna o campo `duration` da OS PAI em todas as linhas de subtarefa,
    inflando HH na Programação Semana. Ex: OS 7255 tem 62 subtarefas; a
    view marcava 95.5h em cada (343800s da OS pai), totalizando 5.921h
    de carga semanal. Cada subtarefa tem duração real ~2h (7200s).

    Estratégia: batched por OS — chamada RPC com filter id_work_order.
    Match por id (que é _id_work_orders_tasks no BD).
    """
    log("Enriquecendo durações por subtarefa (RPC work_orders_tasks_list)...")
    # Indexa rows por (id_work_order, id_work_orders_tasks)
    rows_por_wot: Dict[int, List[Dict[str, Any]]] = {}
    work_orders_alvo = set()
    for r in rows:
        wt_id = r.get("_id_work_orders_tasks")
        wo_id = r.get("_id_work_order")
        if wt_id is None or wo_id is None:
            continue
        try:
            rows_por_wot.setdefault(int(wt_id), []).append(r)
            work_orders_alvo.add(int(wo_id))
        except (ValueError, TypeError):
            continue
    if not rows_por_wot:
        log("  ! Nenhuma linha com _id_work_orders_tasks/_id_work_order - pulando", "WARN")
        return
    log(f"  -> {len(work_orders_alvo)} OSs / {len(rows_por_wot)} subtarefas a corrigir")

    if not FRACTTAL_JWT_TOKEN:
        log("  ! FRACTTAL_JWT_TOKEN ausente. Pulando enriquecimento.", "WARN")
        return

    RPC_URL = "https://app.fracttal.com/rpc/proxy"
    HDRS = {
        "Content-Type": "application/json",
        "Accept": "application/json, text/plain, */*",
        "Authorization": f"Bearer {FRACTTAL_JWT_TOKEN}",
        "Origin": "https://app.fracttal.com",
        "Referer": "https://app.fracttal.com/tasks/wo",
    }

    # FASE 1: descobrir método correto testando com 1 OS conhecida
    # Aceita 4 formatos de retorno:
    #   A) {"result": {"success": true, "data": [...]}}
    #   B) {"result": [...]}  (lista no result)
    #   C) [...]              (lista no top-level — não-compliant)
    #   D) {"data": [...]}    (data no top-level)
    def _extrair_data_rpc(payload):
        """Extrai a lista de itens do payload RPC.

        Trata JSON-RPC batch: o servidor Fracttal retorna [{...}] mesmo
        pra request unico, então desembrulhamos o primeiro elemento.
        """
        # Batch response: lista com responses RPC dentro
        if isinstance(payload, list):
            if not payload:
                return None
            first = payload[0]
            # Se primeiro item parece response RPC (tem result ou error), recurse
            if isinstance(first, dict) and ("result" in first or "error" in first):
                return _extrair_data_rpc(first)
            # Senão, é lista de dados direta
            return payload
        if not isinstance(payload, dict):
            return None
        # data direto no top-level
        if "result" not in payload and isinstance(payload.get("data"), list):
            return payload["data"]
        res = payload.get("result")
        if res is None:
            return None
        if isinstance(res, list):
            return res
        if isinstance(res, dict):
            if res.get("success") is False:
                return None
            d = res.get("data")
            if isinstance(d, list):
                return d
        return None

    def _extrair_msg_rpc(payload):
        """Extrai a mensagem de erro do payload RPC (pra diagnostic)."""
        if isinstance(payload, list) and payload:
            return _extrair_msg_rpc(payload[0])
        if isinstance(payload, dict):
            res = payload.get("result")
            if isinstance(res, dict) and "message" in res:
                return res.get("message")
            err = payload.get("error")
            if isinstance(err, dict):
                return err.get("message")
        return None

    # Helper: dump do response cru pra diagnostic
    DEBUG_RPC_LOG = os.path.join(os.path.dirname(OUTPUT_BD), "_duracoes_rpc_debug.log")
    debug_log_handle = None
    try:
        debug_log_handle = open(DEBUG_RPC_LOG, "w", encoding="utf-8")
    except Exception:
        pass

    def _dbg(msg):
        if debug_log_handle:
            try:
                debug_log_handle.write(msg + "\n")
                debug_log_handle.flush()
            except Exception:
                pass

    primeira_wo = sorted(work_orders_alvo)[0]
    metodos_candidatos = [
        "tasks.work_orders_tasks_list",
        "tasks.work_order_tasks_list",
        "work_orders_tasks_list",
        "tasks.work_orders_tasks",
    ]

    def _params_variantes(wo_id):
        """Diferentes formas de passar o filtro id_work_order."""
        return [
            ("filter_array_eq", {
                "filter": [{"property": "id_work_order",
                            "value": wo_id,
                            "operator": "eq",
                            "condition": "and"}],
                "start": 0, "limit": 5,
            }),
            ("direto", {"id_work_order": wo_id, "start": 0, "limit": 5}),
            ("sem_paginacao", {"id_work_order": wo_id}),
            ("filter_obj", {
                "filter": {"id_work_order": wo_id},
                "start": 0, "limit": 5,
            }),
        ]

    metodo_ok = None
    params_ok = None
    log(f"  FASE 1: testando RPC com id_work_order={primeira_wo}...")
    for metodo in metodos_candidatos:
        if metodo_ok:
            break
        for params_label, params_body in _params_variantes(primeira_wo):
            payload = {"jsonrpc": "2.0", "id": 1, "method": metodo, "params": params_body}
            try:
                resp = client.session.post(RPC_URL, json=payload, headers=HDRS, timeout=30)
                _dbg(f"=== {metodo} | params={params_label} ===")
                _dbg(f"HTTP {resp.status_code}")
                _dbg(f"Request body: {json.dumps(payload)[:500]}")
                _dbg(f"Response (primeiros 600): {resp.text[:600]}")
                _dbg("")
                try:
                    data = resp.json()
                except Exception as e_json:
                    log(f"    ✗ {metodo}/{params_label}: não parseou JSON", "WARN")
                    continue
                itens = _extrair_data_rpc(data)
                if itens:
                    amostra = itens[0] if isinstance(itens[0], dict) else None
                    if not amostra:
                        continue
                    # Confirma que tem duration
                    if amostra.get("duration") is None and amostra.get("id_work_order") is None:
                        # Provavelmente catálogo ou outro objeto. Pula.
                        log(f"    ✗ {metodo}/{params_label}: amostra sem duration/id_work_order", "WARN")
                        continue
                    metodo_ok = metodo
                    params_ok = params_label
                    campos_id = ["id", "id_work_orders_tasks", "id_task"]
                    campo_id_detectado = next((c for c in campos_id if amostra.get(c) is not None), "id")
                    log(f"    ✓ Método OK: {metodo} | params={params_label}")
                    log(f"      Tipo do payload: {type(data).__name__}, itens: {len(itens)}")
                    log(f"      Amostra: {campo_id_detectado}={amostra.get(campo_id_detectado)}, "
                        f"duration={amostra.get('duration')}, "
                        f"id_work_order={amostra.get('id_work_order')}")
                    log(f"      Campos ({len(amostra)}): {list(amostra.keys())[:15]}")
                    break
                else:
                    msg_erro = _extrair_msg_rpc(data) or "sem dados"
                    log(f"    ✗ {metodo}/{params_label}: {str(msg_erro)[:80]}", "WARN")
            except Exception as e:
                log(f"    ✗ {metodo}/{params_label}: {type(e).__name__}: {str(e)[:60]}", "WARN")

    if not metodo_ok:
        if debug_log_handle:
            try:
                debug_log_handle.close()
            except Exception:
                pass
        log(f"  ! Nenhum método RPC funcionou. Veja {DEBUG_RPC_LOG} pra response cru.", "WARN")
        return
    if debug_log_handle:
        try:
            debug_log_handle.close()
        except Exception:
            pass

    # FASE 2: iterar por OS e buscar suas subtarefas
    def _construir_params_fase2(wo_id, params_label):
        """Reconstroi os params do tipo que funcionou na FASE 1, com limit maior."""
        if params_label == "filter_array_eq":
            return {
                "filter": [{"property": "id_work_order", "value": wo_id,
                            "operator": "eq", "condition": "and"}],
                "start": 0, "limit": 500,
            }
        if params_label == "direto":
            return {"id_work_order": wo_id, "start": 0, "limit": 500}
        if params_label == "sem_paginacao":
            return {"id_work_order": wo_id}
        if params_label == "filter_obj":
            return {"filter": {"id_work_order": wo_id}, "start": 0, "limit": 500}
        # fallback
        return {"id_work_order": wo_id, "start": 0, "limit": 500}

    # CHECKPOINT: persiste duração por (wo_id, wt_id) em JSON pra resistir
    # a JWT expirar no meio (Fracttal desloga automaticamente).
    CKPT_PATH = os.path.join(os.path.dirname(OUTPUT_BD), "_duracoes_checkpoint.json")
    duracoes_cache = {}
    wos_processadas = set()
    try:
        if os.path.exists(CKPT_PATH):
            with open(CKPT_PATH, "r", encoding="utf-8") as f:
                ckpt = json.load(f)
            duracoes_cache = {int(k): int(v) for k, v in (ckpt.get("dur") or {}).items()}
            wos_processadas = set(int(x) for x in (ckpt.get("wos") or []))
            log(f"  → Checkpoint encontrado: {len(duracoes_cache)} durações em cache, "
                f"{len(wos_processadas)} OSs já processadas")
    except Exception as e:
        log(f"  ! Falha ao ler checkpoint: {e}", "WARN")
        duracoes_cache = {}
        wos_processadas = set()

    # Aplica cache imediatamente (cobre OSs já processadas em rodadas anteriores)
    aplicadas_do_cache = 0
    for wt_id_int, dur_int in duracoes_cache.items():
        if wt_id_int in rows_por_wot:
            for r in rows_por_wot[wt_id_int]:
                r["Tarefa -> Duração estimada"] = _seg_para_timedelta(dur_int)
            aplicadas_do_cache += 1
    if aplicadas_do_cache:
        log(f"  → {aplicadas_do_cache} durações aplicadas do cache (sem chamar API)")

    def _salvar_checkpoint():
        try:
            with open(CKPT_PATH, "w", encoding="utf-8") as f:
                json.dump({
                    "dur": {str(k): v for k, v in duracoes_cache.items()},
                    "wos": list(wos_processadas),
                    "atualizado": datetime.now().isoformat(),
                }, f)
        except Exception:
            pass

    enriquecidas = aplicadas_do_cache
    erros = 0
    erros_consecutivos = 0
    todas_wos = sorted(work_orders_alvo)
    # Pula as já processadas
    pendentes = [w for w in todas_wos if w not in wos_processadas]
    log(f"  → {len(pendentes)} OSs pendentes ({len(wos_processadas)} já feitas)")
    for idx, wo_id in enumerate(pendentes, 1):
        payload = {
            "jsonrpc": "2.0", "id": idx, "method": metodo_ok,
            "params": _construir_params_fase2(wo_id, params_ok),
        }
        try:
            resp = client.session.post(RPC_URL, json=payload, headers=HDRS, timeout=30)
            itens = _extrair_data_rpc(resp.json())
            if itens is None:
                erros += 1
                continue
            for it in itens:
                wt_id = it.get("id") or it.get("id_work_orders_tasks")
                if wt_id is None:
                    continue
                try:
                    wt_id_int = int(wt_id)
                except (ValueError, TypeError):
                    continue
                if wt_id_int not in rows_por_wot:
                    continue
                dur_seg = it.get("duration")
                if dur_seg in (None, ""):
                    continue
                try:
                    dur_int = int(dur_seg)
                except (ValueError, TypeError):
                    continue
                for r in rows_por_wot[wt_id_int]:
                    r["Tarefa -> Duração estimada"] = _seg_para_timedelta(dur_int)
                duracoes_cache[wt_id_int] = dur_int
                enriquecidas += 1
            # Marca OS como processada SOMENTE se a chamada retornou sucesso
            wos_processadas.add(wo_id)
            erros_consecutivos = 0
        except Exception as e:
            erros += 1
            erros_consecutivos += 1
            if erros < 5:
                log(f"    erro OS {wo_id}: {str(e)[:60]}", "WARN")
            # Se 10 erros consecutivos, JWT provavelmente caiu — para e salva
            if erros_consecutivos >= 10:
                log(f"  ! 10 erros consecutivos. JWT pode ter expirado. "
                    f"Salvando checkpoint e parando.", "WARN")
                _salvar_checkpoint()
                log(f"  ! Re-rode com --so-duracoes pra retomar de onde parou.", "WARN")
                break
        if idx % 200 == 0:
            log(f"  -> {idx}/{len(pendentes)} OSs processadas nesta sessão "
                f"(corrigidas total: {enriquecidas})")
            _salvar_checkpoint()

    # Salva checkpoint final
    _salvar_checkpoint()
    log(f"  -> Total: {enriquecidas}/{len(rows_por_wot)} subtarefas corrigidas "
        f"(OSs com erro: {erros}, checkpoint: {CKPT_PATH})")


def enriquecer_com_execucoes(client, rows):
    """Adiciona Histórico Execução (JSON serializado da timeline completa) para
    OSs com status 1 (Em processo) ou 2 (Verificação).

    Usa JSON-RPC 2.0 via POST /rpc/proxy com method 'companies.transactions_log_wo_list'
    (descoberto via DevTools no app.fracttal.com).

    Mapeia id_wo_log_action → texto legível:
      1=OS criada, 2=Mudança de estado, 3=Tarefa iniciada, 7=Tarefa concluída,
      8=Nota atualizada, 16=Anexo/Form item
    """
    log("Coletando histórico de execução via JSON-RPC (status 1/2 apenas)...")

    # Verifica se JWT está configurado
    if not FRACTTAL_JWT_TOKEN:
        log("  ! FRACTTAL_JWT_TOKEN não configurado no .env.", "WARN")
        log("  ! Capture via DevTools→Application→Local Storage→id_token", "WARN")
        log("  ! Pulando histórico. Modal vai mostrar relatório, sem timeline.", "WARN")
        return rows
    # Decodifica payload do JWT pra checar expiração
    try:
        import base64
        partes = FRACTTAL_JWT_TOKEN.split(".")
        if len(partes) >= 2:
            payload_b64 = partes[1] + "=" * (4 - len(partes[1]) % 4)
            payload = json.loads(base64.urlsafe_b64decode(payload_b64))
            exp = payload.get("exp", 0)
            agora = int(time.time())
            horas_restantes = (exp - agora) / 3600
            email = payload.get("email", "?")
            if horas_restantes <= 0:
                log(f"  ! JWT EXPIRADO! Renove via DevTools. Email do token: {email}", "WARN")
                return rows
            elif horas_restantes < 1:
                log(f"  ! JWT expira em {horas_restantes:.1f}h. Renove em breve.", "WARN")
            else:
                log(f"  → JWT válido por mais {horas_restantes:.1f}h ({email})")
    except Exception as e:
        log(f"  ! Não consegui decodificar JWT: {e}. Tentando mesmo assim...", "WARN")

    # FILTRO INTELIGENTE: só OSs que aparecem no dashboard
    # 1) status Em Verificação (status=2): sempre incluir (precisa fechar)
    # 2) status Em processo (status=1) com date_maintenance recente (últimos 45 dias)
    from datetime import timezone
    corte = datetime.now(timezone.utc) - timedelta(days=45)

    def _eh_relevante(r):
        st = r.get("_id_status_work_order")
        if st not in (1, 2):
            return False
        if not r.get("_id_work_orders_tasks"):
            return False
        # Em Verificação → sempre
        if st == 2:
            return True
        # Em processo → só se date_maintenance >= 45 dias atrás
        dm = r.get("_date_maintenance") or ""
        if not dm:
            return False
        try:
            dm_dt = pd.to_datetime(dm, utc=True)
            return dm_dt >= corte
        except Exception:
            return False

    rows_alvo = [r for r in rows if _eh_relevante(r)]
    total_ativos = sum(1 for r in rows if r.get("_id_status_work_order") in (1, 2))
    log(f"  Filtro: {len(rows_alvo)} OSs relevantes (de {total_ativos} ativas total)")
    log(f"    → Em Verificação + Em processo dos últimos 45 dias")
    if not rows_alvo:
        return rows

    enriquecidas = 0
    sem_historico = 0
    falhas = 0

    # Mapeia id_wo_log_action → descrição
    LOG_ACTIONS = {
        1:  "OS criada",
        2:  "Mudança de estado",
        3:  "Tarefa iniciada",
        4:  "Tarefa pausada",
        5:  "Tarefa retomada",
        6:  "Cancelada",
        7:  "Tarefa concluída",
        8:  "Nota atualizada",
        9:  "Recurso adicionado",
        10: "Recurso removido",
        16: "Anexo / Form item",
    }

    # URL do JSON-RPC proxy (sniffed do app.fracttal.com)
    rpc_url_candidates = [
        "https://app.fracttal.com/rpc/proxy",
    ]

    import uuid

    def chamar_rpc(id_wo_task):
        """Faz 1 chamada JSON-RPC pra pegar timeline de execução de 1 TAREFA da OS.
        Retorna lista de execuções (com array actions IN_PROGRESS/PAUSED) ou [] se falhar.
        Endpoint descoberto: tasks.work_orders_tasks_execution_list, filter id_work_order_task."""
        body = [{
            "id": str(uuid.uuid4()),
            "jsonrpc": "2.0",
            "method": "tasks.work_orders_tasks_execution_list",
            "params": {
                "id_work_order_task": int(id_wo_task),
                "page": 1, "limit": 100,
            }
        }]
        for url in rpc_url_candidates:
            try:
                # /rpc/proxy exige JWT Bearer no header Authorization
                hdrs = {
                    "Content-Type": "application/json",
                    "Accept": "application/json, text/plain, */*",
                    "Authorization": f"Bearer {FRACTTAL_JWT_TOKEN}",
                    "Origin": "https://app.fracttal.com",
                    "Referer": "https://app.fracttal.com/tasks/wo",
                }
                r = client.session.post(url, headers=hdrs, json=body, timeout=20)
                if r.status_code != 200:
                    if id_wo_task == rows_alvo[0]["_id_work_orders_tasks"]:
                        log(f"    RPC {url} → HTTP {r.status_code}: {r.text[:150]}", "WARN")
                    continue
                # Parseia o JSON-RPC response
                try:
                    payload = r.json()
                except Exception:
                    continue
                if isinstance(payload, list) and payload:
                    resp = payload[0]
                else:
                    resp = payload
                if not isinstance(resp, dict):
                    return []
                result = resp.get("result", {})
                # Detecta JWT expirado pelo CAMPO message do JSON-RPC
                # (nao por string match no body inteiro - pode ter false-positive em descricoes)
                if isinstance(result, dict) and not result.get("success", True):
                    msg = result.get("message", "")
                    if msg in ("USER_NOT_LOGIN", "TOKEN_EXPIRED", "INVALID_TOKEN"):
                        if id_wo_task == rows_alvo[0]["_id_work_orders_tasks"]:
                            log(f"    RPC {url} → JWT EXPIRADO ({msg}). Renove no .env.", "WARN")
                        continue
                if isinstance(result, dict):
                    return result.get("data") or result.get("items") or []
                elif isinstance(result, list):
                    return result
                return []
            except Exception as e:
                if id_wo == rows_alvo[0]["_id_work_order"]:
                    log(f"    RPC {url} → erro: {e}", "WARN")
        return []

    # FASE 1: valida endpoint na 1ª OS
    id_teste_task = rows_alvo[0]["_id_work_orders_tasks"]
    log(f"  FASE 1: validando RPC com id_work_order_task={id_teste_task}...")
    eventos_teste = chamar_rpc(id_teste_task)
    if not eventos_teste:
        log("  ! RPC não retornou dados. JWT pode estar errado ou expirado.", "WARN")
        log("  ! Pulando histórico. Modal vai mostrar relatório, sem timeline.", "WARN")
        return rows
    log(f"  -> RPC OK! {len(eventos_teste)} execuções na tarefa de teste")

    # FASE 2: loop por todas as OSs ativas
    def _extrair_nome_usuario(account_log):
        """account_log vem como dict (ou string repr de dict) com chave 'name'."""
        if isinstance(account_log, dict):
            return (account_log.get("name") or "").strip()
        if isinstance(account_log, str):
            # Tenta parse com replace de aspas simples
            try:
                d = json.loads(account_log.replace("'", '"'))
                return (d.get("name") or "").strip()
            except Exception:
                # Regex simples pra 'name': 'X'
                import re
                m = re.search(r"'name'\s*:\s*'([^']+)'", account_log)
                if m:
                    return m.group(1).strip()
        return ""

    def _extrair_detalhe(ev):
        """Tenta produzir uma descrição curta para o evento."""
        try:
            data_action = ev.get("data_action")
            if isinstance(data_action, str):
                data_action = json.loads(data_action.replace("'", '"'))
            if not isinstance(data_action, dict):
                data_action = {}
        except Exception:
            data_action = {}
        action_id = ev.get("id_wo_log_action")
        # Ação 2 = Mudança de estado: pega action string (IN_REVIEW, DONE, etc)
        if action_id == 2:
            a = data_action.get("action") or ""
            mapa = {"IN_REVIEW": "Em Verificação", "IN_PROGRESS": "Em Progresso",
                    "DONE": "Concluída", "CANCELLED": "Cancelada",
                    "NO_STARTED": "Não Iniciada"}
            return f"→ {mapa.get(a, a)}"
        # Ação 7 = Tarefa concluída: duração real
        if action_id == 7:
            d = data_action.get("real_duration") or 0
            if d:
                h = int(d) // 3600
                m = (int(d) % 3600) // 60
                return f"Duração real: {h}h {m}min"
            return ""
        # Ação 8 = Nota
        if action_id == 8:
            note = ev.get("note") or data_action.get("note") or ""
            note = str(note).strip()
            if len(note) > 80:
                note = note[:77] + "..."
            return note
        # Ação 16 = Anexo/Form item
        if action_id == 16:
            files = data_action.get("files") or []
            forms = data_action.get("task_form_items") or []
            if files:
                return f"{len(files)} arquivo(s) anexado(s)"
            if forms:
                return "Form de tarefa atualizado"
        return ""

    total = len(rows_alvo)
    for idx, row in enumerate(rows_alvo):
        i_wot = row["_id_work_orders_tasks"]
        try:
            execucoes = chamar_rpc(i_wot)
            # Cada item de execucoes é uma execução da tarefa (geralmente 1 por OS).
            # Cada execução tem total_active_seconds (MTTR) e actions (timeline).
            mttr_total_s = 0
            tempo_total_s = 0
            historico = []
            ini_global = None
            fim_global = None
            for ex in execucoes:
                # Acumula MTTR (tempo ativo real)
                mttr_total_s += int(ex.get("total_active_seconds") or 0)
                # Pega initial_date / final_date
                ini = ex.get("initial_date")
                fim = ex.get("final_date")
                if ini and (not ini_global or ini < ini_global):
                    ini_global = ini
                if fim and (not fim_global or fim > fim_global):
                    fim_global = fim
                # Extrai actions (timeline IN_PROGRESS/PAUSED)
                acts = ex.get("actions") or []
                for a in acts:
                    historico.append({
                        "tipo": a.get("action_type_description") or "",
                        "ini": a.get("initial_date") or "",
                        "fim": a.get("final_date") or "",
                        "seg": int(a.get("duration_seconds") or 0),
                        "motivo": a.get("paused_reason_description") or "",
                        "usuario": ex.get("name") or "",
                    })
            # Calcula tempo total decorrido (initial_date global → final_date global)
            if ini_global and fim_global:
                try:
                    dt_i = pd.to_datetime(ini_global, utc=True)
                    dt_f = pd.to_datetime(fim_global, utc=True)
                    tempo_total_s = int((dt_f - dt_i).total_seconds())
                except Exception:
                    tempo_total_s = 0
            tempo_pausado_s = max(0, tempo_total_s - mttr_total_s)
            # Ordena historico cronologicamente
            historico.sort(key=lambda x: x.get("ini") or "")
            if historico or mttr_total_s > 0:
                row["Histórico Execução"] = json.dumps(historico, ensure_ascii=False)
                row["MTTR (s)"] = mttr_total_s
                row["Tempo Total (s)"] = tempo_total_s
                row["Tempo Pausado (s)"] = tempo_pausado_s
                row["MTTR (h)"] = round(mttr_total_s / 3600, 2)
                enriquecidas += 1
            else:
                sem_historico += 1
            # Throttling: ~10 req/s
            time.sleep(0.08)
        except Exception as e:
            falhas += 1
            if falhas < 3:
                log(f"    Erro buscando histórico de wot={i_wot}: {e}", "WARN")
        if (idx + 1) % 100 == 0:
            log(f"  -> {idx+1}/{total} processadas (com MTTR: {enriquecidas})")

    log(f"  -> Histórico: {enriquecidas} com timeline, "
        f"{sem_historico} sem dados, {falhas} falhas")
    return rows


def _enriquecer_com_execucoes_OLD(client, rows):
    """DEPRECATED: discovery REST puro - todos endpoints retornaram INVALID_ENDPOINT.
    Mantido apenas pra histórico/referência. Não chamar."""

    def fazer_filter_obj(id_wo):
        return [{"property": "id_work_order",
                 "operator": "eq",
                 "value": int(id_wo),
                 "condition": "and"}]

    def fazer_filter_str(id_wo):
        return json.dumps(fazer_filter_obj(id_wo))

    # IMPORTANTE: a doc REST do Fracttal usa padrão "wo_*" para endpoints relacionados
    # a tarefas dentro de OT (ex: /api/wo_resources/ pra recursos).
    # Mais provável: wo_executions, wo_tasks_execution, wo_tasks_status_log
    id_teste = rows_alvo[0]["_id_work_order"]
    candidatos = [
        # Padrão wo_* (mais provável - confirmado /api/wo_resources/ existe)
        ("GET wo_tasks_execution?filter=JSON",
         lambda b, i: (b + "wo_tasks_execution",
                       {"filter": fazer_filter_str(i), "limit": 100}, None)),
        ("GET wo_executions?filter=JSON",
         lambda b, i: (b + "wo_executions",
                       {"filter": fazer_filter_str(i), "limit": 100}, None)),
        ("GET wo_tasks_log?filter=JSON",
         lambda b, i: (b + "wo_tasks_log",
                       {"filter": fazer_filter_str(i), "limit": 100}, None)),
        ("GET wo_tasks_status?filter=JSON",
         lambda b, i: (b + "wo_tasks_status",
                       {"filter": fazer_filter_str(i), "limit": 100}, None)),
        ("GET wo_tasks_pause?filter=JSON",
         lambda b, i: (b + "wo_tasks_pause",
                       {"filter": fazer_filter_str(i), "limit": 100}, None)),
        ("GET wo_execution_log?filter=JSON",
         lambda b, i: (b + "wo_execution_log",
                       {"filter": fazer_filter_str(i), "limit": 100}, None)),
        # work_orders + sufixos diretos
        ("GET work_orders_execution?filter=JSON",
         lambda b, i: (b + "work_orders_execution",
                       {"filter": fazer_filter_str(i), "limit": 100}, None)),
        ("GET work_orders_executions?filter=JSON",
         lambda b, i: (b + "work_orders_executions",
                       {"filter": fazer_filter_str(i), "limit": 100}, None)),
        ("GET work_orders_pause_log?filter=JSON",
         lambda b, i: (b + "work_orders_pause_log",
                       {"filter": fazer_filter_str(i), "limit": 100}, None)),
        ("GET work_orders_tasks_status_changes?filter=JSON",
         lambda b, i: (b + "work_orders_tasks_status_changes",
                       {"filter": fazer_filter_str(i), "limit": 100}, None)),
        ("GET work_orders_tasks_log?filter=JSON",
         lambda b, i: (b + "work_orders_tasks_log",
                       {"filter": fazer_filter_str(i), "limit": 100}, None)),
        # Tentativa pelo padrão antigo (task singular) - já confirmado INVALID
        ("GET work_orders_task_execution?filter=JSON",
         lambda b, i: (b + "work_orders_task_execution",
                       {"filter": fazer_filter_str(i), "limit": 100}, None)),
    ]

    log(f"  FASE 1: descobrindo URL valida (id_wo={id_teste})...")
    fn_url = None
    formato_desc = ""
    for desc, builder in candidatos:
        url, params, body = builder(client.base_url, id_teste)
        metodo = "POST" if body is not None else "GET"
        try:
            time.sleep(0.2)
            if metodo == "POST":
                hdrs = dict(client._headers())
                hdrs["Content-Type"] = "application/json"
                r = client.session.post(url, headers=hdrs, json=body, timeout=15)
            else:
                r = client.session.get(url, headers=client._headers(),
                                       params=params, timeout=15)
            if r.status_code == 200:
                payload = r.json() if r.text else {}
                items = payload.get("data") or payload.get("items") or []
                if items:
                    fn_url = builder
                    formato_desc = desc
                    log(f"  -> URL valida: {desc} ({len(items)} execs)")
                    break
                else:
                    log(f"    {desc} -> 200 mas vazio", "WARN")
            elif r.status_code in (400, 404):
                # Body curto do erro pra diagnosticar (especialmente 400 = filtro errado)
                body_err = r.text[:200].replace("\n", " ")
                log(f"    {desc} -> HTTP {r.status_code}: {body_err}", "WARN")
            else:
                log(f"    {desc} -> HTTP {r.status_code}: {r.text[:100]}", "WARN")
        except Exception as e:
            log(f"    {desc} -> erro: {e}", "WARN")

    if fn_url is None:
        log("  ! Nenhuma URL respondeu 200. Pulando histórico.", "WARN")
        return rows

    total = len(rows_alvo)
    for idx, row in enumerate(rows_alvo):
        i_wo = row["_id_work_order"]
        try:
            url, params, body = fn_url(client.base_url, i_wo)
            time.sleep(0.04)
            if body is not None:
                hdrs = dict(client._headers())
                hdrs["Content-Type"] = "application/json"
                r = client.session.post(url, headers=hdrs, json=body, timeout=20)
            else:
                r = client.session.get(url, headers=client._headers(),
                                       params=params, timeout=20)
            if r.status_code != 200:
                falhas += 1
                continue
            payload = r.json()
            items = payload.get("data") or payload.get("items") or []
            # Cada item tem array "actions" com timeline
            historico = []
            for ex in items:
                acts = ex.get("actions") or []
                if isinstance(acts, str):
                    try:
                        acts = json.loads(acts.replace("'", '"'))
                    except Exception:
                        acts = []
                for a in acts:
                    historico.append({
                        "tipo": a.get("action_type_description") or "",
                        "ini": a.get("initial_date") or "",
                        "fim": a.get("final_date") or "",
                        "seg": int(a.get("duration_seconds") or 0),
                        "motivo": a.get("paused_reason_description") or "",
                    })
            # Ordena cronologicamente
            historico.sort(key=lambda x: x.get("ini") or "")
            if historico:
                row["Histórico Execução"] = json.dumps(historico, ensure_ascii=False)
                enriquecidas += 1
            else:
                sem_historico += 1
        except Exception as e:
            falhas += 1
            if falhas < 3:
                log(f"    Erro buscando histórico de wo={i_wo}: {e}", "WARN")
        if (idx + 1) % 200 == 0:
            log(f"  -> {idx+1}/{total} processadas (enriquecidas: {enriquecidas})")

    log(f"  -> Histórico: {enriquecidas} com timeline, "
        f"{sem_historico} sem dados, {falhas} falhas")
    return rows


# ============================================================
# Transformação Fracttal → BD
# ============================================================

STATUS_OS_MAP = {1: "Em processo", 2: "Verificação", 3: "Finalizados", 4: "Cancelado"}
TASK_STATUS_MAP = {
    "NO_STARTED": "Não Iniciada", "STARTED": "Em progresso",
    "PAUSED": "pausado", "DONE": "Finalizados",
}
CRITICIDADE_MAP = {"LOW": "Baixo", "MEDIUM": "Médio",
                   "HIGH": "Alto", "VERY_HIGH": "Muito alto"}
TIPO_ATIVO_MAP = {1: "Instalações", 2: "Equipamento",
                  3: "Peça de reposição", 4: "Ferramenta"}

# Override de "Ativo Classificação 2" (Equipe_Cluster) por nome do ativo.
# Usado quando o Fracttal está com groups_2_description vazio ou errado.
# Match é case-insensitive e ignora espaços extras.
OVERRIDE_CLASSIFICACAO_2 = {
    "nova xavantina 1 - mt":       "MT Leste 01",
    "nova xavantina 2 - mt":       "MT Leste 01",
    "são josé do egito 1 - pe":    "PE Oeste 01",
    "sao jose do egito 1 - pe":    "PE Oeste 01",
    "são bento 5 - rj":            "RJ Sul 01",
    "sao bento 5 - rj":            "RJ Sul 01",
    "rodrigues 1 - rn":            "RN Sul 01",
    "aparecida 3 - sp":            "SP Leste 01",
    "guaratinguetá 5 - sp":        "SP Leste 01",
    "guaratingueta 5 - sp":        "SP Leste 01",
    "santo inácio 12 - sp":        "SP Leste 04",
    "santo inacio 12 - sp":        "SP Leste 04",
    "salto pirapora 3 - sp":       "SP Leste 02",
    "barretos 1 - sp":             "SP Norte 02",
    "santa bárbara 1 - sp":        "SP Leste 05",
    "santa barbara 1 - sp":        "SP Leste 05",
    "araçoiaba da serra 1 - sp":   "SP Leste 02",
    "aracoiaba da serra 1 - sp":   "SP Leste 02",
    "araçoiaba da serra 2 - sp":   "SP Leste 02",
    "aracoiaba da serra 2 - sp":   "SP Leste 02",
    "elias fausto 1 - sp":         "SP Leste 07",
    "demerval lobão":              "PI Oeste 01",
    "demerval lobao":              "PI Oeste 01",
    "2c - ipixuna do pará 1":     "PA Norte 03",
    "2c - ipixuna do para 1":     "PA Norte 03",
    "axis - linhares 1":          "ES Norte 02",
    "renogrid - elias fausto":    "SP Leste 07",
    "thopen - alto paraná 1":     "PR Oeste 01",
    "thopen - alto parana 1":     "PR Oeste 01",
    "thopen - alto paraná 2":     "PR Oeste 01",
    "thopen - alto parana 2":     "PR Oeste 01",
    "thopen - ipixuna 1 e 2":     "PA Norte 03",
    "thopen - guaratinguetá":              "SP Leste 01",
    "thopen - guaratingueta":              "SP Leste 01",
    # Variantes com prefixo de cliente (formato completo do BD)
    "renogrid - nova xavantina 1 - mt":    "MT Leste 01",
    "renogrid - nova xavantina 2 - mt":    "MT Leste 01",
    "axis - são josé do egito 1 - pe":     "PE Oeste 01",
    "axis - sao jose do egito 1 - pe":     "PE Oeste 01",
    "thopen - são bento 5 - rj":           "RJ Sul 01",
    "thopen - sao bento 5 - rj":           "RJ Sul 01",
    "thopen - rodrigues 1 - rn":           "RN Sul 01",
    "thopen - aparecida 3 - sp":           "SP Leste 01",
    "thopen - guaratinguetá 5 - sp":       "SP Leste 01",
    "thopen - guaratingueta 5 - sp":       "SP Leste 01",
    "thopen - santo inácio 12 - sp":       "SP Leste 04",
    "thopen - santo inacio 12 - sp":       "SP Leste 04",
    "thopen - salto pirapora 3 - sp":      "SP Leste 02",
    "thopen - barretos 1 - sp":            "SP Norte 02",
    "thopen - santa bárbara 1 - sp":       "SP Leste 05",
    "thopen - santa barbara 1 - sp":       "SP Leste 05",
    "thopen - araçoiaba da serra 1 - sp":  "SP Leste 02",
    "thopen - aracoiaba da serra 1 - sp":  "SP Leste 02",
    "thopen - araçoiaba da serra 2 - sp":  "SP Leste 02",
    "thopen - aracoiaba da serra 2 - sp":  "SP Leste 02",
    "renogrid - elias fausto 1 - sp":      "SP Leste 07",
    "greenyellow - demerval lobão 1 - pi": "PI Oeste 01",
    "greenyellow - demerval lobao 1 - pi": "PI Oeste 01",
    "2c - ipixuna do pará 1 - pa":         "PA Norte 03",
    "2c - ipixuna do para 1 - pa":         "PA Norte 03",
    "axis - linhares 1 - es":              "ES Norte 02",
    "thopen - alto paraná 1 - pr":         "PR Oeste 01",
    "thopen - alto parana 1 - pr":         "PR Oeste 01",
    "thopen - alto paraná 2 - pr":         "PR Oeste 01",
    "thopen - alto parana 2 - pr":         "PR Oeste 01",
    "thopen - ipixuna 1 e 2 - pa":         "PA Norte 03",
}


def _aplicar_override_cls2(ativo: str, valor_atual: str) -> str:
    """Se o ativo está no OVERRIDE_CLASSIFICACAO_2, devolve o cluster correto.
    Senão devolve o valor_atual."""
    if not ativo:
        return valor_atual
    chave = " ".join(str(ativo).strip().lower().split())  # normaliza espaços
    return OVERRIDE_CLASSIFICACAO_2.get(chave, valor_atual)

COLS_BD_SEMANAL = [
    "OSs ID", "Status", "Código", "Ativo", "Fora de serviço", "Dependências",
    "Tarefa", "Ativador", "Localização ou parte de", "Tipo de Ativo",
    "Centro de custo", "Ativo Classificação 1", "Ativo Classificação 2",
    "Estado da Tarefa", "Tarefa -> Duração estimada", "Tempo de execução",
    "Tipo de tarefa", "Iterações", "Data Calculada", "Data Programada",
    "Data inicial", "Data final", "Data de Criação da OS",
    "Data de finalização da OS", "Parou o equipamento?",
    "Tempo de inatividade planejado do ativo", "Total de Horas",
    "Recursos humanos", "Peças de reposição e suprimentos",
    "Serviços Utilizados", "Responsável", "Ordem de Serviço Pai",
    "Criado por", "Tarefa -> Criticidade", "Tarefa -> Severidade das Falhas",
    "Tarefa -> Observação", "Plano de tarefas", "Tarefa -> Classificação 1",
    "Tarefa -> Classificação 2", "Ordem de Serviço -> Observação",
    "Data do Incidente", "Qualificação de OSs", "Número de Solicitação",
    "Motivo do cancelamento", "Tipo de registro de tempo", "OS relacionada",
    "Etiquetas",
    "% Concluída",
    "Descrição subtarefa", "Valor preenchido",
    "Relatório Execução", "Histórico Execução",
    # MTTR real (descontando pausas) — coletado via JSON-RPC tasks.work_orders_tasks_execution_list
    "MTTR (s)",              # Tempo ativo real em segundos (= total_active_seconds)
    "Tempo Total (s)",       # Tempo decorrido total (inicial → final) incluindo pausas
    "Tempo Pausado (s)",     # Tempo Total - MTTR (= duração total das pausas)
    "MTTR (h)",              # MTTR em horas com 2 decimais (pra relatórios)
]


def _seg_para_timedelta(s):
    if s is None or s == "":
        return None
    try:
        return timedelta(seconds=int(float(s)))
    except (ValueError, TypeError):
        return None


def _data_iso_para_naive(s):
    """ISO 8601 → datetime sem timezone (formato que o pandas grava no Excel)."""
    if not s:
        return None
    try:
        dt = pd.to_datetime(s, utc=True)
        # Converte pra horário de São Paulo (UTC-3) e remove tz
        return dt.tz_convert("America/Sao_Paulo").tz_localize(None)
    except Exception:
        return None


def _json_str(v):
    if v is None or v == "" or v == []:
        return ""
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return str(v)


def preservar_colunas_extras_do_bd_anterior(rows: List[Dict[str, Any]],
                                              preservar_duracao: bool = True):
    """Quando rodando com --sem-execucoes, lê o BD existente e copia
    as colunas pesadas (MTTR, Histórico, Relatório, Subtarefa) das OSs
    que já existiam no BD antigo. Evita que o cron rápido apague o
    trabalho do cron diário."""
    if not os.path.exists(OUTPUT_BD):
        log("  Sem BD anterior pra preservar — primeira rodada?")
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(OUTPUT_BD, read_only=True, data_only=True)
        if "Semanal" not in wb.sheetnames:
            return
        ws = wb["Semanal"]
        rows_iter = ws.iter_rows(values_only=True)
        header = list(next(rows_iter))
        # Colunas a preservar (índices no header antigo)
        COLS_PRESERVAR = [
            "Descrição subtarefa", "Valor preenchido",
            "Relatório Execução", "Histórico Execução",
            "MTTR (s)", "Tempo Total (s)", "Tempo Pausado (s)", "MTTR (h)",
        ]
        # Quando preservar_duracao=True (--sem-execucoes), também restaura
        # duração corrigida do BD anterior. Em --so-duracoes, NÃO (acabamos de calcular).
        if preservar_duracao:
            COLS_PRESERVAR.append("Tarefa -> Duração estimada")
        # Estas SOBRESCREVEM o valor novo (mesmo se preenchido).
        COLS_PRESERVAR_OVERRIDE = (
            {"Tarefa -> Duração estimada"} if preservar_duracao else set()
        )
        idx_os = header.index("OSs ID") if "OSs ID" in header else None
        idx_cols = {c: (header.index(c) if c in header else None) for c in COLS_PRESERVAR}
        if idx_os is None:
            log("  BD antigo sem coluna OSs ID — pulando preservação", "WARN")
            return
        # Mapa {os_id: {col: valor}} do BD antigo
        mapa_antigo = {}
        for r in rows_iter:
            try:
                os_id = r[idx_os]
                if os_id in (None, "", 0):
                    continue
                vals = {}
                for c, ix in idx_cols.items():
                    if ix is not None and ix < len(r):
                        v = r[ix]
                        if v not in (None, ""):
                            vals[c] = v
                if vals:
                    mapa_antigo[str(os_id)] = vals
            except Exception:
                continue
        wb.close()
        log(f"  BD anterior: {len(mapa_antigo)} OSs com dados de MTTR/histórico/relatório")
        # Copia pra rows novas
        preservadas = 0
        for r in rows:
            os_id = str(r.get("OSs ID", ""))
            vals_antigos = mapa_antigo.get(os_id)
            if not vals_antigos:
                continue
            for c, v in vals_antigos.items():
                # Override: sobrescreve mesmo se valor novo preenchido (caso da
                # duração estimada, que vem bugada da view 'work_orders').
                # Demais colunas: só preserva se valor novo está vazio.
                if c in COLS_PRESERVAR_OVERRIDE or not r.get(c):
                    r[c] = v
            preservadas += 1
        log(f"  → {preservadas} OSs com colunas extras preservadas do BD anterior")
    except Exception as e:
        log(f"  Falha ao preservar colunas: {e}", "WARN")


def transformar_wo_para_bd(row: Dict[str, Any]) -> Dict[str, Any]:
    """Pega 1 row da view tasks.v_works_orders_list_new e devolve dict no formato BD."""
    ativo_nome = row.get("items_log_description") or ""
    cls2_fonte = row.get("groups_2_description") or ""
    cls2_final = _aplicar_override_cls2(ativo_nome, cls2_fonte)
    return {
        "OSs ID": row.get("wo_folio") or row.get("id_work_order"),  # wo_folio = numeracao curta antiga
        "Status": STATUS_OS_MAP.get(row.get("id_status_work_order"), ""),
        "Código": row.get("code") or "",
        "Ativo": ativo_nome,
        "Fora de serviço": "Sim" if row.get("out_of_service_sec") else "Não",
        "Dependências": _json_str(row.get("dependencies")),
        "Tarefa": row.get("description") or "",
        "Ativador": row.get("trigger_description") or "",
        "Localização ou parte de": row.get("parent_description") or "",
        "Tipo de Ativo": TIPO_ATIVO_MAP.get(row.get("id_type_item"), ""),
        "Centro de custo": row.get("costs_center_description") or "",
        "Ativo Classificação 1": row.get("groups_1_description") or "",
        "Ativo Classificação 2": cls2_final,  # OVERRIDE_CLASSIFICACAO_2 aplicado se houver
        "Estado da Tarefa": TASK_STATUS_MAP.get(row.get("task_status"), row.get("task_status") or ""),
        "Tarefa -> Duração estimada": _seg_para_timedelta(row.get("duration")),
        "Tempo de execução": _seg_para_timedelta(row.get("real_duration")),
        "Tipo de tarefa": row.get("tasks_log_task_type_main") or "",
        "Iterações": row.get("num_iterations") or "",
        "Data Calculada": _data_iso_para_naive(row.get("cal_date_maintenance")),
        "Data Programada": _data_iso_para_naive(row.get("date_maintenance")),
        "Data inicial": _data_iso_para_naive(row.get("initial_date")),
        "Data final": _data_iso_para_naive(row.get("final_date")),
        "Data de Criação da OS": _data_iso_para_naive(row.get("creation_date")),
        "Data de finalização da OS": _data_iso_para_naive(row.get("wo_final_date")),
        "Parou o equipamento?": "Sim" if row.get("stop_assets") else "Não",
        "Tempo de inatividade planejado do ativo": _seg_para_timedelta(row.get("stop_assets_sec")),
        "Total de Horas": _seg_para_timedelta(row.get("tasks_duration")),
        "Recursos humanos": "",                # vazio (decisão do PCM)
        "Peças de reposição e suprimentos": "",
        "Serviços Utilizados": "",
        "Responsável": row.get("user_assigned") or "",
        "Ordem de Serviço Pai": row.get("wo_folio_parent") or "",
        "Criado por": row.get("created_by") or "",
        "Tarefa -> Criticidade": CRITICIDADE_MAP.get(row.get("priorities_description"),
                                                      row.get("priorities_description") or ""),
        "Tarefa -> Severidade das Falhas": row.get("severiry_description") or "",
        "Tarefa -> Observação": row.get("task_note") or "",
        "Plano de tarefas": row.get("group_task_description") or "",
        "Tarefa -> Classificação 1": row.get("tasks_log_types_description") or "",
        "Tarefa -> Classificação 2": row.get("tasks_log_types_2_description") or "",
        "Ordem de Serviço -> Observação": row.get("note") or "",
        "Data do Incidente": _data_iso_para_naive(row.get("event_date")),
        "Qualificação de OSs": row.get("rating") or "",
        "Número de Solicitação": row.get("id_request") or "",
        "Motivo do cancelamento": _json_str(row.get("reschedule_causes")),
        "Tipo de registro de tempo": row.get("execution_type") or "",
        "OS relacionada": row.get("code_wo_related") or _json_str(row.get("related_items")),
        "Etiquetas": _json_str(row.get("labels")),
        "% Concluída": row.get("completed_percentage") or row.get("percentage_completed") or 0,
        "Descrição subtarefa": "",
        "Valor preenchido": "",
        "Relatório Execução": "",
        "Histórico Execução": "",
        "MTTR (s)": "",
        "Tempo Total (s)": "",
        "Tempo Pausado (s)": "",
        "MTTR (h)": "",
        "_id_work_order": row.get("id_work_order"),
        "_id_work_orders_tasks": row.get("id_work_orders_tasks") or row.get("id_task"),
        "_id_status_work_order": row.get("id_status_work_order"),
        "_date_maintenance": row.get("date_maintenance") or row.get("cal_date_maintenance"),
    }


# ============================================================
# Coleta principal
# ============================================================

def coletar_todas_ots(client: FracttalClient, max_items=None,
                     com_subtarefas=True, com_execucoes=True,
                     so_duracoes=False) -> pd.DataFrame:
    log("Coletando OTs do Fracttal (paginação)...")
    rows = []
    total = 0
    for it in client.paginar("work_orders", page_size=100, max_items=max_items):
        rows.append(transformar_wo_para_bd(it))
        total += 1
        if total % 500 == 0:
            log(f"  -> {total} OTs processadas...")
    log(f"  -> Total: {total} OTs (brutas)")
    # FIX #167: Dedup defensivo — bug intermitente da API Fracttal pode fazer
    # a paginação repetir a mesma página, gerando 180x cada OS. Aqui garantimos
    # que cada (id_work_orders_tasks, ou wo_folio+code+description fallback)
    # apareça só 1x no BD final.
    seen = set()
    rows_dedup = []
    for r in rows:
        wt_id = r.get("_id_work_orders_tasks")
        if wt_id:
            key = ("wt", wt_id)
        else:
            # Fallback: combo de campos identificadores
            key = ("combo", r.get("OSs ID"), r.get("Código"), r.get("Tarefa"))
        if key in seen:
            continue
        seen.add(key)
        rows_dedup.append(r)
    if len(rows_dedup) != len(rows):
        log(f"  ! ALERTA: {len(rows) - len(rows_dedup)} linhas duplicadas removidas "
            f"({len(rows)} → {len(rows_dedup)}). "
            f"Provável bug de paginação intermitente da API.", "WARN")
    rows = rows_dedup
    # Detecta paginação travada: se muitas OSs aparecem N vezes EM SEQUÊNCIA, é red flag
    folios_unicos = len(set(r.get("OSs ID") for r in rows if r.get("OSs ID")))
    if total > 200 and folios_unicos < total / 5:
        log(f"  ! AVISO: {total} OTs processadas mas só {folios_unicos} folios únicos. "
            f"API pode ter ficado em loop.", "WARN")
    if so_duracoes:
        # Modo --so-duracoes: pula subtarefas/execuções, só corrige durações,
        # preserva MTTR/Histórico do BD anterior.
        log("Modo --so-duracoes: rodando só durações + preservando MTTR do BD anterior...")
        enriquecer_com_duracoes_subtarefas(client, rows)
        # Preserva todas colunas extras MENOS Duração (que acabou de ser calculada)
        preservar_colunas_extras_do_bd_anterior(rows, preservar_duracao=False)
    elif com_subtarefas:
        enriquecer_com_subtarefas(client, rows)
        if com_execucoes:
            # FIX 2026-06-15: durações POR subtarefa (não da OS pai).
            enriquecer_com_duracoes_subtarefas(client, rows)
            enriquecer_com_execucoes(client, rows)
        else:
            log("Modo --sem-execucoes: preservando MTTR/Histórico/Relatório/Duração do BD anterior...")
            preservar_colunas_extras_do_bd_anterior(rows)
    elif com_execucoes:
        # FIX 2026-06-15: durações POR subtarefa (não da OS pai). Só no modo full.
        enriquecer_com_duracoes_subtarefas(client, rows)
        enriquecer_com_execucoes(client, rows)
    else:
        # Modo --sem-execucoes: preserva colunas MTTR/Histórico/Duração do BD anterior
        log("Modo --sem-execucoes: preservando MTTR/Histórico/Relatório/Duração do BD anterior...")
        preservar_colunas_extras_do_bd_anterior(rows)
    for r in rows:
        r.pop("_id_work_order", None)
        r.pop("_id_work_orders_tasks", None)
        r.pop("_id_status_work_order", None)
        r.pop("_date_maintenance", None)
    return pd.DataFrame(rows, columns=COLS_BD_SEMANAL)


def carregar_aba_auxiliar() -> Optional[pd.DataFrame]:
    if not AUXILIAR_SOURCE_PATH:
        log("AUXILIAR_SOURCE_PATH não configurado — aba AUXILIAR ficará vazia.", "WARN")
        return None
    if not os.path.exists(AUXILIAR_SOURCE_PATH):
        log(f"Arquivo AUXILIAR não encontrado: {AUXILIAR_SOURCE_PATH}", "ERROR")
        return None
    log(f"Lendo AUXILIAR de {os.path.basename(AUXILIAR_SOURCE_PATH)}...")
    sheet_pref = os.environ.get("AUXILIAR_SHEET", "").strip()
    candidatos = []
    if sheet_pref:
        candidatos.append(sheet_pref)
    candidatos.extend(["Operações", "Operacoes", "OPERAÇÃO", "AUXILIAR", "Auxiliar", 0])
    df = None
    aba_usada = None
    last_err = None
    for cand in candidatos:
        try:
            df = pd.read_excel(AUXILIAR_SOURCE_PATH, sheet_name=cand)
            aba_usada = cand
            break
        except (ValueError, KeyError) as e:
            last_err = e
            continue
        except Exception as e:
            last_err = e
            break
    if df is None:
        log(f"Falha ao ler AUXILIAR: {last_err}", "ERROR")
        return None
    renomeacoes = {}
    for col in df.columns:
        c_clean = str(col).strip()
        if c_clean.upper() in ("OPERAÇÃO", "OPERACAO"):
            renomeacoes[col] = "UFV"
            break
    if renomeacoes:
        df = df.rename(columns=renomeacoes)
        log(f"  → Renomeado '{list(renomeacoes.keys())[0]}' → 'UFV'")
    log(f"  → Aba: '{aba_usada}', {len(df)} linhas, {len(df.columns)} colunas")
    return df


def backup_arquivo(path):
    if not os.path.exists(path):
        return
    os.makedirs(OUTPUT_BD_BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.basename(path)
    nm, ext = os.path.splitext(base)
    dest = os.path.join(OUTPUT_BD_BACKUP_DIR, f"{nm}_{ts}{ext}")
    try:
        import shutil
        shutil.copy2(path, dest)
        log(f"Backup: {os.path.basename(dest)}")
        bkps = sorted(pathlib.Path(OUTPUT_BD_BACKUP_DIR).glob(f"{nm}_*{ext}"),
                      key=lambda p: p.stat().st_mtime, reverse=True)
        for b in bkps[10:]:
            try:
                b.unlink()
            except OSError:
                pass
    except Exception as e:
        log(f"Backup falhou (não crítico): {e}", "WARN")


def gravar_excel(df_semanal: pd.DataFrame, df_auxiliar: Optional[pd.DataFrame]):
    backup_arquivo(OUTPUT_BD)
    log(f"Gravando {os.path.basename(OUTPUT_BD)}...")
    with pd.ExcelWriter(OUTPUT_BD, engine="openpyxl") as writer:
        df_semanal.to_excel(writer, sheet_name="Semanal", index=False)
        if df_auxiliar is not None:
            df_auxiliar.to_excel(writer, sheet_name="AUXILIAR", index=False)
    log(f"  → {len(df_semanal)} linhas em Semanal" +
        (f" + {len(df_auxiliar)} em AUXILIAR" if df_auxiliar is not None else ""))


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--test", action="store_true")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--no-auxiliar", action="store_true")
    parser.add_argument("--com-subtarefas", action="store_true")
    parser.add_argument("--sem-execucoes", action="store_true",
                        help="Pula coleta de MTTR/Timeline.")
    parser.add_argument("--so-duracoes", action="store_true",
                        help="Só roda enriquecer_com_duracoes_subtarefas. "
                             "Preserva MTTR/Histórico do BD anterior. "
                             "Use após logoff/JWT cair — retoma do checkpoint.")
    args = parser.parse_args()

    log("=" * 70)
    log("GERAR BD via API Fracttal")
    log("=" * 70)
    log(f"Base URL:  {BASE_URL}")
    log(f"Output:    {OUTPUT_BD}")
    log(f"AUXILIAR:  {AUXILIAR_SOURCE_PATH or '(nao configurado)'}")
    log(f"JWT:       {'configurado (' + str(len(FRACTTAL_JWT_TOKEN)) + ' chars)' if FRACTTAL_JWT_TOKEN else 'NAO configurado'}")
    log("")

    try:
        client = FracttalClient(CLIENT_ID, CLIENT_SECRET, BASE_URL)
        client.autenticar()
    except Exception as e:
        log(f"FALHA na autenticacao: {e}", "ERROR")
        sys.exit(1)

    if args.test:
        log("[--test] Autenticacao OK. Encerrando.")
        sys.exit(0)

    try:
        df_semanal = coletar_todas_ots(client, max_items=args.limit,
                                         com_subtarefas=args.com_subtarefas,
                                         com_execucoes=not args.sem_execucoes,
                                         so_duracoes=args.so_duracoes)
    except Exception as e:
        import traceback
        log(f"FALHA ao coletar OTs: {e}", "ERROR")
        log(traceback.format_exc(), "ERROR")
        sys.exit(2)

    df_auxiliar = None
    if not args.no_auxiliar:
        df_auxiliar = carregar_aba_auxiliar()

    try:
        gravar_excel(df_semanal, df_auxiliar)
    except Exception as e:
        log(f"FALHA ao gravar Excel: {e}", "ERROR")
        sys.exit(3)

    log("=" * 70)
    log("CONCLUIDO")
    log("=" * 70)


if __name__ == "__main__":
    main()
