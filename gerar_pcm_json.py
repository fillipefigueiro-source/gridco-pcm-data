# -*- coding: utf-8 -*-
"""
gerar_pcm_json.py
-----------------
Lê Programação Semana XX.xlsx (semana corrente) e gera banco_dados.json
no formato esperado pelo index.html.

Substitui o Office Script Acumulador_PCM + Power Automate. 100% local.

Saída: banco_dados.json no mesmo diretório, formato:
{
  "geradoEm": "...",
  "semanas": [
    {
      "week": "2026-W24",
      "num": 24,
      "label": "Semana 24 · 08–12 Jun 2026",
      "dates": {"seg": "08/06", "ter": "09/06", "qua": "10/06", "qui": "11/06", "sex": "12/06"},
      "resumo": { "<cluster>": {"hh_util": X, "hh_disp": 44, "pend": Y} },
      "rows": [
        { "cliente": "...", "tipo": "...", "usina": "...", "cluster": "...",
          "dia": "...", "os_id": "...", "responsavel": "...", "criticidade": "...",
          "status_bd": "...", "duracao": X, "h_ini": "...", "h_fim": "...",
          "reprog": "...", "nova_os": "...", "vezes": X, "status": "..." }
      ]
    }
  ]
}

Uso:
   py -3 gerar_pcm_json.py
   py -3 gerar_pcm_json.py --semana 24
"""
from __future__ import annotations
import os, sys, json, re, argparse, datetime as dt
from pathlib import Path
from collections import defaultdict

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import openpyxl

HH_DISP = 44
MESES_PT = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
DIAS_KEY = {"Segunda-feira":"seg","Terça-feira":"ter","Quarta-feira":"qua",
            "Quinta-feira":"qui","Sexta-feira":"sex","Sábado":"sab","Domingo":"dom"}


def _log(msg):
    print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


# ─── Grafia canônica de cluster ──────────────────────────────────────────────
# No Fracttal, cada cluster tem DUAS entradas na "Ativo Classificação 2" — uma
# para instalação (tipo 1) e outra para equipamento (tipo 2). Em três delas a
# segunda foi digitada em caixa alta, e o efeito era partir a equipe em duas:
# capacidade dividida ao meio (PA Leste 01 aparecia como 8,4h E 46,5h de 44h) e
# 297 tarefas sem destinatário de alerta, porque o alertas_destinatarios.json só
# conhece a grafia normal.
#
# Decisão de 21/08/2026: o cadastro do Fracttal NÃO será corrigido (a API só lê;
# a interface exigiria três renomeações manuais). Então este mapa é a REGRA, não
# um remendo — e por isso normaliza EM SILÊNCIO. Gritar a cada rodada por uma
# correção que ninguém vai fazer só ensina a ignorar o log.
#
# O que grita é grafia NOVA: se aparecer um par divergente fora deste mapa, o
# _conferir_clusters() denuncia e já entrega a linha a colar aqui.
_CLUSTER_CANONICO = {
    "PA LESTE 01": "PA Leste 01",
    "PA NORTE 01": "PA Norte 01",
    "MA LESTE 02": "MA Leste 02",
}
_CLUSTER_VISTO = {}          # UPPER -> {grafias vistas}
_CLUSTER_CORRIGIDO = {}      # grafia bruta -> quantas vezes


def _cluster_norm(bruto):
    """Grafia canônica do cluster. Silencioso para os casos conhecidos."""
    c = str(bruto or "").strip()
    if not c:
        return c
    _CLUSTER_VISTO.setdefault(c.upper(), set()).add(c)
    alvo = _CLUSTER_CANONICO.get(c.upper())
    if alvo and c != alvo:
        _CLUSTER_CORRIGIDO[c] = _CLUSTER_CORRIGIDO.get(c, 0) + 1
        return alvo
    return c


def _conferir_clusters():
    """Resumo do que foi normalizado; alarme só para grafia nova.

    Sem a segunda parte, o mapa resolveria os três de hoje e o próximo apareceria
    em silêncio — que é exatamente como o problema chegou até aqui.
    """
    if _CLUSTER_CORRIGIDO:
        det = ", ".join(f"{k} x{v}" for k, v in sorted(_CLUSTER_CORRIGIDO.items()))
        _log(f"cluster: grafia canônica aplicada em "
             f"{sum(_CLUSTER_CORRIGIDO.values())} linhas ({det})")

    novos = [(k, sorted(g)) for k, g in sorted(_CLUSTER_VISTO.items())
             if len(g) > 1 and k not in _CLUSTER_CANONICO]
    for _chave, grafias in novos:
        # Escolhe como canônica a que segue "UF Regiao NN" com a região em Título.
        # Um [A-Za-z]+ aqui aceitava "SUL" tão bem quanto "Sul" e, como sorted()
        # põe maiúscula primeiro, a sugestão saía inútil ("CE SUL 01" -> ela mesma).
        pref = next((g for g in grafias
                     if re.match(r"^[A-Z]{2} [A-ZÀ-Ú][a-zà-ú]+ \d{2}$", g)),
                    grafias[0])
        _log(f"ATENÇÃO: cluster com grafias divergentes NÃO tratadas: {grafias} "
             f"— a equipe está sendo contada em dobro e pode ficar sem alerta.")
        for g in grafias:
            if g != pref:
                _log(f'         acrescente em _CLUSTER_CANONICO:  '
                     f'"{g.upper()}": "{pref}",')


def _detectar_pasta() -> Path:
    env = os.getenv("PCM_PROG_DIR")
    if env and Path(env).exists():
        return Path(env)
    here = Path(__file__).resolve().parent
    if (here / "atualizacao_semanal.py").exists():
        return here
    for c in [
        Path.home() / "GRID CO" / "Grid Co. - Gridco" / "4. O&M"
            / "11.Pré-Operação" / "6. PCM" / "09. Programação Semanal",
        Path.home() / "OneDrive - Grid Co" / "4. O&M"
            / "11.Pré-Operação" / "6. PCM" / "09. Programação Semanal",
    ]:
        if c.exists():
            return c
    raise FileNotFoundError("Pasta PCM não encontrada.")


def _semana_iso(d: dt.date):
    return d.isocalendar()


def _label_semana(mon: dt.date) -> str:
    fim = mon + dt.timedelta(days=4)
    _, num, _ = mon.isocalendar()
    if mon.month == fim.month:
        return f"Semana {num} · {mon.day:02d}–{fim.day:02d} {MESES_PT[mon.month-1]} {mon.year}"
    return f"Semana {num} · {mon.day:02d} {MESES_PT[mon.month-1]} – {fim.day:02d} {MESES_PT[fim.month-1]} {mon.year}"


def _dates_dict(mon: dt.date) -> dict:
    out = {}
    for i, key in enumerate(["seg","ter","qua","qui","sex"]):
        d = mon + dt.timedelta(days=i)
        out[key] = d.strftime("%d/%m")
    return out


def _extrair_cliente(ativo: str) -> str:
    if not ativo:
        return ""
    s = str(ativo).strip()
    # Rejeita nomes de equipamentos (contém chaves ou sem separador " - ")
    if "{" in s or "}" in s or " - " not in s:
        return ""
    # Padrão "Cliente - Usina - UF ..."
    parts = s.split(" - ", 1)
    if not parts:
        return ""
    return parts[0].strip()


def _safe_float(v, default=0.0):
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        try:
            return float(str(v).replace(",", "."))
        except (ValueError, TypeError):
            return default


def _safe_int(v, default=0):
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


def _norm_dia(v):
    """Normaliza 'Segunda-feira (08/06)' → 'Segunda-feira'."""
    if not v:
        return ""
    s = str(v).strip()
    if "(" in s:
        s = s.split("(")[0].strip()
    return s


def _hora_to_str(v):
    if v is None or v == "":
        return ""
    if isinstance(v, dt.time):
        return v.strftime("%H:%M")
    if isinstance(v, dt.datetime):
        return v.strftime("%H:%M")
    if isinstance(v, (int, float)):
        # fração de dia (Excel)
        if 0 <= v < 1:
            total_min = int(round(v * 24 * 60))
            return f"{total_min//60:02d}:{total_min%60:02d}"
        return str(v)
    s = str(v).strip()
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return s


def _norm_tar(s: str) -> str:
    """Normaliza texto de tarefa p/ casar Excel↔API (minúsculo, espaços colapsados, – vira -)."""
    s = str(s or "").lower().replace("–", "-").replace("—", "-")
    return re.sub(r"\s+", " ", s).strip()


_RANK_EST = {"não iniciada": 0, "nao iniciada": 0, "": 0, "pausado": 1, "pausada": 1,
             "em progresso": 2, "em andamento": 2, "iniciada": 2,
             "finalizada": 3, "finalizados": 3, "finalizado": 3, "concluída": 3, "concluida": 3}


def _rank_est(e: str) -> int:
    """Ordena estados p/ resolver duplicatas de tarefa (mais avançado vence)."""
    return _RANK_EST.get(str(e or "").strip().lower(), 0)


def _carregar_mapa_extras(pasta: Path) -> dict:
    """Monta {os_id: {pct, statusPai, datas, relatorio, historico, mttr...}} DIRETO da
    API Fracttal (fonte_bd_api.df_semanal), SEM ler o BD_Relatório Semanal.xlsx.
    MOTIVO: a leitura do BD via openpyxl travava no PC da automação (OneDrive
    cloud-only / arquivo meio-escrito) e congelava o banco_dados.json (desde 02/07).
    Campos ENRIQUECIDOS (historico/mttr/relatório detalhado) vêm vazios no fluxo base
    da API — só o daily full run os preenchia; o relatório cai no fallback da Observação."""
    out = {}
    try:
        import fonte_bd_api
        df = fonte_bd_api.df_semanal()   # cacheado (o atualizacao_semanal acabou de popular)
    except Exception as e:
        _log(f"  ! Falha ao carregar mapa extras da API: {e}")
        return out

    # AUXILIAR p/ resolver Responsável O&M por usina ATUAL (mesma lógica da Gestão PCM)
    try:
        import gerar_gestao_pcm_json as _gp
        _uidx, _cmaj = _gp.carregar_auxiliar(_gp.AUXILIAR_PATH)
    except Exception:
        _gp, _uidx, _cmaj = None, {}, {}

    def _nn(v):   # None p/ NaN/NaT do pandas
        if v is None:
            return None
        try:
            if v != v:
                return None
        except Exception:
            pass
        return v

    def _date_iso(v):
        v = _nn(v)
        if v in (None, ""):
            return ""
        try:
            return v.isoformat() if hasattr(v, "isoformat") else str(v)
        except Exception:
            return str(v)

    for r in df.to_dict("records"):
        try:
            os_id = int(_nn(r.get("OSs ID")))
        except (ValueError, TypeError):
            continue
        entrada = out.setdefault(os_id, {"pct": 0.0, "relatorio": "", "historico": [],
                                          "mttr_s": 0, "mttr_h": 0.0,
                                          "tempo_total_s": 0, "pausado_s": 0,
                                          "dataCriacao": "", "dataFinal": "", "dataProgramada": "",
                                          "statusPai": ""})
        sp = _nn(r.get("Status"))
        if sp:
            entrada["statusPai"] = str(sp).strip()
        # ITEM 1 (status VIVO): estado da tarefa por (os_id, tarefa), direto da API.
        # Duplicatas de mesma descrição: mantém o estado MAIS avançado (não sub-declara).
        _tar_t = _nn(r.get("Tarefa"))
        if _tar_t:
            _d_est = entrada.setdefault("estadoTarefa", {})
            _k = _norm_tar(str(_tar_t))
            _novo = str(_nn(r.get("Estado da Tarefa")) or "").strip()
            if _k not in _d_est or _rank_est(_novo) >= _rank_est(_d_est[_k]):
                _d_est[_k] = _novo
        # Classificação ATUAL do ativo + responsável fresco (via API + AUXILIAR)
        _ua = _nn(r.get("Ativo Classificação 1"))
        _ca = _nn(r.get("Ativo Classificação 2"))
        if _ua:
            entrada["usinaAtual"] = str(_ua).strip()
        if _ca:
            entrada["clusterAtual"] = _cluster_norm(_ca)
        if _gp is not None and _ua:
            try:
                entrada["responsavelAtual"] = _gp.resolver_responsavel(
                    str(_ua), str(_ca or ""), _uidx, _cmaj)
            except Exception:
                pass
        # Filtro "Responsável" do painel (13/08/2026): a PESSOA dona da OS no
        # Fracttal (coluna "Responsável" da API). Não confundir com o
        # "Responsável O&M" acima, que vem do AUXILIAR e é o contato da usina.
        _ro = _nn(r.get("Responsável"))
        if _ro and str(_ro).strip():
            entrada["respOs"] = " ".join(str(_ro).split())
        entrada["dataCriacao"]    = _date_iso(r.get("Data de Criação da OS"))
        entrada["dataFinal"]      = _date_iso(r.get("Data de finalização da OS"))
        entrada["dataProgramada"] = _date_iso(r.get("Data Programada"))
        pct = _nn(r.get("% Concluída"))
        if pct not in (None, ""):
            entrada["pct"] = _safe_float(pct)
        rel = _nn(r.get("Relatório Execução"))
        if rel and str(rel).strip():
            entrada["relatorio"] = str(rel).strip()
        if not entrada["relatorio"]:
            obs = _nn(r.get("Ordem de Serviço -> Observação"))   # Task #147: fallback
            if obs and str(obs).strip():
                entrada["relatorio"] = str(obs).strip()
        hist_raw = _nn(r.get("Histórico Execução"))
        if hist_raw and str(hist_raw).strip():
            try:
                entrada["historico"] = json.loads(str(hist_raw))
            except (ValueError, TypeError):
                pass
        for k, c in (("mttr_s", "MTTR (s)"), ("tempo_total_s", "Tempo Total (s)"),
                     ("pausado_s", "Tempo Pausado (s)")):
            v = _nn(r.get(c))
            if v not in (None, ""):
                entrada[k] = int(_safe_float(v))
        v = _nn(r.get("MTTR (h)"))
        if v not in (None, ""):
            entrada["mttr_h"] = _safe_float(v)
    return out


def _carregar_mapa_concluida(pasta: Path) -> dict:
    """Compat: devolve só {os_id: pct} pra código antigo. Usa _carregar_mapa_extras."""
    return {k: v.get("pct", 0.0) for k, v in _carregar_mapa_extras(pasta).items()}


def ler_planilha(xlsx_path: Path, mon: dt.date, mapa_extras: dict = None) -> dict:
    if not xlsx_path.exists():
        _log(f"ERRO: {xlsx_path} não existe.")
        return None
    _log(f"Lendo {xlsx_path.name}...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    rows = []
    resumo = defaultdict(lambda: {"hh_util": 0.0, "hh_disp": HH_DISP, "pend": 0})
    pendentes_total = 0
    pendentes = []   # linhas da aba _Pendentes (não couberam) → mostradas no dashboard
    qualidade = []   # linhas da aba _Qualidade (portão 0.7) → aviso no dashboard

    # FAST: usa iter_rows (values_only) que é ordens de magnitude mais rápido
    # que ws.cell() linha por linha em arquivo grande.
    sheet_names = wb.sheetnames
    total_sheets = len([s for s in sheet_names if not s.startswith("_")])
    _log(f"  → {total_sheets} abas de equipe a processar...")

    for sheet_idx, sheet_name in enumerate(sheet_names):
        if sheet_name.startswith("_"):
            # Aba especial: _Pendentes alimenta o resumo
            if sheet_name == "_Pendentes":
                ws_p = wb[sheet_name]
                rows_p = list(ws_p.iter_rows(values_only=True))
                if len(rows_p) >= 2:
                    header_p = list(rows_p[0])
                    def _gp(n): return header_p.index(n) if n in header_p else None
                    ie, io_, ia = _gp("Equipe"), _gp("OSs ID"), _gp("Ativo")
                    it, itp, idur, imot = _gp("Tarefa"), _gp("Tipo"), _gp("Duração (h)"), _gp("Motivo")
                    for r in rows_p[1:]:
                        def _v(i): return (r[i] if (i is not None and i < len(r)) else None)
                        eq = _v(ie)
                        eq = eq.strip() if (isinstance(eq, str) and eq.strip()) else "(sem equipe)"
                        resumo[eq]["pend"] += 1
                        pendentes_total += 1
                        if io_ is None or _v(io_) is None:
                            continue
                        _usina = str(_v(ia) or "").strip()
                        pendentes.append({
                            "os_id":   str(_v(io_)).split(".")[0].strip(),
                            "usina":   _usina,
                            "cliente": _extrair_cliente(_usina) if _usina else "",
                            "cluster": _cluster_norm(eq),
                            "tarefa":  str(_v(it) or "").strip(),
                            "tipo":    str(_v(itp) or "").strip(),
                            "duracao": _safe_float(_v(idur)),
                            "motivo":  str(_v(imot) or "").strip(),
                        })
            # Aba especial: _Qualidade (portão 0.7) → aviso no dashboard
            elif sheet_name == "_Qualidade":
                rows_q = list(wb[sheet_name].iter_rows(values_only=True))
                for r in rows_q[1:]:
                    tipo = str(r[0] or "").strip()
                    if tipo in ("REMOVIDA", "AVISO"):
                        qualidade.append({
                            "tipo": tipo,
                            "item": str(r[1] or "").strip(),
                            "detalhe": str(r[2] or "").strip(),
                            "acao": str(r[3] or "").strip() if len(r) > 3 else "",
                        })
            continue

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 2:
            continue
        header = list(all_rows[0])
        if "OSs ID" not in header:
            continue
        col = {h: i for i, h in enumerate(header) if h}

        def G(name):
            return col.get(name)

        i_os       = G("OSs ID")
        i_equipe   = G("Equipe")
        i_dia      = G("Dia")
        i_ativo    = G("Ativo (Usina)") if G("Ativo (Usina)") is not None else G("Ativo")
        i_codigo   = G("Código Equipamento")
        i_tipo     = G("Tipo")
        i_tarefa   = G("Tarefa")
        i_rpn      = G("RPN/Prioridade")
        i_etiq     = G("Etiquetas")
        i_reprog   = G("Reprogramada")
        i_vezes    = G("Nº vezes programada")
        i_termo    = G("Termografia")
        i_hi       = G("Hora Início")
        i_hf       = G("Hora Fim")
        i_dur      = G("Duração (h)")
        i_desloc   = G("Desloc (h)")
        i_paralelo = G("Paralelo (terceirizada)")
        i_status   = G("Status Atual (BD)")
        i_resp     = G("Responsável")
        i_nova     = G("Nova OS")
        i_solic    = G("Nº Solicitação Origem")
        i_rolagem  = G("Rolagem")            # M1: badge ↻ no painel

        def get(r, idx):
            if idx is None or idx >= len(r):
                return None
            return r[idx]

        n_aba = 0
        for r in all_rows[1:]:
            v_os = get(r, i_os)
            if v_os in (None, "", 0):
                continue
            estado = get(r, i_status) or ""
            if isinstance(estado, str) and "cancelad" in estado.lower():
                continue
            # Lookup do extras (pct, relatório, histórico, statusPai)
            extras_os = None
            try:
                os_id_int = int(get(r, i_os))
                if mapa_extras:
                    extras_os = mapa_extras.get(os_id_int)
            except (ValueError, TypeError):
                os_id_int = None
            # FIX #141: filtra OSs cuja OS pai está Cancelada (mesmo que tarefa não esteja)
            status_pai_raw = (extras_os or {}).get("statusPai", "")
            if "cancelad" in status_pai_raw.lower():
                continue
            # ITEM 1: badge = ESTADO DA TAREFA VIVO da API (casado por tarefa);
            # fallback p/ a coluna "Status Atual (BD)" do Excel só se a API não tiver.
            _live_est = None
            if extras_os:
                _live_est = (extras_os.get("estadoTarefa") or {}).get(_norm_tar(str(get(r, i_tarefa) or "")))
            estado_badge = _live_est if (_live_est not in (None, "")) else str(estado).strip()
            # NOTA: o badge principal reflete o ESTADO DA TAREFA puro (Não Iniciada /
            # Em progresso / Pausada / Finalizada). O Status da OS (Em processo /
            # Verificação / Finalizados) vai num SELO SEPARADO no frontend (statusPai).
            # (removida a heurística que trocava para "Em Verificação" por %/Status)
            ativo = get(r, i_ativo) or ""
            cliente = _extrair_cliente(ativo)
            equipe = get(r, i_equipe) or sheet_name
            dur = _safe_float(get(r, i_dur))
            hi  = _hora_to_str(get(r, i_hi))
            hf  = _hora_to_str(get(r, i_hf))
            _ex = extras_os or {}
            _usina_at = _ex.get("usinaAtual")
            row = {
                "cliente":     (_extrair_cliente(_usina_at) if _usina_at else cliente),
                "usina":       (_usina_at or str(ativo).strip()),
                "cluster":     _cluster_norm(_ex.get("clusterAtual") or equipe),
                "tipo":        str(get(r, i_tipo) or "").strip(),
                "dia":         _norm_dia(get(r, i_dia)),
                "os_id":       str(v_os).strip(),
                "codigo":      str(get(r, i_codigo) or "").strip(),
                "tarefa":      str(get(r, i_tarefa) or ""),
                "responsavel": (_ex.get("responsavelAtual") or str(get(r, i_resp) or "").strip()),
                "resp_os": str(_ex.get("respOs") or ""),
                "criticidade": str(get(r, i_rpn) or ""),
                "etiquetas":   str(get(r, i_etiq) or ""),
                "status_bd":   estado_badge,
                "status":      estado_badge,
                "duracao":     dur,
                "h_ini":       hi,
                "h_fim":       hf,
                "desloc":      _safe_float(get(r, i_desloc)),
                "reprog":      str(get(r, i_reprog) or ""),
                "vezes":       _safe_int(get(r, i_vezes), 1),
                "termo":       str(get(r, i_termo) or ""),
                "paralelo":    str(get(r, i_paralelo) or ""),
                "nova_os":     str(get(r, i_nova) or ""),
                "solic_orig":  get(r, i_solic) or "",
                "rolagem":     str(get(r, i_rolagem) or ""),

                "relatorio":   (extras_os or {}).get("relatorio", ""),
                "historico":   (extras_os or {}).get("historico", []),
                "mttr_s":      (extras_os or {}).get("mttr_s", 0),
                "mttr_h":      (extras_os or {}).get("mttr_h", 0.0),
                "tempo_total_s": (extras_os or {}).get("tempo_total_s", 0),
                "pausado_s":   (extras_os or {}).get("pausado_s", 0),
                "dataCriacao": (extras_os or {}).get("dataCriacao", ""),
                "dataFinal":   (extras_os or {}).get("dataFinal", ""),
                "dataProgramada": (extras_os or {}).get("dataProgramada", ""),
                "statusPai":   (extras_os or {}).get("statusPai", ""),
            }
            rows.append(row)
            resumo[row["cluster"]]["hh_util"] += dur
            n_aba += 1
        if n_aba > 0:
            _log(f"    [{sheet_idx+1}/{len(sheet_names)}] {sheet_name}: {n_aba} OSs")

    # Limpa o defaultdict pra dict normal
    resumo_final = {}
    for k, v in resumo.items():
        resumo_final[k] = {
            "hh_util": round(v["hh_util"], 1),
            "hh_disp": v["hh_disp"],
            "pend":    v["pend"],
        }

    _log(f"  → {len(rows)} linhas / {len(resumo_final)} equipes / {pendentes_total} pendentes")
    return {
        "week":   f"{mon.year}-W{mon.isocalendar()[1]:02d}",
        "num":    mon.isocalendar()[1],
        "label":  _label_semana(mon),
        "dates":  _dates_dict(mon),
        "resumo": resumo_final,
        "rows":   rows,
        "pendentes": pendentes,
        "qualidade": qualidade,
    }


_DIA_FULL = ['Segunda-feira', 'Terça-feira', 'Quarta-feira', 'Quinta-feira', 'Sexta-feira']


def _iso_dt(v):
    try:
        if v is None or (v != v):
            return ""
        return v.isoformat() if hasattr(v, "isoformat") else str(v)
    except Exception:
        return str(v) if v else ""


def linhas_fora_do_plano(df, mon: dt.date, plan_keys: set) -> list:
    """Tarefas EXECUTADAS (Data final) na semana [seg..sex] que NÃO estavam no plano.
    Placa cada uma no dia em que foi concluída. Retorna linhas no mesmo schema do plano,
    com foraDoPlano=True. Casa por (os_id, código) para não colidir com tarefas planejadas."""
    import pandas as pd
    d0, d1 = pd.Timestamp(mon), pd.Timestamp(mon) + pd.Timedelta(days=5)  # seg 00h → sáb 00h
    dfin = pd.to_datetime(df.get('Data final'), errors='coerce')
    di = pd.to_datetime(df.get('Data inicial'), errors='coerce')
    out = []
    for i, r in enumerate(df.to_dict("records")):
        if str(r.get('Status') or '').strip() == 'Cancelado':
            continue
        fim = dfin.iat[i]
        if pd.isna(fim) or not (d0 <= fim < d1):        # executada nesta semana útil
            continue
        try:
            os_id = str(int(float(str(r.get('OSs ID')).strip())))
        except (ValueError, TypeError):
            continue
        cod = str(r.get('Código') or '').strip()
        if (os_id, cod) in plan_keys or os_id in plan_keys:   # já está no plano
            continue
        usina = str(r.get('Ativo Classificação 1') or '').strip()
        if not usina or usina.lower() == 'nan' or 'teste' in usina.lower():
            continue
        ini = di.iat[i]
        # duração real de execução (cap 24h: spans multi-dia são tempo de OS aberta, não trabalho)
        dur = round((fim - ini).total_seconds() / 3600.0, 2) if (pd.notna(ini) and fim >= ini) else 0.0
        if dur > 24:
            dur = 0.0
        estado = str(r.get('Estado da Tarefa') or '').strip()
        out.append({
            "cliente": _extrair_cliente(usina), "usina": usina,
            "cluster": _cluster_norm(r.get('Ativo Classificação 2')),
            "tipo": str(r.get('Tipo de tarefa') or '').strip(),
            "dia": _DIA_FULL[fim.weekday()], "os_id": os_id, "codigo": cod,
            "tarefa": str(r.get('Tarefa') or ''), "responsavel": "",
            "resp_os": " ".join(str(r.get('Responsável') or '').split()),
            "criticidade": str(r.get('Tarefa -> Criticidade') or ''),
            "etiquetas": str(r.get('Etiquetas') or ''),
            "status_bd": estado, "status": estado,
            "duracao": dur,
            "h_ini": ini.strftime('%H:%M') if pd.notna(ini) else '',
            "h_fim": fim.strftime('%H:%M') if pd.notna(fim) else '',
            "desloc": 0, "reprog": "", "vezes": 0, "termo": "", "paralelo": "",
            "nova_os": "", "solic_orig": "", "relatorio": "", "historico": [],
            "mttr_s": 0, "mttr_h": 0.0, "tempo_total_s": 0, "pausado_s": 0,
            "dataCriacao": _iso_dt(r.get('Data de Criação da OS')),
            "dataFinal": _iso_dt(r.get('Data final')),
            "dataProgramada": _iso_dt(r.get('Data Programada')),
            "statusPai": str(r.get('Status') or '').strip(),
            "foraDoPlano": True,
        })
    return out


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--semana", type=int, help="semana ISO ativa (default: semana corrente)")
    p.add_argument("--semanas", type=int, default=4, help="quantas semanas incluir no JSON (default: 4)")
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    pasta = _detectar_pasta()
    _log(f"Pasta: {pasta}")

    # Determina a semana "ativa" (padrão do seletor no dashboard)
    ano_hoje = dt.date.today().year
    if args.semana:
        mon_ativa = dt.date.fromisocalendar(ano_hoje, args.semana, 1)
    else:
        hoje = dt.date.today()
        mon_ativa = hoje - dt.timedelta(days=hoje.weekday())
    _, iso_w_ativa, _ = mon_ativa.isocalendar()
    semana_ativa_key = f"{mon_ativa.year}-W{iso_w_ativa:02d}"

    # Descobre todos os arquivos "Programação Semana XX.xlsx" disponíveis
    disponiveis = []
    for arq in pasta.glob("Programação Semana ??.xlsx"):
        try:
            n = int(arq.stem.rsplit(None, 1)[-1])
            disponiveis.append((n, arq))
        except (ValueError, IndexError):
            continue
    disponiveis.sort(key=lambda x: x[0], reverse=True)  # mais recente primeiro

    # Seleciona semanas a incluir: ativa + as mais recentes até completar o limite
    nums_incluir = {iso_w_ativa}
    for n, _ in disponiveis:
        if len(nums_incluir) >= args.semanas:
            break
        nums_incluir.add(n)

    mapa_extras = _carregar_mapa_extras(pasta)
    n_rel = sum(1 for v in mapa_extras.values() if v.get("relatorio"))
    n_hist = sum(1 for v in mapa_extras.values() if v.get("historico"))
    _log(f"  → Mapa Extras: {len(mapa_extras)} OSs / {n_rel} c/relatório / {n_hist} c/histórico")

    # API (cacheada — mapa_extras já chamou) p/ derivar as executadas fora do plano
    try:
        import fonte_bd_api
        _df_api = fonte_bd_api.df_semanal()
    except Exception as e:
        _df_api = None
        _log(f"  ! Fora-do-plano desativado (sem API): {e}")

    semanas_json = []
    for n, arq in sorted(disponiveis, key=lambda x: x[0], reverse=True):
        if n not in nums_incluir:
            continue
        mon = dt.date.fromisocalendar(ano_hoje, n, 1)
        semana = ler_planilha(arq, mon, mapa_extras)
        if semana:
            if _df_api is not None:
                plan_keys = set()
                for _r in semana["rows"]:
                    plan_keys.add((str(_r.get("os_id", "")), str(_r.get("codigo", ""))))
                    plan_keys.add(str(_r.get("os_id", "")))
                extra = linhas_fora_do_plano(_df_api, mon, plan_keys)
                semana["rows"].extend(extra)
                _log(f"    + {len(extra)} executadas FORA do plano ({semana['week']})")
            semanas_json.append(semana)

    if not semanas_json:
        _log("ERRO: não foi possível ler nenhuma planilha.")
        return 1

    def _ler_alertas(pasta_):
        try:
            p = pasta_ / "_alertas_do_dia.json"
            if p.exists():
                return json.loads(p.read_text(encoding="utf-8"))
        except Exception as e:
            _log(f"  ! alertas ilegíveis: {e}")
        return None

    def _ler_motivos(pasta_):
        """Linhas '# motivo OS 10371: Aguarda peça/material' das observações.
        São comentário para o parser de pins (texto livre), mas viram DADO aqui
        — alimentam o Pareto de causas do painel."""
        out = []
        pad = re.compile(r"^#\s*motivo\s+OS\s*#?\s*(\d+)\s*(?:\(([^)]*)\))?\s*:\s*(.+)$", re.I)
        for arq in ("Observacoes_Semana_Atual.txt", "Observacoes_Semana.txt"):
            p = pasta_ / arq
            if not p.exists():
                continue
            try:
                for ln in p.read_text(encoding="utf-8").splitlines():
                    m = pad.match(ln.strip())
                    if m:
                        out.append({"os": m.group(1), "equipe": (m.group(2) or "").strip(),
                                    "motivo": m.group(3).strip()[:80]})
            except Exception as e:
                _log(f"  ! motivos ilegíveis em {arq}: {e}")
        if out:
            _log(f"  → {len(out)} motivo(s) capturado(s) nas observações")
        return out

    payload = {
        "geradoEm":     dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "fonte":        f"gerar_pcm_json.py ({len(semanas_json)} semanas)",
        "semana_ativa": semana_ativa_key,
        "semanas":      semanas_json,
        # M3 — alertas do dia (STEP A5 grava o arquivo; o estado "janelas já
        # emitidas hoje" viaja aqui dentro, que é o artefato commitado)
        "alertas":      _ler_alertas(pasta),
        # M3 — motivos capturados ("# motivo OS X: Y" nas observações) → Pareto
        "motivos":      _ler_motivos(pasta),
    }
    _conferir_clusters()
    total_rows = sum(len(s.get("rows", [])) for s in semanas_json)
    semanas_str = ", ".join(s["week"] for s in semanas_json)
    if args.dry_run:
        out = json.dumps(payload, ensure_ascii=False, indent=2)
        print(out[:4000])
        _log(f"DRY-RUN: {len(out)} bytes | semanas: {semanas_str}")
        return 0
    out_path = pasta / "banco_dados.json"
    out_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    _log(f"OK: gravado {out_path.name} ({total_rows} linhas | semanas: {semanas_str})")
    _log(f"  → Semana ativa: {semana_ativa_key}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
