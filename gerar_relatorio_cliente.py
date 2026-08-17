# -*- coding: utf-8 -*-
"""
gerar_relatorio_cliente.py — Melhoria 2 (relatório semanal do cliente)
----------------------------------------------------------------------
Um e-mail por cliente, só com as usinas dele, gerado na sexta depois do
programacao_v7. Traz o FECHAMENTO da semana que termina e a PRÉVIA da semana
que começa, e convida o cliente a pedir ajustes — que deságuam na gramática de
observações que já existe (§11), sem mecanismo novo.

Ciclo definido com o PCM em 13/08/2026:
  · sexta ao MEIO-DIA  — PCM revisa a programação e dispara os .eml
  · segunda ao MEIO-DIA — prazo do cliente para pedir ajuste
  · terça em diante     — o pedido entra na programação
A segunda-feira não é ajustável: quando o retorno chega, a equipe já está em
campo com o dia fechado desde sexta. Por isso o e-mail promete "a partir de
terça" — prometer a própria segunda seria vender o que a operação não entrega.

Decisões do pacote de 08/08/2026:
  12 — contatos vêm de `contatos_relatorio.json`; cliente sem contato NÃO é
       enviado e aparece no aviso final
  13 — do não-feito: reprogramada mostra a nova data (vira compromisso);
       pendente sem data aparece UMA vez como "em replanejamento"
  15 — nada sai direto: o script gera .eml para revisão do PCM
  16 — corte automático de layout: > 8 usinas com atividade OU > 40 linhas -> B

Regras de linguagem (do mockup):
  · sem jargão interno — nada de RPN, tier, cluster, "Nº vezes programada"
  · período (manhã/tarde/noturno), nunca hora exata — a cascata muda horários
  · PROIBIDO afirmar impacto em geração (MWh): o PCM não tem esse dado

Uso:
    py -3 gerar_relatorio_cliente.py                 # semana ativa do JSON
    py -3 gerar_relatorio_cliente.py --semana 34
    py -3 gerar_relatorio_cliente.py --cliente Athon
"""
import argparse
import datetime as dt
import html
import json
import os
import re
import sys
from email.message import EmailMessage

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

AQUI = os.path.dirname(os.path.abspath(__file__))
CORTE_USINAS = int(os.environ.get("PCM_REL_CORTE_USINAS", 8))    # decisão 16
CORTE_LINHAS = int(os.environ.get("PCM_REL_CORTE_LINHAS", 40))
VERDE, ESCURO, CINZA = "#A2CA40", "#1C1F3B", "#666666"


def log(m):
    print(f"[{dt.datetime.now():%H:%M:%S}] {m}", flush=True)


def pasta_dados():
    p = os.environ.get("PCM_PROG_DIR", "").strip()
    if p and os.path.isdir(p):
        return p
    c = os.path.join(os.path.expanduser("~"), "GRID CO", "GRID CO. - Gridco",
                     "4. O&M", "11.Pré-Operação", "6. PCM", "09. Programação Semanal")
    return c if os.path.isdir(c) else AQUI


def esc(s):
    return html.escape(str(s or ""))


def periodo(row):
    """Período em vez de hora exata (decisão do mockup). Noturno é o único que
    ganha janela real, porque muda a rotina da usina.

    A flag [NOTURNO] existe na planilha mas NÃO chega ao banco_dados.json — por
    isso a MPA é reconhecida pelo tipo + horário na janela 15:30→01:30."""
    dia = str(row.get("dia") or "")
    hi = str(row.get("h_ini") or "")
    if str(row.get("paralelo") or "").lower() == "sim" or "ZELADORIA" in dia.upper():
        return "dia todo*"
    try:
        h, mn = (int(x) for x in hi.split(":")[:2])
    except (ValueError, IndexError):
        return "—"
    minutos = h * 60 + mn
    if "NOTURNO" in dia.upper() or (str(row.get("tipo")) == "MPA" and (minutos >= 930 or h < 3)):
        return "noturno 15:30→01:30"
    return "manhã" if h < 12 else ("tarde" if h < 18 else "noite")


TIPO_CLIENTE = {                      # jargão interno -> linguagem de cliente
    "Corretiva": ("CORR", "Corretiva"), "Corretiva Emergencial": ("CORR", "Corretiva"),
    "MPA": ("MPA", "Manutenção anual"), "MPS": ("PREV", "Manutenção semestral"),
    "MPM": ("PREV", "Manutenção mensal"), "MPM-Inversor": ("PREV", "Manutenção mensal — inversores"),
    "MPM-Mod/Tracker": ("PREV", "Manutenção mensal — módulos e trackers"),
    "MPT": ("PREV", "Manutenção trimestral"),
    "MPQ": ("PREV", "Manutenção quinzenal"), "Inspeção": ("INSP", "Inspeção"),
    "Preventiva": ("PREV", "Preventiva"), "Zeladoria": ("ZELAD.", "Zeladoria"),
    "Handover": ("PREV", "Passagem de turno"), "Administrativa": ("ADM", "Administrativa"),
}


def rotulo(tipo):
    return TIPO_CLIENTE.get(str(tipo), ("PREV", str(tipo or "Atividade")))


def nome_usina(r):
    """O campo vem como "Cliente - Usina - UF"; o cliente só quer a usina.

    O separador NÃO é confiável: o cadastro tem "Araçoiaba da Serra 1- SP" sem
    espaço antes do traço, então a UF é retirada por padrão e não por posição."""
    t = re.sub(r"\s*-\s*[A-Z]{2}\s*$", "", str(r.get("usina") or "").strip())
    partes = [x.strip() for x in t.split(" - ") if x.strip()]
    return partes[-1] if partes else ""


def escopo_tarefa(r):
    """Duas linhas da mesma OS não são blocos partidos: são subtarefas com
    escopos diferentes (MPA - Relé de Proteção vs MPA - String e Módulos).
    Sem isso a tabela mostra duas linhas idênticas e parece erro."""
    t = str(r.get("tarefa") or "")
    t = re.sub(r"^\s*\[[^\]]*\]\s*-\s*", "", t)          # tira "[Grid Co.] - "
    t = re.sub(r"^\s*(MPA|MPS|MPM|MPT|MPQ|MPW)\b\s*-?\s*", "", t, flags=re.I)
    t = t.strip(" -.")
    # o rótulo do tipo já diz "Corretiva"/"Manutenção X" — repetir vira ruído
    if len(t) < 3 or t.lower().startswith(("corretiva", "manuten")):
        return ""
    return re.sub(r"\s*-\s*corretiva\.?\s*$", "", t, flags=re.I)


def finalizada(r):
    e = str(r.get("status_bd") or r.get("status") or "").lower()
    return "finaliz" in e or "conclu" in e or "verifica" in e


def carregar_contatos():
    p = os.path.join(pasta_dados(), "contatos_relatorio.json")
    if not os.path.exists(p):
        return {}, p
    try:
        with open(p, encoding="utf-8") as f:
            d = json.load(f)
        return {str(k).strip().lower(): v for k, v in (d.get("contatos") or {}).items()}, p
    except Exception as e:
        log(f"! contatos_relatorio.json ilegível ({e})")
        return {}, p


ASSINATURA_NOMES = ("assinatura.png", "assinatura.jpg", "assinatura.jpeg", "assinatura.gif")


def assinatura_arquivo():
    """Caminho do cartão de assinatura, se o PCM tiver colocado na pasta."""
    base = pasta_dados()
    for n in ASSINATURA_NOMES:
        p = os.path.join(base, n)
        if os.path.exists(p):
            return p
    return None


def assinatura_html():
    """Assinatura NO FIM do e-mail (14/08/2026).

    O Outlook injeta a assinatura automática no TOPO do corpo que já vem pronto
    no .eml — por isso o cartão aparecia antes do relatório. Como o novo Outlook
    guarda a assinatura na nuvem (sem arquivo local para ler), a saída é o PCM
    desligar a assinatura automática e deixar o cartão como imagem na pasta:
    daí ele vai embutido aqui, no rodapé, onde deve ficar."""
    p = assinatura_arquivo()
    if not p:
        return ""
    return ('<div style="margin-top:22px">'
            '<img src="cid:assinaturagridco" alt="" '
            'style="display:block;max-width:520px;width:100%;height:auto;border:0"></div>')


def dedup_blocos(rows):
    """O motor parte tarefas longas (MPA) em blocos de 4 h — para o cliente isso
    é a MESMA visita. Colapsa por (OS, dia) e mantém o período mais cedo, salvo
    quando algum bloco é noturno (aí o noturno é o que muda a rotina da usina)."""
    saida, vistos = [], {}
    for r in sorted(rows, key=lambda x: str(x.get("h_ini") or "")):
        k = (str(r.get("os_id")), str(r.get("dia") or "").split(" [")[0],
             str(r.get("tarefa") or ""))
        if k in vistos:
            if "noturno" in periodo(r) and "noturno" not in periodo(vistos[k]):
                vistos[k].update({"h_ini": r.get("h_ini"), "tipo": r.get("tipo")})
            continue
        vistos[k] = dict(r)
        saida.append(vistos[k])
    return saida


def bloco_fechamento(rows_ant, semana_ant, rows_atual):
    """Fechamento da semana que termina.

    Decisão 13 — do que não foi feito, o cliente precisa saber o DESTINO:
    se a OS reaparece na semana nova, isso é um compromisso com data; se não
    reaparece, é honesto dizer "em replanejamento" em vez de omitir.

    O campo `reprog` do BD NÃO serve aqui: ele marca "já foi reprogramada
    alguma vez" (histórico da OS), não "reprogramada para esta semana" — usá-lo
    inflaria o número (8 contra 3 reais na Utragaz/S32)."""
    if not rows_ant:
        return ""
    # Tudo em OS, nunca misturando com subtarefa: contar o total por linha e o
    # pendente por OS produzia "445 tarefas · 89 reprogramadas · 13 em
    # replanejamento", onde as partes não somam com o todo e o cliente cobra.
    # Uma OS só conta como concluída quando TODAS as suas linhas fecharam.
    por_os = {}
    for r in rows_ant:
        por_os.setdefault(str(r.get("os_id")), []).append(r)
    tot = len(por_os)
    feitas = {o for o, rs in por_os.items() if all(finalizada(x) for x in rs)}
    fin = len(feitas)
    pct = round(100 * fin / tot) if tot else 0

    dias_novos = {}
    for r in rows_atual:
        dias_novos.setdefault(str(r.get("os_id")), str(r.get("dia") or "").split(" [")[0])
    pend = {o: rs[0] for o, rs in por_os.items() if o not in feitas}
    volta = [(r, dias_novos[o]) for o, r in pend.items() if o in dias_novos]
    replan = {o for o in pend if o not in dias_novos}

    det = []
    if volta:
        det.append(f'<b>{len(volta)}</b> reprogramada(s) para esta semana')
    if replan:
        det.append(f'<b>{len(replan)}</b> em replanejamento')
    lista = ""
    if volta and len(volta) <= 12:
        itens = "".join(
            f'<li style="margin:2px 0">OS <b>{esc(r.get("os_id"))}</b> &middot; '
            f'{esc(nome_usina(r))} &rarr; {esc(d)}</li>'
            for r, d in sorted(volta, key=lambda x: str(x[0].get("os_id"))))
        lista = (f'<ul style="margin:8px 0 0 16px;padding:0;font-size:12px;'
                 f'color:#5a5566;line-height:1.5">{itens}</ul>')

    # O que FOI concluído (pedido do PCM em 14/08). O bloco só mostrava o não
    # feito, o que dava ao cliente uma leitura injusta da semana. A forma muda
    # com o tamanho: listar as 139 OS da Thopen arruinaria o e-mail.
    feito = ""
    if feitas:
        # Sempre agrupado por usina, e o número é de TAREFAS, não de OS (pedido
        # do PCM em 14/08): uma OS pode ter 20 subtarefas, então contar OS
        # subdimensiona o que a equipe realmente entregou na semana.
        porusi = {}
        for o in feitas:
            for r in por_os[o]:
                u = nome_usina(r)
                porusi[u] = porusi.get(u, 0) + 1
        # Sem teto (14/08): o "e mais 40 usina(s)" escondia justamente as usinas
        # que o cliente quer ver reconhecidas. Em texto corrido, 50+ usinas ainda
        # cabem num parágrafo — é bem menos volume que listar OS a OS.
        top = sorted(porusi.items(), key=lambda x: (-x[1], x[0]))
        corpo = (f'<div style="margin:6px 0 0;font-size:12px;color:#5a5566;line-height:1.6">'
                 + " &middot; ".join(f'{esc(u)} <b>({n})</b>' for u, n in top)
                 + '</div>')
        feito = (f'<div style="margin-top:10px;padding-top:9px;border-top:1px solid #e4e4dc">'
                 f'<div style="font-size:11px;color:{CINZA};text-transform:uppercase;'
                 f'letter-spacing:.05em;font-weight:700">Conclu&iacute;das '
                 f'&middot; tarefas por usina</div>{corpo}</div>')
    return (
        f'<div style="background:#f5f5f2;border-radius:8px;padding:12px 14px;margin:0 0 18px">'
        f'<div style="font-size:11px;color:{CINZA};text-transform:uppercase;letter-spacing:.05em;'
        f'font-weight:700;margin-bottom:6px">Como foi a {esc(semana_ant)}</div>'
        f'<div style="font-size:14px;color:{ESCURO}">'
        f'<b>{fin} de {tot} ordens de servi&ccedil;o conclu&iacute;das ({pct}%)</b>'
        + (' &middot; ' + ' &middot; '.join(det) if det else '')
        + f'</div>{feito}'
        + (f'<div style="margin-top:10px;padding-top:9px;border-top:1px solid #e4e4dc">'
           f'<div style="font-size:11px;color:{CINZA};text-transform:uppercase;'
           f'letter-spacing:.05em;font-weight:700">Reprogramadas para esta semana</div>'
           f'{lista}</div>' if lista else '')
        + '</div>')


def layout_a(cli, rows, semana_lbl, periodo_lbl, fechamento, resp):
    """Detalhado: tabela OS a OS, agrupada por dia."""
    por_dia = {}
    for r in rows:
        d = str(r.get("dia") or "").split(" [")[0]
        por_dia.setdefault(d, []).append(r)
    ordem = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta"]
    def chave(d):
        for i, x in enumerate(ordem):
            if d.startswith(x):
                return i
        return 9
    linhas = []
    for d in sorted(por_dia, key=chave):
        linhas.append(
            f'<tr><td colspan="4" style="background:{ESCURO};color:#fff;padding:6px 10px;'
            f'font-size:12px;font-weight:700">{esc(d)}</td></tr>')
        for r in sorted(por_dia[d], key=lambda x: str(x.get("h_ini") or "")):
            tag, desc = rotulo(r.get("tipo"))
            linhas.append(
                f'<tr>'
                f'<td style="padding:6px 10px;font-size:12px;border-bottom:1px solid #eee"><b>{esc(r.get("os_id"))}</b></td>'
                f'<td style="padding:6px 10px;font-size:12px;border-bottom:1px solid #eee">'
                f'{esc(nome_usina(r))}</td>'
                f'<td style="padding:6px 10px;font-size:12px;border-bottom:1px solid #eee">{esc(desc)}'
                + (f' <span style="color:{CINZA}">&mdash; {esc(escopo_tarefa(r))}</span>'
                   if escopo_tarefa(r) else '')
                + (f' <span style="background:#eef2e0;color:#5E8C1A;font-size:9px;font-weight:700;'
                   f'padding:1px 5px;border-radius:3px">{tag}</span></td>')
                + f'<td style="padding:6px 10px;font-size:12px;border-bottom:1px solid #eee;'
                  f'color:{CINZA}">{periodo(r)}</td></tr>')
    tem_zel = any(str(r.get("paralelo") or "").lower() == "sim" for r in rows)
    tem_not = any("noturno" in periodo(r) for r in rows)
    notas = []
    if tem_zel:
        notas.append("* zeladoria executada por equipe terceirizada, em paralelo.")
    if tem_not:
        notas.append("A manuten&ccedil;&atilde;o anual em turno noturno ocorre com a usina operando normalmente.")
    return (fechamento
            + f'<div style="font-size:11px;color:{CINZA};text-transform:uppercase;letter-spacing:.05em;'
              f'font-weight:700;margin:0 0 8px">O que est&aacute; programado &mdash; {esc(semana_lbl)}</div>'
            + '<table style="width:100%;border-collapse:collapse;border:1px solid #e6e6ee">'
            + '<tr><th style="text-align:left;padding:7px 10px;font-size:10px;color:#8a8a94;'
              'text-transform:uppercase">OS</th>'
              '<th style="text-align:left;padding:7px 10px;font-size:10px;color:#8a8a94;'
              'text-transform:uppercase">Usina</th>'
              '<th style="text-align:left;padding:7px 10px;font-size:10px;color:#8a8a94;'
              'text-transform:uppercase">Atividade</th>'
              '<th style="text-align:left;padding:7px 10px;font-size:10px;color:#8a8a94;'
              'text-transform:uppercase">Per&iacute;odo</th></tr>'
            + "".join(linhas) + '</table>'
            + (f'<div style="font-size:11px;color:{CINZA};margin-top:8px">'
               + "<br>".join(notas) + '</div>' if notas else ''))


def layout_b(cli, rows, semana_lbl, periodo_lbl, fechamento, resp):
    """Portfólio: visão de frota + destaques; o detalhe vai no anexo."""
    usinas = {str(r.get("usina") or "") for r in rows}
    prev = sum(1 for r in rows if rotulo(r.get("tipo"))[0] == "PREV")
    corr = sum(1 for r in rows if rotulo(r.get("tipo"))[0] == "CORR")
    mpa = [r for r in rows if rotulo(r.get("tipo"))[0] == "MPA"]
    kpi = (f'<div style="display:block;margin:0 0 16px">'
           f'<span style="display:inline-block;background:#f5f5f2;border-radius:7px;padding:9px 14px;'
           f'margin:0 6px 6px 0"><b style="font-size:17px;color:{ESCURO}">{len(usinas)}</b>'
           f'<span style="font-size:11px;color:{CINZA}"> usinas</span></span>'
           f'<span style="display:inline-block;background:#f5f5f2;border-radius:7px;padding:9px 14px;'
           f'margin:0 6px 6px 0"><b style="font-size:17px;color:{ESCURO}">{len(rows)}</b>'
           f'<span style="font-size:11px;color:{CINZA}"> atividades</span></span>'
           f'<span style="display:inline-block;background:#f5f5f2;border-radius:7px;padding:9px 14px;'
           f'margin:0 6px 6px 0"><b style="font-size:17px;color:{ESCURO}">{prev}</b>'
           f'<span style="font-size:11px;color:{CINZA}"> preventivas</span></span>'
           f'<span style="display:inline-block;background:#f5f5f2;border-radius:7px;padding:9px 14px;'
           f'margin:0 6px 6px 0"><b style="font-size:17px;color:{ESCURO}">{corr}</b>'
           f'<span style="font-size:11px;color:{CINZA}"> corretivas</span></span>'
           f'<span style="display:inline-block;background:#f5f5f2;border-radius:7px;padding:9px 14px;'
           f'margin:0 6px 6px 0"><b style="font-size:17px;color:{ESCURO}">{len(mpa)}</b>'
           f'<span style="font-size:11px;color:{CINZA}"> anuais</span></span></div>')
    dest = []
    if mpa:
        us = sorted({str(r.get("usina") or "").split(" - ")[-2] for r in mpa
                     if " - " in str(r.get("usina") or "")})[:4]
        dest.append(f"{len(mpa)} manuten&ccedil;&atilde;o(&otilde;es) anual(is) em turno noturno: {esc(', '.join(us))}")
    if corr:
        dest.append(f"{corr} corretiva(s) programada(s) &mdash; detalhe no anexo")
    # Nada de contar `reprog` aqui: ele é histórico da OS ("já foi reprogramada
    # alguma vez"), e como quase toda OS antiga tem esse selo o destaque dizia
    # "334 de 364 reprogramadas da semana anterior" — falso e alarmante. O que
    # vale é o que o bloco de fechamento já calcula pelo reencontro real.
    zel = sum(1 for r in rows if str(r.get("paralelo") or "").lower() == "sim")
    if zel:
        dest.append(f"{zel} atividade(s) de zeladoria em paralelo, por equipe terceirizada")
    destaques = ('<div style="font-size:11px;color:' + CINZA + ';text-transform:uppercase;'
                 'letter-spacing:.05em;font-weight:700;margin:0 0 8px">Destaques</div><ul style="margin:0 0 16px;'
                 'padding-left:18px;font-size:13px;color:' + ESCURO + ';line-height:1.7">'
                 + "".join(f"<li>{d}</li>" for d in dest) + '</ul>') if dest else ""
    # mapa usina x dia (só as 12 mais movimentadas)
    ordem = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta"]
    cont = {}
    for r in rows:
        u = str(r.get("usina") or "")
        d = str(r.get("dia") or "")
        di = next((i for i, x in enumerate(ordem) if d.startswith(x)), None)
        if di is None:
            continue
        cont.setdefault(u, [0] * 5)[di] += 1
    top = sorted(cont.items(), key=lambda x: -sum(x[1]))[:12]
    linhas = []
    for u, v in top:
        cel = "".join(f'<td style="padding:5px 8px;text-align:center;font-size:12px;'
                      f'border-bottom:1px solid #eee">{n if n else "&middot;"}</td>' for n in v)
        nome = u.split(" - ")[-2] if " - " in u else u
        linhas.append(f'<tr><td style="padding:5px 8px;font-size:12px;border-bottom:1px solid #eee">'
                      f'{esc(nome)}</td>{cel}</tr>')
    mapa = ('<div style="font-size:11px;color:' + CINZA + ';text-transform:uppercase;letter-spacing:.05em;'
            'font-weight:700;margin:0 0 8px">Mapa da semana &mdash; atividades por dia</div>'
            '<table style="width:100%;border-collapse:collapse;border:1px solid #e6e6ee">'
            '<tr><th style="text-align:left;padding:6px 8px;font-size:10px;color:#8a8a94">Usina</th>'
            + "".join(f'<th style="padding:6px 8px;font-size:10px;color:#8a8a94">{d[:3]}</th>'
                      for d in ordem) + '</tr>' + "".join(linhas) + '</table>'
            + (f'<div style="font-size:11px;color:{CINZA};margin-top:6px">'
               f'&hellip; e mais {len(cont) - len(top)} usina(s). Detalhe completo, OS a OS, no anexo.</div>'
               if len(cont) > len(top) else ''))
    return fechamento + kpi + destaques + mapa


def montar_email(cli, rows, semana_lbl, semana_ant_lbl, rows_ant, resp, anexo_nome):
    rows = dedup_blocos(rows)
    grande = (len({str(r.get("usina") or "") for r in rows}) > CORTE_USINAS
              or len(rows) > CORTE_LINHAS)                      # decisão 16
    fech = bloco_fechamento(rows_ant, semana_ant_lbl, rows)
    corpo = (layout_b if grande else layout_a)(cli, rows, semana_lbl, "", fech, resp)
    ajuste = (
        f'<div style="background:#fffaf0;border-left:4px solid {VERDE};border-radius:6px;'
        f'padding:12px 14px;margin:18px 0 0">'
        f'<div style="font-size:13px;font-weight:700;color:{ESCURO};margin-bottom:4px">'
        f'Precisa ajustar algum ponto?</div>'
        f'<div style="font-size:12.5px;color:#5a5566;line-height:1.6">'
        f'Responda este e-mail citando a OS e o que precisa &mdash; por exemplo: '
        f'<i>OS 10390 &rarr; quinta</i> ou <i>OS 10415: adiar, a usina recebe visita na segunda</i>.<br>'
        f'Retorno at&eacute; <b>segunda ao meio-dia</b> entra na programa&ccedil;&atilde;o '
        f'<b>a partir de ter&ccedil;a</b>; depois disso, entra no ciclo de '
        f'atualiza&ccedil;&atilde;o do meio da semana.</div></div>')
    rodape = (f'<div style="margin-top:16px;padding-top:10px;border-top:1px solid #e6e6ee;'
              f'font-size:11px;color:{CINZA}">Respons&aacute;vel O&amp;M: {esc(resp)}'
              + (f'<br>Detalhe completo no anexo <b>{esc(anexo_nome)}</b>.' if grande and anexo_nome else '')
              + '</div>'
              + assinatura_html())
    html_ = (f'<div style="font-family:Segoe UI,Arial,sans-serif;max-width:760px">'
             f'<div style="background:{ESCURO};color:#fff;padding:16px 18px;border-radius:8px 8px 0 0">'
             f'<div style="font-size:11px;color:{VERDE};letter-spacing:.06em">GRID CO. &mdash; O&amp;M</div>'
             f'<div style="font-size:19px;font-weight:700;margin-top:2px">Programa&ccedil;&atilde;o Semanal &mdash; {esc(semana_lbl)}</div>'
             f'<div style="font-size:12px;color:#b9b9c6;margin-top:3px">Cliente {esc(cli)}</div></div>'
             f'<div style="border:1px solid #e6e6ee;border-top:none;border-radius:0 0 8px 8px;padding:18px">'
             f'{corpo}{ajuste}{rodape}</div></div>')
    return html_, grande


def anexo_xlsx(cli, rows, destino, rows_ant=None):
    """Layout B: uma linha por OS, com abas separadas (decisão 13).

    Passa pelo mesmo dedup do corpo, senão o anexo mostra 460 linhas contra as
    364 do KPI e o cliente pergunta qual das duas está certa."""
    import openpyxl
    rows = dedup_blocos(rows)
    wb = openpyxl.Workbook()
    def escreve(ws, dados, titulo):
        ws.append(["OS", "Usina", "Atividade", "Tipo", "Dia", "Período"])
        for c in ws[1]:
            c.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            c.fill = openpyxl.styles.PatternFill("solid", fgColor="1C1F3B")
        for r in dados:
            tag, desc = rotulo(r.get("tipo"))
            ws.append([r.get("os_id"), nome_usina(r), str(r.get("tarefa") or "")[:90],
                       desc.replace("&ccedil;", "ç").replace("&atilde;", "ã").replace("&otilde;", "õ"),
                       str(r.get("dia") or "").split(" [")[0],
                       periodo(r).replace("&rarr;", "→").replace("&atilde;", "ã").replace("&mdash;", "—")])
        for col, w in zip("ABCDEF", (10, 34, 52, 20, 24, 20)):
            ws.column_dimensions[col].width = w
    ws = wb.active; ws.title = "Programação"
    escreve(ws, rows, "Programação")
    # Aba de reprogramadas pelo REENCONTRO real (mesma lógica do fechamento),
    # não pelo selo histórico `reprog` — este marcava 334 das 364 linhas.
    if rows_ant:
        pend = {str(r.get("os_id")) for r in rows_ant if not finalizada(r)}
        volta = [r for r in rows if str(r.get("os_id")) in pend]
        if volta:
            escreve(wb.create_sheet("Vieram da semana anterior"), volta, "")
    wb.save(destino)
    return os.path.basename(destino)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--semana", type=int, default=None)
    ap.add_argument("--cliente", default=None)
    a = ap.parse_args()

    base = pasta_dados()

    def _ler_db():
        with open(os.path.join(base, "banco_dados.json"), encoding="utf-8") as f:
            return json.load(f)

    def _escolher(semanas):
        if a.semana:
            return next((s for s in semanas if s.get("num") == a.semana), None)
        return max(semanas, key=lambda s: s.get("num", 0)) if semanas else None

    db = _ler_db()
    semanas = db.get("semanas") or []
    alvo = _escolher(semanas)

    # A planilha da semana nova nasce ANTES de o banco_dados.json saber dela: o
    # motor grava o .xlsx e só a rodada seguinte do robô (~54 min) regenera o
    # JSON. Na sexta o PCM gera a programação e clica em seguida no relatório do
    # cliente — sem isto, ou dá erro, ou (pior) manda em silêncio a semana
    # PASSADA para o cliente. Então: se a planilha existe e o JSON não a tem,
    # regenera o JSON aqui.
    _no_json = max((s.get("num", 0) for s in semanas), default=0)
    _nums_xlsx = [int(m.group(1)) for m in
                  (re.match(r"^Programação Semana (\d+)\.xlsx$", f) for f in os.listdir(base))
                  if m]
    _xlsx_alvo = None
    if a.semana and alvo is None and a.semana in _nums_xlsx:
        _xlsx_alvo = f"Programação Semana {a.semana}.xlsx"      # pediram, existe, JSON não tem
    elif not a.semana and _nums_xlsx and max(_nums_xlsx) > _no_json:
        _xlsx_alvo = f"Programação Semana {max(_nums_xlsx)}.xlsx"  # planilha mais nova que o JSON

    if _xlsx_alvo:
        log(f"banco_dados.json não conhece a {_xlsx_alvo} — regenerando antes de continuar...")
        import subprocess
        _r = subprocess.run([sys.executable, "-u",
                             os.path.join(AQUI, "gerar_pcm_json.py")],
                            cwd=base, env={**os.environ, "PCM_PROG_DIR": base})
        if _r.returncode == 0:
            db = _ler_db()
            semanas = db.get("semanas") or []
            alvo = _escolher(semanas)
        else:
            log("! falha ao regenerar o banco_dados.json — seguindo com o que existe")

    if not semanas:
        log("ERRO: banco_dados.json sem semanas."); return 1
    if alvo is None:
        log(f"ERRO: semana {a.semana} não está no banco_dados.json, e não existe "
            f"'Programação Semana {a.semana}.xlsx' na pasta. Gere a programação "
            "dessa semana primeiro."); return 1
    ant = next((s for s in sorted(semanas, key=lambda s: -s.get("num", 0))
                if s.get("num", 0) < alvo.get("num", 0)), None)
    log(f"Semana alvo: {alvo['label']}" + (f" | fechamento: {ant['label']}" if ant else ""))

    contatos, path_cont = carregar_contatos()
    if not contatos:
        log(f"! Sem contatos: crie {os.path.basename(path_cont)} (decisão 12 — "
            "cliente sem contato não é enviado)")

    destino = os.path.join(base, "_relatorios_cliente")
    os.makedirs(destino, exist_ok=True)
    porcli, porcli_ant = {}, {}
    for r in alvo.get("rows", []):
        c = str(r.get("cliente") or "").strip()
        if c:
            porcli.setdefault(c, []).append(r)
    for r in (ant or {}).get("rows", []):
        c = str(r.get("cliente") or "").strip()
        if c:
            porcli_ant.setdefault(c, []).append(r)

    feitos, sem_contato = [], []
    for cli, rows in sorted(porcli.items(), key=lambda x: -len(x[1])):
        if a.cliente and cli.lower() != a.cliente.lower():
            continue
        info = contatos.get(cli.lower(), {})
        para = info.get("para") or []
        resp = info.get("responsavel") or "Coordenação de O&M — Grid Co."
        nome_anexo = ""
        # mesmo critério (e mesmas linhas) que montar_email usa, senão o corte
        # de layout aqui e lá podem discordar na fronteira
        vis = dedup_blocos(rows)
        grande = (len({str(r.get("usina") or "") for r in vis}) > CORTE_USINAS
                  or len(vis) > CORTE_LINHAS)
        if grande:
            nome_anexo = anexo_xlsx(
                cli, rows,
                os.path.join(destino, f"Programacao_{alvo['week']}_{cli.replace(' ','_')}.xlsx"),
                rows_ant=porcli_ant.get(cli, []))
        corpo, _ = montar_email(cli, rows, alvo["label"],
                                (ant or {}).get("label", ""), porcli_ant.get(cli, []),
                                resp, nome_anexo)
        m = EmailMessage()
        if para:
            m["To"] = ", ".join(para)
        else:
            sem_contato.append(cli)
            m["To"] = ""
        m["Subject"] = f"Grid Co. — Programação Semanal | {cli} | {alvo['label']}"
        # X-Unsent: 1 faz o Outlook abrir o .eml como RASCUNHO (com botão
        # Enviar) em vez de mensagem recebida (que só oferece Responder/
        # Encaminhar). É o que entrega a decisão 15 — "rascunho no Outlook" —
        # sem depender de COM/MAPI, que o novo Outlook não expõe.
        m["X-Unsent"] = "1"
        m.set_content("Este e-mail tem versão HTML.")
        m.add_alternative(corpo, subtype="html")
        # cartão de assinatura como imagem INLINE (cid), não como anexo solto
        _sig = assinatura_arquivo()
        if _sig:
            import mimetypes
            _sub = (mimetypes.guess_type(_sig)[0] or "image/png").split("/")[-1]
            with open(_sig, "rb") as f:
                m.get_payload()[1].add_related(
                    f.read(), maintype="image", subtype=_sub, cid="<assinaturagridco>")
        if nome_anexo:
            with open(os.path.join(destino, nome_anexo), "rb") as f:
                m.add_attachment(f.read(), maintype="application", filename=nome_anexo,
                                 subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        arq = os.path.join(destino, f"{alvo['week']}_{cli.replace(' ', '_')}.eml")
        with open(arq, "wb") as f:
            f.write(bytes(m))
        with open(arq.replace(".eml", ".html"), "w", encoding="utf-8") as f:
            f.write('<html><meta charset="utf-8"><body style="background:#f3f3f7;padding:20px">'
                    + corpo + "</body></html>")
        feitos.append((cli, len(rows), len({str(r.get('usina') or '') for r in rows}),
                       "B (portfólio)" if grande else "A (detalhado)", bool(para)))

    print()
    print("  %-14s %6s %7s  %-14s %s" % ("CLIENTE", "linhas", "usinas", "LAYOUT", "destinatário"))
    for c, n, u, lay, tem in feitos:
        print("  %-14s %6d %7d  %-14s %s" % (c, n, u, lay, "ok" if tem else "SEM CONTATO"))
    if sem_contato:
        print()
        log(f"! {len(sem_contato)} cliente(s) sem contato — .eml gerado sem destinatário, "
            f"NÃO envie antes de preencher: {', '.join(sem_contato)}")
    log(f"OK: {len(feitos)} relatório(s) em {os.path.basename(destino)}/ "
        "(.eml para revisar e enviar · .html para conferir)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
