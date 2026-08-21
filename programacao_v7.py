"""
Programação semanal - v7 (motor v8 de priorização, 31/07/2026)
========================
Mudanças v8 (aprovadas pelo Fabrício):
 - Hierarquia por TIER: Religamento > Corretiva Emergencial > Corretiva >
   Inspeção/Preditiva > Preventivas (MPA>MPS>MPT>MPM>MPQ) > Administrativas/Handover.
   Corrige inspeções/preditivas que caíam em 'OUTRA' (atrás de toda MPM) e
   move Handover para o fim da fila (decisão: lista de 38, posição 37).
 - Regra dos 10 dias: tarefa em andamento (Não Iniciada/Em progresso/pausado)
   há >= AGING_DIAS entra na semana (mesmo fora do carry-over de 2 semanas,
   até AGING_JANELA_MAX dias) e fura para o topo do SEU tier (não pula tiers).
 - Modo fim de mês (dia >= FIM_DE_MES_DIA, padrão 20): preventivas do mês
   corrente ainda pendentes sobem um degrau (acima da Inspeção; NUNCA acima
   de corretivas — decisão do usuário).

Mudanças sobre v6:
 - BUGFIX: corretivas (e preventivas) com Status='Em processo' e Data Programada
   de semanas anteriores agora são incluídas como carry-over.
   Antes, o filtro de semana exigia Data Programada >= WEEK_START, excluindo
   tarefas não iniciadas que ficaram para trás (ex: OS 6501/6500 de MA Leste 01).
import sys as _sys_init
try:
    _sys_init.stdout.reconfigure(encoding='utf-8', errors='replace')
    _sys_init.stderr.reconfigure(encoding='utf-8', errors='replace')
except (AttributeError, Exception):
    pass

Mudanças sobre v5 (herdadas do v6):
 - Distribuição proporcional de dias por usina (intercalar prioritário)
 - MPM Inversor / Mod+Tracker entram no rodízio normal (sem dia dedicado)
 - Chave única (OSs ID, Código Equipamento) para tracking de reprogramação
 - Histórico semanal acumulativo (HISTORICO_PATH)
   * Para cada execução: lê histórico, atualiza contagem, regrava
   * "Nº de vezes programada" = quantas semanas a chave já apareceu

Algoritmo (alto nível):
 1. Filtra tarefas da semana
 2. Separa zeladoria (paralela na segunda) e MPA noturno
 3. Para o resto (corretivas + preventivas):
    a. Calcula carga por usina
    b. Atribui dias úteis a usinas proporcionalmente (cada usina ganha
       ceil(carga_usina / carga_total × dias_disp) dias, mínimo 1)
    c. Em cada dia atribuído à usina, escalona as tarefas dela em ordem
       (corretivas primeiro por RPN, depois preventivas por MPA>MPS>MPT>MPM>MPQ)
    d. Tarefas que não cabem nos dias atribuídos viram pendência
"""

import math
import os
import re
import unicodedata
from datetime import date, timedelta

import pandas as pd
import fonte_bd_api  # fonte de dados via API Fracttal (substitui leitura do BD xlsx)

# ====================== Configuração ======================
def _next_monday(ref=None):
    """Retorna a próxima segunda-feira a partir de `ref` (default: hoje).
    Se hoje já for segunda, retorna a segunda da semana seguinte."""
    ref = ref or date.today()
    delta = (7 - ref.weekday()) % 7
    if delta == 0:
        delta = 7
    return ref + timedelta(days=delta)


# WEEK_START é calculado automaticamente como a próxima segunda-feira.
# Para forçar uma data específica, descomente a linha abaixo:
# WEEK_START = date(2026, 5, 11)
#
# OU defina a env var PCM_WEEK_FORCE com o numero ISO da semana (ex: "25").
# O PCM Painel (gui_pcm.py) usa isso quando o usuario seleciona uma semana
# diferente do default. Suporta numero ISO ("25") ou "YYYY-Wxx" ("2026-W25").
def _resolver_week_start():
    forced = os.environ.get("PCM_WEEK_FORCE", "").strip()
    if not forced:
        # Também aceita --semana XX via CLI
        import sys as _sys
        for _i, _a in enumerate(_sys.argv[1:], 1):
            if _a in ('--semana', '-s') and _i < len(_sys.argv):
                forced = _sys.argv[_i]
                break
    if not forced:
        return _next_monday()
    try:
        ano_default = date.today().year
        if "-W" in forced.upper():
            # Formato ISO "2026-W25"
            parte_ano, parte_w = forced.upper().split("-W", 1)
            ano = int(parte_ano)
            num = int(parte_w)
        else:
            ano = ano_default
            num = int(forced)
        # Segunda-feira (weekday=1 em isocalendar) da semana ISO `num`
        return date.fromisocalendar(ano, num, 1)
    except (ValueError, TypeError):
        # Valor invalido — cai no comportamento padrao
        return _next_monday()


# Flag --incremental: preserva schedule existente, só encaixa OSs novas
import sys as _sys_argv
_INCREMENTAL = '--incremental' in _sys_argv.argv


WEEK_START = _resolver_week_start()
DAYS = [WEEK_START + timedelta(days=i) for i in range(5)]
DAY_NAMES = ['Segunda-feira', 'Terça-feira', 'Quarta-feira',
             'Quinta-feira', 'Sexta-feira']
_ISO_YEAR, _ISO_WEEK, _ = WEEK_START.isocalendar()

# Última semana do mês: semana que contém o dia 30 ou 31
# (MPM também recebe prioridade de corretiva nessa semana)
ULTIMA_SEMANA_MES = any(d.day >= 30 for d in DAYS)

# Carregado dinamicamente da pasta Feriados/ (ver _carregar_feriados_dict abaixo)
BR_HOLIDAYS_2026 = set()  # NACIONAL — bloqueia todas equipes
ESTADUAL_FERIADOS = {}     # {date: set(UF)} — bloqueia equipes daquela UF

WORK_START_MIN   = 7 * 60
FIRST_OS_MIN     = 7 * 60 + 30
WORK_END_MIN     = 17 * 60
LUNCH_START_MIN  = 12 * 60
LUNCH_END_MIN    = 13 * 60 + 12
LUNCH_BREAK_MIN  = LUNCH_END_MIN - LUNCH_START_MIN
PCM_END_MIN      = 9 * 60 + 30
NIGHT_START_MIN  = 18 * 60
NIGHT_END_MIN    = 27 * 60
POST_NIGHT_START_MIN = 14 * 60
TERMO_START_MIN  = 11 * 60
TERMO_END_MIN    = 15 * 60
GAP_BETWEEN_OS   = 15
MIN_PREVENTIVA_H = 1.0
TRAVEL_KMH       = 70.0
TRAVEL_FACTOR    = 1.2
# Melhoria 0.1 (13/08/2026) — deslocamento entre usinas:
#   off    = comportamento antigo (deslocamento ignorado)
#   sombra = calcula e reporta na coluna "Desloc (h)", mas NÃO consome capacidade
#   on     = calcula e consome capacidade (decisão 1: só após 1 semana de sombra)
DESLOC_MODO    = os.environ.get('PCM_DESLOC', 'sombra').strip().lower()
DESLOC_CONSOME = (DESLOC_MODO == 'on')
DESLOC_CALCULA = (DESLOC_MODO in ('on', 'sombra'))
USEFUL_DAY_MIN     = int((WORK_END_MIN - WORK_START_MIN - LUNCH_BREAK_MIN) * 0.80)
NIGHT_CAPACITY_MIN = int((NIGHT_END_MIN - NIGHT_START_MIN) * 0.80)

# ---- Janela da MPA (v8, 31/07/2026) ----
# MPA roda em turno estendido: 15:30 até 01:30 do dia seguinte (10h).
# A MPA é SEMPRE inserida na programação, mesmo estourando a janela — nesse
# caso a linha recebe a flag EXCEDE JANELA MPA (nunca vira pendência por
# capacidade). O dia que recebe MPA tem o expediente diurno truncado às 15:30.
MPA_START_MIN    = 15 * 60 + 30      # 15:30
MPA_END_MIN      = 25 * 60 + 30      # 01:30 do dia seguinte
MPA_CAPACITY_MIN = MPA_END_MIN - MPA_START_MIN     # 600 min = 10 h
# Expediente diurno de um dia com MPA (mesma proporção 80% dos demais)
MPA_DAY_USEFUL_MIN = int((MPA_START_MIN - WORK_START_MIN - LUNCH_BREAK_MIN) * 0.80)





def _carregar_observacoes_semana():
    """Lê Observacoes_Semana.txt da pasta base. Cada linha = uma observação.
    Linhas vazias e que começam com # são ignoradas.
    Retorna lista de strings (observações limpas).
    """
    path = os.path.join(BASE_DIR, "Observacoes_Semana.txt")
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            linhas = f.readlines()
    except Exception as e:
        print(f"[AVISO] Falha ao ler observacoes: {e}")
        return []
    # v9 (eco): guarda o nº da linha física junto do texto, para que o relatório
    # de observações consiga apontar ONDE está o erro. Retorna [(lineno, texto)].
    obs = []
    for _i, ln in enumerate(linhas, start=1):
        ln = ln.strip()
        if not ln or ln.startswith("#"):
            continue
        obs.append((_i, ln))
    if obs:
        print(f"[INFO] Observacoes da semana carregadas: {len(obs)}")
    return obs


MUNICIPAL_FERIADOS = {}      # {date: set(cidade_normalizada)} — v9 (0.5)


def _norm_cidade_fer(s):
    """Normaliza nome de cidade p/ casar Feriados x AUXILIAR (acentos/caixa)."""
    s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode()
    return re.sub(r'[^a-z]+', ' ', s.lower()).strip()


def _carregar_feriados_dict():
    """Carrega Feriados/FERIADOS ESTADUAIS, MUNICIPAIS E NACIONAIS 2026.xlsx
    Popula BR_HOLIDAYS_2026 (NACIONAL) e ESTADUAL_FERIADOS (por UF).
    Se o arquivo não existir, mantém sets vazios (sem bloqueio)."""
    global BR_HOLIDAYS_2026, ESTADUAL_FERIADOS
    try:
        import pandas as _pd
        # Localiza o arquivo: pasta Feriados dentro de BASE_DIR
        _ferpath = os.path.join(BASE_DIR, "Feriados",
                                "FERIADOS ESTADUAIS, MUNICIPAIS E NACIONAIS 2026.xlsx")
        if not os.path.exists(_ferpath):
            print(f"[AVISO] Arquivo de feriados não encontrado em {_ferpath}")
            return
        df = _pd.read_excel(_ferpath, sheet_name="Feriados 2026", header=1)
        nac = set()
        est = {}
        for _, row in df.iterrows():
            tipo = str(row.get("Tipo") or "").strip().upper()
            uf = str(row.get("Estado") or "").strip().upper()
            data_ = row.get("Data")
            if data_ is None:
                continue
            try:
                d = data_.date() if hasattr(data_, "date") else data_
            except Exception:
                continue
            if tipo == "NACIONAL":
                nac.add(d)
            elif tipo == "ESTADUAL" and uf and uf != "TODOS":
                est.setdefault(d, set()).add(uf)
        # v9 (0.5): MUNICIPAIS ficam no BLOCO DIREITO da mesma aba (colunas
        # Tipo.1 / MUNICIPIO / Data.1) — nunca foram lidos por ninguém.
        # Grão correto: feriado municipal bloqueia a CIDADE (as usinas dela),
        # não a equipe inteira. Popula MUNICIPAL_FERIADOS {date: {cidade_norm}}.
        mun = {}
        _c_tipo2 = next((c for c in df.columns if str(c).startswith("Tipo.")), None)
        _c_mun = next((c for c in df.columns if "MUNICIPIO" in str(c).upper()), None)
        _c_data2 = next((c for c in df.columns if str(c).startswith("Data.")), None)
        if _c_tipo2 and _c_mun and _c_data2:
            for _, row in df.iterrows():
                if "MUNICIP" not in str(row.get(_c_tipo2) or "").upper():
                    continue
                data_ = row.get(_c_data2)
                cid = _norm_cidade_fer(row.get(_c_mun))
                if data_ is None or not cid:
                    continue
                try:
                    d = data_.date() if hasattr(data_, "date") else data_
                except Exception:
                    continue
                mun.setdefault(d, set()).add(cid)
        MUNICIPAL_FERIADOS.clear()
        MUNICIPAL_FERIADOS.update(mun)
        BR_HOLIDAYS_2026 = nac
        ESTADUAL_FERIADOS = est
        print(f"[INFO] Feriados: {len(nac)} nacionais, "
              f"{sum(len(v) for v in est.values())} estaduais, "
              f"{sum(len(v) for v in mun.values())} municipais")
    except Exception as _e:
        print(f"[AVISO] Falha ao carregar feriados ({type(_e).__name__}: {_e}). "
              f"Prosseguindo sem bloqueio.")


def _is_feriado_para_equipe(data_, equipe):
    """True se a data é feriado nacional OU estadual da UF da equipe."""
    if data_ in BR_HOLIDAYS_2026:
        return True
    if data_ in ESTADUAL_FERIADOS and isinstance(equipe, str):
        partes = equipe.strip().split()
        if partes and partes[0].upper() in ESTADUAL_FERIADOS[data_]:
            return True
    return False

# ====================== Caminhos ======================
# Pasta principal de trabalho.
# Prioridade: variável de ambiente PCM_PROG_DIR > auto-detecção > erro.
import os
import pathlib

def _detectar_base_dir():
    # 1. Variável de ambiente explícita (qualquer SO)
    if 'PCM_PROG_DIR' in os.environ:
        return os.environ['PCM_PROG_DIR']

    _subpath = pathlib.Path('4. O&M') / '11.Pré-Operação' / '6. PCM' / '09. Programação Semanal'

    # 2. Windows — pasta sincronizada do OneDrive (nomes comuns de tenant)
    home = pathlib.Path.home()
    _win_candidates = [
        home / 'GRID CO' / 'Grid Co. - Gridco' / _subpath,
        home / 'Grid Co. - Gridco' / _subpath,
        home / 'OneDrive - Grid Co' / _subpath,
        home / 'OneDrive - Gridco' / _subpath,
        home / 'OneDrive' / _subpath,
    ]
    for c in _win_candidates:
        if c.exists():
            return str(c)

    # 3. Cowork / Linux — mount FUSE do OneDrive (caminho varia por sessão)
    sessions = pathlib.Path('/sessions')
    if sessions.exists():
        for _mnt in sessions.glob('*/mnt/09. Programação Semanal'):
            return str(_mnt)

    raise RuntimeError(
        "Pasta PCM não encontrada automaticamente.\n"
        "Defina a variável de ambiente PCM_PROG_DIR com o caminho completo da pasta\n"
        "'09. Programação Semanal' e tente novamente."
    )

BASE_DIR = _detectar_base_dir()
_carregar_feriados_dict()
OBSERVACOES_SEMANA = _carregar_observacoes_semana()


def _carregar_pins_incremental(xlsx_path: str) -> dict:
    """Lê a Programação Semana XX.xlsx existente e retorna pins para todas as OSs
    já agendadas, preservando o dia em que cada uma estava.
    Retorna dict {os_id (int): {'dia': day_idx, 'tarefas': [], 'start_min': None}}
    """
    import openpyxl
    _DIA_IDX = {'segunda': 0, 'terca': 1, 'terça': 1, 'quarta': 2,
                'quinta': 3, 'sexta': 4}
    pins = {}
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        print(f'[INCREMENTAL] Não foi possível ler {xlsx_path}: {e}')
        return pins

    for sheet in wb.sheetnames:
        if sheet.startswith('_'): continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        header = [str(c).strip() if c is not None else '' for c in rows[0]]
        # Encontra colunas
        try:
            col_os  = header.index('OSs ID')
            col_dia = header.index('Dia')
        except ValueError:
            continue
        for r in rows[1:]:
            if not r or r[col_os] is None: continue
            try:
                os_id = int(r[col_os])
            except (ValueError, TypeError):
                continue
            dia_str = str(r[col_dia] or '').lower()
            # Remove flags como [EXCEDE HH], [PARCIAL] etc.
            dia_str = dia_str.split('(')[0].strip()
            dia_str = dia_str.split('[')[0].strip()
            day_idx = None
            for key, idx in _DIA_IDX.items():
                if key in dia_str:
                    day_idx = idx
                    break
            if day_idx is None: continue
            # Captura TODOS os dias distintos em que a OS aparecia (Modo A re-espalha)
            lst = pins.setdefault(os_id, [])
            if not any(p['dia'] == day_idx for p in lst):
                lst.append({'dia': day_idx, 'tarefas': None, 'start_min': None})
    wb.close()
    print(f'[INCREMENTAL] {len(pins)} OSs travadas do schedule anterior')
    return pins


# Regra "dias de atendimento por usina": @usina <nome> = seg, qua, sex
# Preenchida em _carregar_pins(). {usina_norm: set(dias 0..4)}
USINA_DIAS: dict = {}


def _norm_usina(s):
    import unicodedata as _u
    s = _u.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()
    return re.sub(r"\s+", " ", s).strip()


def _usina_match(ativo, regra_norm):
    """True se a regra (ex.: 'maraba 1') casa com o ativo (ex.: 'Athon - Marabá 1 - PA').
    Usa fronteira de palavra p/ 'Marabá 1' NÃO casar com 'Marabá 10'."""
    a = _norm_usina(ativo)
    return re.search(r"\b" + re.escape(regra_norm) + r"\b", a) is not None


# Siglas de planos de manutenção que são SEMPRE OS com muitas tarefas dentro.
# OS cujo nome de tarefa contém uma destas → MODO A (distribui o bloco pelos dias).
# Qualquer outra OS pinada → MODO B (fixa tarefa específica no dia).
#   MPM=mensal, MPW=semanal, MPT=trimestral, MPS=semestral,
#   MPA=anual, MPB=bianual, MPQ=quinzenal, Handover
SIGLAS_MULTI = ('HANDOVER', 'MPM', 'MPW', 'MPT', 'MPS', 'MPA', 'MPB', 'MPQ')
_RE_SIGLA = re.compile(r"\b(" + "|".join(SIGLAS_MULTI) + r")\b")


def _os_eh_multi(tks):
    """True se ALGUMA tarefa da OS carrega uma sigla de plano (Handover/MPx).
    Nesse caso a OS é tratada em MODO A (bloco distribuído pelos dias)."""
    for tk in tks:
        if _RE_SIGLA.search((tk.get('tarefa_txt') or '').upper()):
            return True
    return False


# Exclusão de tarefas internas por OS: "OS; dias; SIGLA; sem: QGBT, Trafo, Cabine"
# {os_id: [palavra-chave, ...]} — tarefa cujo nome contém a chave não é programada.
OS_TAREFAS_EXCLUIR: dict = {}

# ── v9: eco das observações ──────────────────────────────────────────────────
# O parser passa a registrar o que ENTENDEU e o que NÃO entendeu, para relatar de
# uma vez no fim da geração — nunca no meio do log, onde se perde. Antes, linha com
# erro de digitação era descartada em silêncio: o pin sumia e o log dizia "Pins
# ativos: [...]" como se tivesse funcionado.
OBS_ENTENDIDAS: list = []   # [str] — descrição em linguagem humana
OBS_PROBLEMAS:  list = []   # [(lineno|None, trecho, motivo)]


def relatorio_observacoes():
    """Eco das observações. Chamado no FIM da geração, para não se perder no log."""
    if not OBS_ENTENDIDAS and not OBS_PROBLEMAS:
        return
    print()
    if OBS_ENTENDIDAS:
        print(f'[OBS] {len(OBS_ENTENDIDAS)} linha(s) entendida(s):')
        for _d in OBS_ENTENDIDAS:
            print(f'        OK  {_d}')
    if OBS_PROBLEMAS:
        print(f'[AVISO] {len(OBS_PROBLEMAS)} linha(s) NAO aplicada(s):')
        for _lineno, _trecho, _motivo in OBS_PROBLEMAS:
            _onde = f'linha {_lineno}' if _lineno else 'pin'
            print(f'        !!  {_onde}: "{_trecho[:60]}"')
            print(f'            -> {_motivo}')
        print('        Corrija em Observacoes_Semana.txt e gere de novo.')


def _tarefa_excluida(tk):
    """True se a tarefa deve ser pulada por causa de um filtro 'sem:' da sua OS."""
    kws = OS_TAREFAS_EXCLUIR.get(tk['os_id'])
    if not kws:
        return False
    txt = _norm_usina(tk['tarefa_txt'])
    return any(_norm_usina(k) in txt for k in kws)


def _carregar_pins():
    """Lê Observacoes_Semana.txt e separa linhas no formato estruturado:
        OS_ID;DIA;TAREFAS;TURNO
    Campos:
        OS_ID   — número da OS (obrigatório)
        DIA     — SEG/TER/QUA/QUI/SEX (opcional; vazio = melhor dia disponível)
        TAREFAS — nomes separados por vírgula (opcional; vazio = todas da OS)
        TURNO   — manhã/tarde/noite (opcional; define horário de início)

    Formato especial:
        OS_ID;nao;  — exclui a OS da programação da semana (não aparece nem como pendente)

    Retorna dict: {os_id (int): {'dia': int|None, 'tarefas': [str]|None, 'start_min': int|None}}
    Linhas que não começam com número são tratadas como texto livre (ignoradas aqui).
    """
    # Aceita abreviado E por extenso, com/sem acento (via _norm_usina que tira acento)
    DIA_MAP = {'SEG': 0, 'SEGUNDA': 0, 'TER': 1, 'TERCA': 1, 'QUA': 2, 'QUARTA': 2,
               'QUI': 3, 'QUINTA': 3, 'SEX': 4, 'SEXTA': 4}

    def _dia_idx(s):
        return DIA_MAP.get(_norm_usina(s).upper())

    TURNO_MAP = {
        'MANHA': 7 * 60 + 30, 'MANHÃ': 7 * 60 + 30,
        'TARDE': 13 * 60,
        'NOITE': 18 * 60,
    }
    pins = {}
    for _lineno, ln in OBSERVACOES_SEMANA:
        ln_s = (ln or "").strip()
        if ln_s.lower().startswith('@usina'):
            # @usina <nome> = seg, qua, sex  → fixa os dias de atendimento da usina
            corpo = ln_s[6:]
            if '=' not in corpo:
                OBS_PROBLEMAS.append((_lineno, ln_s,
                    "regra de usina sem '=' — use: @usina Nome = seg, qua"))
                continue
            nome, _, dias_s = corpo.partition('=')
            nome_norm = _norm_usina(nome)
            _toks_u = [d for d in re.split(r'[,\s]+', dias_s.strip()) if d]
            dias = {i for i in (_dia_idx(d) for d in _toks_u) if i is not None}
            _ruins_u = [d for d in _toks_u if _dia_idx(d) is None]
            if not nome_norm:
                OBS_PROBLEMAS.append((_lineno, ln_s, "nome da usina vazio"))
            elif not dias:
                OBS_PROBLEMAS.append((_lineno, ln_s,
                    f"nenhum dia reconhecido em '{dias_s.strip()}' — use seg/ter/qua/qui/sex"))
            else:
                USINA_DIAS[nome_norm] = dias
                _dn_u = {0: 'seg', 1: 'ter', 2: 'qua', 3: 'qui', 4: 'sex'}
                _av = f" (ignorado: {', '.join(_ruins_u)})" if _ruins_u else ""
                OBS_ENTENDIDAS.append(
                    f"@usina {nome.strip()} → {', '.join(_dn_u[i] for i in sorted(dias))}{_av}")
            continue
        parts = [p.strip() for p in ln_s.split(';')]
        if not parts[0].isdigit():
            # Só reclama se a linha PARECE estruturada (tem ';' ou começa com "OS <nº>").
            # Comentário e texto livre continuam passando calados, como sempre.
            if ';' in ln_s or re.match(r'^\s*os\s*[\s#nº:.-]*\d', ln_s, re.I):
                OBS_PROBLEMAS.append((_lineno, ln_s,
                    "comece pelo número da OS, sem prefixo (ex.: 10369; qua)"))
            continue
        os_id = int(parts[0])

        # Campo "sem: QGBT, Trafo, Cabine" — exclui tarefas internas da OS (Modo A)
        campos = parts[1:]
        _sem = [c for c in campos if re.match(r'^\s*sem\s*:', c, re.I)]
        if _sem:
            _kws = re.split(r'[;,]', re.sub(r'^\s*sem\s*:', '', _sem[0], flags=re.I))
            OS_TAREFAS_EXCLUIR[os_id] = [k.strip() for k in _kws if k.strip()]
            campos = [c for c in campos if c not in _sem]
        # Campo "só: Inversor, Cabine" — nesta linha, agenda SÓ as tarefas que batem (Modo A)
        _so = [c for c in campos if re.match(r'^\s*(s[óo]|apenas)\s*:', c, re.I)]
        _incluir = None
        if _so:
            _kws = re.split(r'[;,]', re.sub(r'^\s*(s[óo]|apenas)\s*:', '', _so[0], flags=re.I))
            _incluir = [k.strip() for k in _kws if k.strip()] or None
            campos = [c for c in campos if c not in _so]

        dia_str = campos[0] if len(campos) > 0 else ''

        # Formato "OS_ID; nao" — exclui a OS da programação
        if _norm_usina(dia_str).upper() in ('NAO', 'NO'):
            OS_EXCLUIDAS.add(os_id)
            OBS_ENTENDIDAS.append(f"OS {os_id} → fora da semana (nao)")
            continue

        tarefas_s = campos[1] if len(campos) > 1 else ''
        turno_str = (campos[2] if len(campos) > 2 else '').replace('Ã', 'A').replace('ã', 'a').upper().strip()

        # DIA pode trazer vários dias: "seg, ter, qua" ou "qua e qui" → um pin por dia.
        _toks = [t for t in re.split(r'[,;/\s]+', dia_str.strip()) if t]
        _dias = [d for d in (_dia_idx(t) for t in _toks) if d is not None]
        # Token que não é dia nem conectivo ("e") indica erro de digitação.
        _ruins = [t for t in _toks if _dia_idx(t) is None and _norm_usina(t).upper() != 'E']
        if not _dias:
            if _toks:
                # v9 (eco): havia algo escrito no campo dia e NADA foi reconhecido.
                # Antes isto virava um pin SEM dia — parecia funcionar. Agora é erro.
                OBS_PROBLEMAS.append((_lineno, ln_s,
                    f"dia não reconhecido ({', '.join(_toks)}) — use seg/ter/qua/qui/sex"))
                continue
            _dias = [None]   # campo dia realmente vazio → melhor dia disponível
        elif _ruins:
            OBS_PROBLEMAS.append((_lineno, ln_s,
                f"trecho ignorado no campo dia: {', '.join(_ruins)}"))
        _tarefas = [t.strip() for t in tarefas_s.split(',') if t.strip()] or None
        _turno = TURNO_MAP.get(turno_str)
        if turno_str and _turno is None:
            OBS_PROBLEMAS.append((_lineno, ln_s,
                f"turno não reconhecido ('{turno_str}') — use manhã/tarde/noite"))
        # Múltiplos pins por OS (uma linha por dia e/ou por tarefa) → LISTA
        for _di in _dias:
            pins.setdefault(os_id, []).append({
                'dia': _di,
                'tarefas': _tarefas,
                'start_min': _turno,
                'incluir': _incluir,   # filtro "só:" desta linha (Modo A)
            })
        # v9 (eco): descreve em linguagem humana o que foi entendido desta linha
        _dn = {0: 'segunda', 1: 'terça', 2: 'quarta', 3: 'quinta', 4: 'sexta'}
        _d_txt = ' e '.join(_dn[d] for d in _dias if d is not None) or 'melhor dia disponível'
        _t_txt = (f'só "{", ".join(_incluir)}"' if _incluir else
                  (f'tarefas: {", ".join(_tarefas)}' if _tarefas else 'todas as tarefas'))
        _tu_txt = {7 * 60 + 30: 'manhã', 13 * 60: 'tarde', 18 * 60: 'noite'}.get(_turno, 'turno padrão')
        _sem_txt = ''
        if os_id in OS_TAREFAS_EXCLUIR:
            _sem_txt = f" · sem: {', '.join(OS_TAREFAS_EXCLUIR[os_id])}"
        OBS_ENTENDIDAS.append(f"OS {os_id} → {_d_txt} · {_t_txt} · {_tu_txt}{_sem_txt}")

    # v9 (eco): se havia linhas com cara de estruturada e NENHUMA foi entendida, não é
    # erro de digitação — é arquivo corrompido ou salvo em codificação errada.
    _estruturadas = sum(1 for _, l in OBSERVACOES_SEMANA if ';' in l)
    if _estruturadas >= 3 and not pins and not OS_EXCLUIDAS and not USINA_DIAS:
        relatorio_observacoes()
        raise SystemExit(
            f'ABORTADO: {_estruturadas} linhas com ";" em Observacoes_Semana.txt e '
            f'nenhuma foi entendida. Verifique a codificação do arquivo (deve ser UTF-8).')
    return pins


OS_EXCLUIDAS: set = set()   # preenchido por _carregar_pins()
OS_PINS = _carregar_pins()
# Incremental merge é feito depois de _default_output ser definido (linha ~401)

# Map UFV-normalizada → Responsavel O&M (lookup da aba AUXILIAR do BD)
def _carregar_map_responsavel():
    import unicodedata, re
    mp = {}
    try:
        df_aux = fonte_bd_api.df_auxiliar()   # AUXILIAR - FABRICIO.xlsx (era aba AUXILIAR do BD)
        col_ufv = None
        col_resp = None
        for c in df_aux.columns:
            if str(c).strip().upper() == 'UFV':
                col_ufv = c
            cn = str(c).strip().upper()
            if ('RESPONS' in cn and ('O&M' in cn or 'OM' in cn)
                    and 'URL' not in cn and 'IMAGEM' not in cn):
                col_resp = c
        if col_ufv and col_resp:
            for _, r in df_aux.iterrows():
                u = r.get(col_ufv)
                v = r.get(col_resp)
                if isinstance(u, str) and isinstance(v, str) and v.strip():
                    nrm = ''.join(c for c in unicodedata.normalize('NFD', u.lower().strip())
                                  if unicodedata.category(c) != 'Mn')
                    nrm = re.sub(r'\s+', ' ', nrm)
                    mp[nrm] = v.strip()
        print(f"[INFO] Map Responsavel O&M: {len(mp)} UFVs")
    except Exception as e:
        print(f"[AVISO] Falha ao carregar map Responsavel: {e}")
    return mp

MAP_RESPONSAVEL = _carregar_map_responsavel()

def _resolver_responsavel(ativo_str):
    import unicodedata, re
    if not isinstance(ativo_str, str) or not MAP_RESPONSAVEL:
        return ''
    nrm = ''.join(c for c in unicodedata.normalize('NFD', ativo_str.lower().strip())
                  if unicodedata.category(c) != 'Mn')
    nrm = re.sub(r'\s+', ' ', nrm)
    if nrm in MAP_RESPONSAVEL:
        return MAP_RESPONSAVEL[nrm]
    # Tenta match por prefixo (UFV pode ter sufixos como "1 - SP")
    for k, v in MAP_RESPONSAVEL.items():
        if nrm.startswith(k) or k in nrm:
            return v
    return ''
INPUT_BD       = os.path.join(BASE_DIR, 'BD_Relatório Semanal.xlsx')
INPUT_CONFIAB  = os.path.join(BASE_DIR, 'Planilha Confiabilidade R00.xlsx')
INPUT_PRIOR    = os.path.join(BASE_DIR, 'Lista_Prioridades_GridCo.xlsx')
_default_output = os.path.join(BASE_DIR, f'Programação Semana {_ISO_WEEK:02d}.xlsx')
OUTPUT         = os.environ.get('PCM_OUTPUT', _default_output)

# Merge incremental: trava OSs já agendadas na planilha existente como pins
if _INCREMENTAL:
    _inc_pins = _carregar_pins_incremental(_default_output)
    for _os_id, _pin in _inc_pins.items():
        if _os_id not in OS_PINS:   # pins manuais têm prioridade
            OS_PINS[_os_id] = _pin
    print(f'[INCREMENTAL] Total de pins ativos após merge: {len(OS_PINS)}')
HISTORICO      = os.path.join(BASE_DIR, 'Historico_Programacoes.xlsx')  # acumulativo, persistente

PREVENTIVA_ORDER = {'MPA': 0, 'MPS': 1, 'MPT': 2, 'MPM': 3, 'MPQ': 4, 'OUTRA': 5}

KEYWORDS_ZELADORIA = ['roçagem','rocagem','lavagem','limpeza',
                       'poda','capina','jardim','zelador',
                       'supress','vegetal','vegetação','roça']

# ====================== Helpers ======================
def t(m):
    m = int(round(m))
    h, mn = divmod(m, 60)
    d = h // 24                       # v8: MPA longa pode passar de D+1
    return f'{h%24:02d}:{mn:02d}' + (f' (D+{d})' if d else '')


def norm(s):
    if pd.isna(s): return ''
    s = str(s).strip().lower()
    s = unicodedata.normalize('NFKD', s).encode('ascii','ignore').decode('ascii')
    return re.sub(r'\s+', ' ', s)


def haversine_km(c1, c2):
    if c1 is None or c2 is None: return 0.0
    R = 6371.0
    p1, p2 = math.radians(c1[0]), math.radians(c2[0])
    dphi = math.radians(c2[0] - c1[0])
    dlmb = math.radians(c2[1] - c1[1])
    a = math.sin(dphi/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dlmb/2)**2
    return 2 * R * math.asin(math.sqrt(a))


def norm_equipe(s):
    if pd.isna(s): return None
    parts = str(s).strip().split()
    if not parts: return None
    if len(parts) == 1: return parts[0]
    return ' '.join([parts[0].upper()] + [p.capitalize() for p in parts[1:-1]] + [parts[-1]])


def parse_dms(s):
    m = re.match(r"\s*(\d+(?:\.\d+)?)[°º]\s*(\d+(?:\.\d+)?)?'?\s*(\d+(?:\.\d+)?)?\"?\s*([NSEWLO])?", s, re.IGNORECASE)
    if not m: return None
    g, mn, sec, hemi = m.groups()
    v = float(g) + (float(mn or 0)/60) + (float(sec or 0)/3600)
    if hemi and hemi.upper() in ('S','W','O'): v = -v
    return v


def clean_coord(v, is_lat=True):
    if pd.isna(v): return None
    s = str(v).strip().replace('−','-').replace('–','-').replace('—','-')
    s2 = s.replace('°','').replace('º','')
    if s.count(',') == 1 and not re.search(r'-?\d+\.\d+\s*,\s*-?\d+', s):
        s2 = s2.replace(',', '.')
    if ',' in s2 and re.search(r'\d\s*,\s*-?\d', s2):
        s2 = s2.split(',')[0]
    s2 = s2.strip()
    if "'" in s2 or '"' in s2 or re.search(r'[NSEWLO]', s2, re.I):
        v_dms = parse_dms(s)
        if v_dms is not None: return v_dms
    try: x = float(s2)
    except ValueError: return None
    limit = 90 if is_lat else 180
    if abs(x) >= limit: x = x / 1_000_000.0
    if x > 0: x = -x
    return x


def parse_duration_to_hours(v):
    if pd.isna(v): return None
    if isinstance(v, (int, float)):
        try: return float(v) * 24.0
        except: return None
    s = str(v).strip()
    m = re.match(r'(?:(\d+)\s*days?\s+)?(\d+):(\d+):(\d+)', s)
    if m:
        days = int(m.group(1) or 0); hh = int(m.group(2)); mm = int(m.group(3)); ss = int(m.group(4))
        return days * 24 + hh + mm/60 + ss/3600
    try: return float(s) * 24.0
    except: return None


def extract_city_from_usina(usina):
    s = str(usina).strip()
    s = re.sub(r'\s*-\s*[A-Z]{2}\s*$', '', s)
    s = re.sub(r'\s+\d+(\s+e\s+\d+)?\s*$', '', s)
    return norm(s)


def task_key(os_id, codigo):
    return f"{int(os_id)}|{str(codigo).strip()}"


# ====================== Carregamento ======================
def _safe_read_excel(path, **kwargs):
    """pandas.read_excel com múltiplos fallbacks:
    1. Leitura direta.
    2. Cópia para /tmp via Python open() — contorna BadZipFile de FUSE mounts.
    3. Cópia via ctypes/WinAPI com FILE_SHARE_READ|WRITE|DELETE — contorna
       PermissionError de arquivos bloqueados pelo OneDrive ou Excel aberto.
    """
    try:
        return pd.read_excel(path, **kwargs)
    except Exception:
        import shutil, tempfile, os as _os, sys
        fd, tmp = tempfile.mkstemp(suffix=_os.path.splitext(path)[1])
        _os.close(fd)
        try:
            with open(path, 'rb') as fi, open(tmp, 'wb') as fo:
                fo.write(fi.read())
        except (PermissionError, OSError):
            if sys.platform != 'win32':
                raise
            # Fallback Windows: abre com flags de compartilhamento via ctypes
            import ctypes, ctypes.wintypes
            GENERIC_READ               = 0x80000000
            FILE_SHARE_READ            = 0x1
            FILE_SHARE_WRITE           = 0x2
            FILE_SHARE_DELETE          = 0x4
            OPEN_EXISTING              = 3
            INVALID_HANDLE_VALUE       = ctypes.wintypes.HANDLE(-1).value
            k32 = ctypes.windll.kernel32
            h = k32.CreateFileW(
                path, GENERIC_READ,
                FILE_SHARE_READ | FILE_SHARE_WRITE | FILE_SHARE_DELETE,
                None, OPEN_EXISTING, 0, None,
            )
            if h == INVALID_HANDLE_VALUE:
                raise PermissionError(f'CreateFileW falhou: {path}')
            try:
                size = _os.path.getsize(path)
                buf  = (ctypes.c_char * size)()
                n_read = ctypes.wintypes.DWORD(0)
                k32.ReadFile(h, buf, size, ctypes.byref(n_read), None)
                with open(tmp, 'wb') as fo:
                    fo.write(buf.raw[:n_read.value])
            finally:
                k32.CloseHandle(h)
        return pd.read_excel(tmp, **kwargs)


print('Lendo arquivos...')
df_bd = fonte_bd_api.df_semanal()   # OS/tarefas direto da API Fracttal (era BD 'Semanal')

# Lista de prioridades (RPN + MTTR fallback para preventivas/zeladoria)
df_pri_raw = _safe_read_excel(INPUT_PRIOR, sheet_name='📋 Lista de Prioridades', header=7)
df_pri = df_pri_raw.iloc[:, [0, 1, 2, 3, 6, 7]].copy()
df_pri.columns = ['prio', 'atividade', 'tipo', 'ativo_principal', 'qtd_hist', 'mttr_h']
df_pri = df_pri.dropna(subset=['prio'])
df_pri['prio'] = pd.to_numeric(df_pri['prio'], errors='coerce')
df_pri['mttr_h'] = pd.to_numeric(df_pri['mttr_h'], errors='coerce')
df_pri = df_pri.dropna(subset=['prio']).reset_index(drop=True)

# Planilha Confiabilidade — tempo médio (h) por categoria para corretivas
# Header está na linha 4 do Excel (header=3 em 0-index). Última linha é "TOTAL" sem média.
df_confiab = _safe_read_excel(INPUT_CONFIAB, sheet_name='Resumo por Categoria', header=3)
df_confiab = df_confiab.dropna(subset=['Categoria', 'Média (h)']).copy()
df_confiab = df_confiab[df_confiab['Categoria'].astype(str).str.upper() != 'TOTAL']
df_confiab['Média (h)'] = pd.to_numeric(df_confiab['Média (h)'], errors='coerce')
df_confiab = df_confiab.dropna(subset=['Média (h)']).reset_index(drop=True)
print(f'Planilha Confiabilidade: {len(df_confiab)} categorias carregadas')


def find_corretiva_mttr(tarefa_txt):
    """Procura a categoria de Confiabilidade que melhor casa com a tarefa.
    Retorna Média (h) ou None se nenhum match.
    Estratégia: pontua por palavras-chave (≥4 letras) compartilhadas entre
    a tarefa e o nome da categoria; também considera o Tipo de Ativo."""
    txt = norm(tarefa_txt)
    if not txt:
        return None
    best_h, best_score = None, 0
    for _, row in df_confiab.iterrows():
        cat_norm = norm(row['Categoria'])
        ativo_norm = norm(row.get('Tipo de Ativo', ''))
        # palavras significativas (≥4 letras)
        cat_words = [w for w in re.split(r'[^\w]+', cat_norm) if len(w) >= 4]
        ativo_words = [w for w in re.split(r'[^\w]+', ativo_norm) if len(w) >= 4]
        score = sum(1 for w in cat_words if w in txt)
        score += sum(0.5 for w in ativo_words if w in txt)
        if score > best_score:
            best_score = score
            best_h = float(row['Média (h)'])
    return best_h if best_score >= 1 else None


# ---------- v9 (0.5): feriado MUNICIPAL bloqueia a CIDADE, não a equipe ----------
def _carregar_usina_cidade():
    """Mapa usina_norm -> cidade_norm via AUXILIAR (colunas UFV e CIDADE)."""
    mp = {}
    try:
        df_aux = fonte_bd_api.df_auxiliar()
        cu = next((c for c in df_aux.columns if str(c).strip().upper() == 'UFV'), None)
        cc = next((c for c in df_aux.columns if str(c).strip().upper() == 'CIDADE'), None)
        if cu is not None and cc is not None:
            for _, _r in df_aux.iterrows():
                _u, _c = _r[cu], _r[cc]
                if isinstance(_u, str) and _u.strip() and isinstance(_c, str) and _c.strip():
                    _cidn = _norm_cidade_fer(_c)
                    mp[_norm_usina(_u)] = _cidn
                    # indexa também a grafia sem " - UF" (Fracttal usa com UF)
                    mp[_norm_usina(re.sub(r'\s*-\s*[A-Z]{2}\s*$', '', _u.strip()))] = _cidn
    except Exception as _e:
        print(f"[AVISO] mapa usina→cidade indisponível ({_e}) — municipal usa o nome do ativo")
    return mp


USINA_CIDADE = _carregar_usina_cidade()
_MUN_LOGADOS = set()


def _mun_dias_bloqueados(usina):
    """Índices de dia (0-4) desta semana em que a CIDADE da usina tem feriado
    municipal. Fonte da cidade: AUXILIAR; fallback: nome do ativo."""
    if not MUNICIPAL_FERIADOS:
        return set()
    # O ativo vem "Cliente - Usina N - UF"; o UFV do AUXILIAR vem sem UF.
    # (1ª versão fazia lookup direto e falhava SEMPRE, em silêncio — pego no
    # teste dirigido da W35: o bloqueio nunca disparava.)
    _base = re.sub(r'\s*-\s*[A-Z]{2}\s*$', '', str(usina).strip())
    cid = (USINA_CIDADE.get(_norm_usina(_base))
           or USINA_CIDADE.get(_norm_usina(usina)))
    if not cid:
        # fallback: última parte do nome sem número ("Thopen - Matão 1"→"matao")
        _sem_num = re.sub(r'\s+\d+(\s+e\s+\d+)?\s*$', '', _base)
        cid = _norm_cidade_fer(_sem_num.split(' - ')[-1])
    if not cid:
        return set()
    bloq = set()
    for _di, _d in enumerate(DAYS):
        if cid in MUNICIPAL_FERIADOS.get(_d, ()):
            bloq.add(_di)
            if (str(usina), _di) not in _MUN_LOGADOS:
                _MUN_LOGADOS.add((str(usina), _di))
                print(f"[FERIADO MUNICIPAL] {_d.strftime('%d/%m')} em "
                      f"{cid.title()}: '{usina}' fora da distribuição nesse dia")
    return bloq


# ====================== Coordenadas (Melhoria 0.1) ======================
# Vêm da API Fracttal (latitude/longitud do ativo de LOCALIZAÇÃO nível usina),
# colhidas pelo fonte_bd_api no MESMO laço que carrega as classificações —
# custo de API zero. A chave é o NOME DA USINA (Ativo Classificação 1), não a
# cidade: duas usinas na mesma cidade são pontos distintos.
#
# Leitura PREGUIÇOSA de propósito: no import deste módulo o df_semanal() ainda
# não rodou e o mapa do fonte_bd_api está vazio — capturar aqui congelaria {}.
MAPA_COORD = {}
USINAS_SEM_COORD = set()      # preenchido em runtime, reportado no fim
_COORD_SEM_UF = {}


def resolve_coord(usina):
    """(lat, lon) da usina, ou None se não houver coordenada cadastrada.
    Sem coordenada => trecho vale 0 h, mas a usina entra em USINAS_SEM_COORD e
    aparece no resumo do fim da geração. Nunca falha em silêncio."""
    if not usina:
        return None
    if not MAPA_COORD:
        MAPA_COORD.update(fonte_bd_api.mapa_coordenadas())
        for _k, _v in MAPA_COORD.items():
            _COORD_SEM_UF.setdefault(
                re.sub(r'\s*-\s*[A-Z]{2}\s*$', '', _k).strip().lower(), _v)
    u = str(usina).strip()
    xy = MAPA_COORD.get(u)
    if xy is None:   # grafia com/sem " - UF" diverge entre planilha e cadastro
        xy = _COORD_SEM_UF.get(re.sub(r'\s*-\s*[A-Z]{2}\s*$', '', u).lower())
    if xy is None:
        USINAS_SEM_COORD.add(u)
    return xy


# ====================== Histórico ======================
# Estrutura: { task_key: {'first_week': 'YYYY-WW', 'last_week': 'YYYY-WW', 'count': int, 'weeks': 'YYYY-WW,...'} }
historico = {}
try:
    df_h = pd.read_excel(HISTORICO)
    for _, r in df_h.iterrows():
        k = str(r['task_key'])
        historico[k] = {
            'first_week': r.get('first_week', ''),
            'last_week':  r.get('last_week', ''),
            'count':      int(r.get('count', 0)),
            'weeks':      str(r.get('weeks', '')),
        }
    print(f'Histórico carregado: {len(historico)} tarefas anteriormente programadas')
except FileNotFoundError:
    print('Histórico ainda não existe (1ª semana). Será criado.')


def current_week_str():
    iso = WEEK_START.isocalendar()
    return f'{iso[0]}-W{iso[1]:02d}'


CURRENT_WEEK = current_week_str()


# ====================== Filtragem ======================
df_bd['Data Programada'] = pd.to_datetime(df_bd['Data Programada'], errors='coerce')
df_bd['Data Calculada']  = pd.to_datetime(df_bd['Data Calculada'], errors='coerce')

mask_keep_verif = ((df_bd['Status'] == 'Verificação') &
                   (df_bd['Estado da Tarefa'] == 'Não Iniciada') &
                   (df_bd['Tipo de tarefa'] == 'Preventiva'))
df_filt = df_bd[(df_bd['Status'] == 'Em processo') | mask_keep_verif].copy()
df_filt = df_filt[df_filt['Estado da Tarefa'].isin(['Não Iniciada', 'Em progresso', 'pausado'])]
# Exclui religamentos — não entram na programação semanal
_TIPOS_EXCLUIDOS = {'religamento', 'religamento remoto'}
_mask_excluir = df_filt['Tipo de tarefa'].apply(
    lambda t: str(t).strip().lower() in _TIPOS_EXCLUIDOS)
if _mask_excluir.any():
    print(f'[INFO] Excluídas {_mask_excluir.sum()} tarefas de Religamento/Religamento Remoto')
df_filt = df_filt[~_mask_excluir].copy()
WEEK_END         = WEEK_START + timedelta(days=5)
# BUGFIX v7: tarefas "Em processo" não iniciadas de semanas anteriores são
# carry-overs legítimos. Janela de carry-over: até CARRYOVER_WEEKS semanas
# atrás (padrão 2) — evita puxar todo o histórico mas cobre atrasos recentes.
# Tarefas de outros status (ex: Verificação de preventivas) permanecem
# restritas à semana corrente.
CARRYOVER_WEEKS   = 2
CARRYOVER_START   = WEEK_START - timedelta(weeks=CARRYOVER_WEEKS)
mask_carry_over   = ((df_filt['Status'] == 'Em processo') &
                     (df_filt['Data Programada'] >= pd.Timestamp(CARRYOVER_START)) &
                     (df_filt['Data Programada'] <  pd.Timestamp(WEEK_END)))
mask_semana_atual = ((df_filt['Data Programada'] >= pd.Timestamp(WEEK_START)) &
                     (df_filt['Data Programada'] <  pd.Timestamp(WEEK_END)))
# v8 — Regra dos 10 dias: tarefa em andamento (o filtro de Estado upstream já
# garante Não Iniciada/Em progresso/pausado) há >= AGING_DIAS entra na semana
# mesmo fora da janela de carry-over. AGING_JANELA_MAX evita puxar histórico
# morto (tarefas paradas há meses são caso de revisão manual, não de programação).
AGING_DIAS       = int(os.environ.get('PCM_AGING_DIAS', 10))
AGING_JANELA_MAX = int(os.environ.get('PCM_AGING_MAX', 60))
_hoje_ts = pd.Timestamp('today').normalize()
_idade_filt = (_hoje_ts - df_filt['Data Programada']).dt.days
mask_aging_in = ((df_filt['Status'] == 'Em processo') &
                 (_idade_filt >= AGING_DIAS) & (_idade_filt <= AGING_JANELA_MAX))

# v8 — Modo fim de mês: ativo quando a sexta da semana cai no dia >= FIM_DE_MES_DIA
# (ou quando a semana vira o mês — é a última semana dele). Preventivas do mês
# corrente pendentes sobem um degrau na fila (ver task_sort).
FIM_DE_MES_DIA = int(os.environ.get('PCM_FIM_MES_DIA', 20))
_sexta_ref = WEEK_START + timedelta(days=4)
if _sexta_ref.month == WEEK_START.month:
    FIM_MES_ATIVO = _sexta_ref.day >= FIM_DE_MES_DIA
else:
    FIM_MES_ATIVO = True   # semana que vira o mês = última semana do mês
if FIM_MES_ATIVO:
    print(f'[FIM DE MÊS] Ativo (sexta dia {_sexta_ref.day} >= {FIM_DE_MES_DIA}): '
          f'preventivas do mês corrente sobem acima da Inspeção')

# OS pinadas em Observacoes_Semana.txt entram independente da Data Programada.
# BUGFIX: 'OSs ID' vem da API como string ('7975') e OS_PINS usa int (7975);
# isin() direto casava ZERO. Normaliza ambos p/ string canônica antes de comparar.
_os_id_norm = df_filt['OSs ID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
mask_pinned = _os_id_norm.isin({str(k) for k in OS_PINS.keys()})

# v9 (eco): pin apontando para OS que não está na base da semana. Antes era
# registrado, impresso em "Pins ativos" e não fazia nada — ninguém percebia.
_ids_presentes = set(_os_id_norm)
for _k in OS_PINS:
    if str(_k) not in _ids_presentes:
        OBS_PROBLEMAS.append((None, f"OS {_k}",
            "pinada mas ausente da semana (finalizada? cancelada? número errado?)"))
for _k in OS_EXCLUIDAS:
    if str(_k) not in _ids_presentes:
        OBS_PROBLEMAS.append((None, f"OS {_k}",
            "marcada 'nao' mas já não estava na semana — linha desnecessária"))
# ─────────────────────────────────────────────────────────────────────────────
# FILA DE REPROGRAMAÇÃO ENTRE SEMANAS (21/08/2026 — pedido do PCM)
#
# O carry-over acima resolve 2 semanas e o aging para em AGING_JANELA_MAX (60
# dias). Tarefa vencida há mais tempo saía do plano e não aparecia em semana
# nenhuma — 1.334 tarefas em 21/08. Esta máscara traz de volta o que ainda faz
# sentido executar, com regra por sigla, e o que não couber na capacidade
# continua saindo como PENDENTE (visível), nunca sumindo.
#
# Não mexe no filtro de topo: OS com Status='Finalizados' seguem fora. As 1.059
# tarefas abertas dentro de OS fechada são problema de cadastro e têm relatório
# próprio — despejá-las na semana seria trocar um buraco por ruído.
# ATIVO desde 21/08 por decisão do PCM: a regra é que MPM do mês corrente NUNCA
# pode ficar invisível — ou está na semana, ou aparece em "Não couberam na
# semana". Em sombra ela não aparecia em lugar nenhum, que é o pior dos casos.
# PCM_FILA=sombra volta a desligar.
FILA_MODO = os.environ.get('PCM_FILA', 'ativo').strip().lower()
if FILA_MODO not in ('sombra', 'ativo'):
    FILA_MODO = 'sombra'

_SIG_MES  = ('MPM', 'MPW', 'MPQ', 'MPT')   # só do mês corrente
_SIG_LIVRE = ('MPS', 'MPA')                # de qualquer mês
_TIPO_SEMPRE = ('corretiva', 'inspe')      # tier 0 e 1 da régua da rolagem
_TIPO_NUNCA  = ('handover', 'administrativa')

_rx_sig = re.compile(r'\b(MP[MSAQWT])\b')


def _fila_elegivel(tarefa, tipo, dt_prog):
    """True se a tarefa vencida deve voltar para a fila desta semana."""
    if pd.isna(dt_prog):
        return False
    t = str(tipo or '').strip().lower()
    if any(k in t for k in _TIPO_NUNCA):
        return False
    m = _rx_sig.search(str(tarefa or ''))
    if m:
        sig = m.group(1)
        if sig in _SIG_LIVRE:
            return True
        if sig in _SIG_MES:
            return (dt_prog.month == WEEK_START.month and
                    dt_prog.year == WEEK_START.year)
        return False
    return any(k in t for k in _TIPO_SEMPRE)


_venc = df_filt['Data Programada'] < pd.Timestamp(WEEK_START)
mask_fila = _venc & df_filt.apply(
    lambda r: _fila_elegivel(r.get('Tarefa'), r.get('Tipo de tarefa'),
                             r.get('Data Programada')), axis=1)

_n_fila = int(mask_fila.sum())
if _n_fila:
    _det = df_filt[mask_fila]
    _por = _det['Tipo de tarefa'].astype(str).value_counts().head(6).to_dict()
    _idade_max = int((_hoje_ts - _det['Data Programada']).dt.days.max())
    print(f'[FILA REPROG · {FILA_MODO.upper()}] {_n_fila} tarefa(s) vencida(s) '
          f'voltam para a semana · mais antiga há {_idade_max} dias · {_por}')
    for _, _r in _det.iterrows():
        OBS_PROBLEMAS.append((None,
            f"OS {_r['OSs ID']}",
            f"reprogramada: vencida em {_r['Data Programada'].date()} "
            f"({(_hoje_ts - _r['Data Programada']).days}d) · {str(_r.get('Tarefa',''))[:60]}"))
else:
    print(f'[FILA REPROG · {FILA_MODO.upper()}] nada vencido elegível nesta semana')

if FILA_MODO != 'ativo':
    mask_fila = mask_fila & False          # sombra: contabiliza, não entra
# ─────────────────────────────────────────────────────────────────────────────

mask_week = mask_carry_over | mask_semana_atual | mask_pinned | mask_aging_in | mask_fila
# OS marcadas com "nao" em Observacoes_Semana.txt são excluídas da programação
if OS_EXCLUIDAS:
    mask_excluir_obs = _os_id_norm.isin({str(k) for k in OS_EXCLUIDAS})
    df_filt = df_filt[~mask_excluir_obs].copy()
    mask_week = mask_week[~mask_excluir_obs]
df_tasks = df_filt[mask_week].copy().reset_index(drop=True)

print(f'Tarefas a programar: {len(df_tasks)} (de {df_tasks["OSs ID"].nunique()} OSs)')

df_tasks['Equipe'] = df_tasks['Ativo Classificação 2'].apply(norm_equipe)
df_tasks['_ativo'] = df_tasks['Ativo Classificação 1'].fillna(df_tasks['Ativo'])
df_tasks['_city']  = df_tasks['_ativo'].apply(extract_city_from_usina)
df_tasks['_key']   = df_tasks.apply(lambda r: task_key(r['OSs ID'], r.get('Código','')), axis=1)


# ====================== v9 — Portão de qualidade do cadastro (Melhoria 0.7) ======================
# Decisão 6/6.3 do pacote de 08/08/2026: REMOVE o inequívoco (ativo de teste —
# não existe interpretação em que programá-lo esteja certo), AVISA o ambíguo
# (quase-duplicata, UF divergente — só o PCM sabe). Remove a LINHA, nunca
# aborta: a semana sempre sai. Tudo que o portão faz vai para a aba _Qualidade
# e para o painel — log que ninguém lê é o mesmo que silêncio.
QUALIDADE = []          # [{'Tipo','Item','Detalhe','Ação'}]
ATIVOS_TESTE = {'grid co', 'usina teste', 'teste'}   # comparação normalizada EXATA


def _qnorm(s):
    s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode().lower()
    return re.sub(r'[^a-z0-9]+', '', s)


_ATIVOS_TESTE_N = {_qnorm(a) for a in ATIVOS_TESTE}

# a) REMOVE: tarefa em ativo de teste (ou equipe "TESTE")
_mask_teste = (df_tasks['_ativo'].astype(str).apply(lambda a: _qnorm(a) in _ATIVOS_TESTE_N)
               | df_tasks['Equipe'].astype(str).str.strip().str.lower().eq('teste'))
if _mask_teste.any():
    for _, _r in df_tasks[_mask_teste].iterrows():
        _it = str(_r['_ativo']) if pd.notna(_r['_ativo']) else f"(sem ativo) · equipe {_r['Equipe']}"
        QUALIDADE.append({'Tipo': 'REMOVIDA', 'Item': _it,
                          'Detalhe': f"OS {_r['OSs ID']} · {str(_r.get('Tarefa', ''))[:70]}",
                          'Ação': 'Linha excluída da programação (ativo/equipe de teste)'})
    print(f'[QUALIDADE] {int(_mask_teste.sum())} tarefa(s) em ativo de teste '
          f'removida(s) da programação (ver aba _Qualidade)')
    df_tasks = df_tasks[~_mask_teste].copy().reset_index(drop=True)

# b) AVISA: grafias que normalizam igual (ex.: "Coração 1 - SC" vs "Coração 1- SC")
_por_norm = {}
for _a in sorted(set(df_tasks['_ativo'].astype(str))):
    _por_norm.setdefault(_qnorm(_a), []).append(_a)
for _lst in _por_norm.values():
    if len(_lst) > 1:
        QUALIDADE.append({'Tipo': 'AVISO', 'Item': ' | '.join(_lst),
                          'Detalhe': 'Grafias diferentes que normalizam igual — '
                                     'contam como DUAS usinas na distribuição de dias',
                          'Ação': 'Conferir no Fracttal se é uma usina duplicada'})

# c) AVISA: UF do nome do ativo difere da UF da equipe (ex.: "Nobres 1 - CE" no MT Sul)
for _, _r in df_tasks[['_ativo', 'Equipe']].drop_duplicates().iterrows():
    _m = re.search(r'-\s*([A-Z]{2})\s*$', str(_r['_ativo']).strip())
    _ufa = _m.group(1) if _m else None
    _eq = str(_r['Equipe']).strip()
    _ufe = _eq[:2].upper() if len(_eq) >= 2 and _eq[:2].isalpha() else None
    if _ufa and _ufe and _ufa != _ufe:
        QUALIDADE.append({'Tipo': 'AVISO', 'Item': str(_r['_ativo']),
                          'Detalhe': f'UF do nome ({_ufa}) difere da equipe ({_eq})',
                          'Ação': 'Conferir nome do ativo / Classificação 2 no Fracttal'})
if QUALIDADE:
    _nav = sum(1 for q in QUALIDADE if q['Tipo'] == 'AVISO')
    print(f'[QUALIDADE] {_nav} aviso(s) de cadastro (ver aba _Qualidade)')


# ====================== v8 — Placar Mensal por usina ======================
# Mede o plano preventivo do MÊS CORRENTE por usina/cluster: quantas estavam
# previstas, quantas já foram feitas e quantas ainda faltam.
# REGRA DE OURO: "feita" se mede pelo ESTADO DA TAREFA (nível tarefa), nunca
# pelo Status da OS. Canceladas saem da meta (não são trabalho a fazer).
def _montar_placar_mensal(df_bd_full):
    d = df_bd_full.copy()
    d['_dt'] = pd.to_datetime(d['Data Programada'], errors='coerce')
    no_mes = (d['_dt'].dt.month == WEEK_START.month) & (d['_dt'].dt.year == WEEK_START.year)
    eh_prev = d['Tipo de tarefa'].astype(str).str.strip().str.lower() == 'preventiva'
    cancel  = d['Status'].astype(str).str.strip().str.lower() == 'cancelado'
    base = d[no_mes & eh_prev & ~cancel].copy()
    if base.empty:
        return pd.DataFrame(), {}
    base['_ativo']  = base['Ativo Classificação 1'].fillna(base['Ativo'])
    base['Equipe']  = base['Ativo Classificação 2'].apply(norm_equipe)
    base['_feita']  = (base['Estado da Tarefa'].astype(str).str.strip().str.lower()
                       == 'finalizados')
    g = base.groupby(['Equipe', '_ativo'], dropna=False)
    placar = g.agg(Meta=('_feita', 'size'), Feitas=('_feita', 'sum')).reset_index()
    placar['Faltam'] = placar['Meta'] - placar['Feitas']
    placar['% do mês'] = (placar['Feitas'] / placar['Meta'] * 100).round(0)
    # chave p/ o motor: quais usinas ainda têm déficit no mês
    deficit = {str(r['_ativo']): int(r['Faltam'])
               for _, r in placar.iterrows() if r['Faltam'] > 0}
    return placar, deficit


PLACAR_MENSAL, DEFICIT_MENSAL = _montar_placar_mensal(df_bd)
if len(PLACAR_MENSAL):
    _mt, _ft = int(PLACAR_MENSAL['Meta'].sum()), int(PLACAR_MENSAL['Feitas'].sum())
    print(f'[PLANO DO MÊS] {_mt} preventivas previstas p/ {WEEK_START.month:02d}/{WEEK_START.year} '
          f'| {_ft} feitas ({_ft/max(_mt,1)*100:.0f}%) | {_mt-_ft} faltam '
          f'em {len(DEFICIT_MENSAL)} usina(s)')


# ====================== Classificações ======================
def classify_preventiva(tarefa):
    if pd.isna(tarefa): return 'OUTRA'
    t_ = str(tarefa).lower()
    for code in ['MPA','MPS','MPT','MPM','MPQ']:
        if re.search(rf'\b{code.lower()}\b', t_): return code
    if 'anual' in t_:     return 'MPA'
    if 'semestral' in t_: return 'MPS'
    if 'trimestral' in t_:return 'MPT'
    if 'mensal' in t_:    return 'MPM'
    if 'quinzenal' in t_: return 'MPQ'
    return 'OUTRA'


def is_corretiva(tipo):
    return str(tipo).strip().lower().startswith('corretiva')


# ---------- v8: hierarquia por TIER ----------
TIER_RELIGAMENTO  = 0
TIER_EMERGENCIAL  = 1
TIER_CORRETIVA    = 2
TIER_INSPECAO     = 3
TIER_PREVENTIVA   = 4
TIER_ADMIN        = 5   # Administrativas + Handover (decisão: fim de fila)


def classify_tier(tipo_bd, tarefa, mp_cat):
    """Hierarquia da programação (v8):
    0 Religamento > 1 Corretiva Emergencial > 2 Corretiva >
    3 Inspeção/Preditiva > 4 Preventivas > 5 Administrativas/Handover.
    Detecção pelo Tipo de tarefa do Fracttal; descrição como fallback.
    (Religamentos normalmente já são excluídos do filtro de entrada.)"""
    t_ = norm(tipo_bd)
    d_ = norm(tarefa)
    if t_.startswith('religamento'):
        return TIER_RELIGAMENTO
    if t_.startswith('corretiva emerg'):
        return TIER_EMERGENCIAL
    if t_.startswith('corretiva'):
        return TIER_CORRETIVA
    if 'handover' in t_ or 'handover' in d_:
        return TIER_ADMIN
    if t_.startswith(('administrativa', 'teste', 'informativa')):
        return TIER_ADMIN
    if t_.startswith(('inspec', 'predit')):
        return TIER_INSPECAO
    if t_.startswith(('preventiva', 'zeladoria')):
        return TIER_PREVENTIVA
    # Fallback pela descrição/categoria
    if mp_cat and mp_cat != 'OUTRA':
        return TIER_PREVENTIVA
    if 'termografi' in d_ or 'inspec' in d_ or 'mapeamento' in d_:
        return TIER_INSPECAO
    return TIER_PREVENTIVA


def is_termografia(tarefa):
    return 'termografi' in str(tarefa).lower()


def is_zeladoria(tarefa):
    t_ = norm(tarefa)
    return any(kw in t_ for kw in (norm(k) for k in KEYWORDS_ZELADORIA))


def find_rpn_and_mttr(tarefa_txt, tipo_bd):
    txt = norm(tarefa_txt)
    best_prio, best_mttr, best_score = None, None, 0
    for _, p in df_pri.iterrows():
        atv_norm = norm(p['atividade'])
        tipo_p = str(p['tipo']).lower() if pd.notna(p['tipo']) else ''
        atv_words = [w for w in re.split(r'[^\w]+', atv_norm) if len(w) >= 4]
        score = sum(1 for w in atv_words if w in txt)
        if 'corretiva' in tipo_bd.lower() and 'corretiva' in tipo_p: score += 1
        if 'preventiva' in tipo_bd.lower() and 'preventiva' in tipo_p: score += 1
        if score > best_score:
            best_score = score; best_prio = int(p['prio']); best_mttr = p['mttr_h']
    return (best_prio if best_prio else 99,
            float(best_mttr) if pd.notna(best_mttr) else None)


rpn_list, mttr_list = [], []
for _, r in df_tasks.iterrows():
    rpn, mttr = find_rpn_and_mttr(r.get('Tarefa',''), str(r.get('Tipo de tarefa','')))
    rpn_list.append(rpn); mttr_list.append(mttr)
df_tasks['rpn'] = rpn_list
df_tasks['mttr_h_fallback'] = mttr_list


# ============ v9 — Melhoria 4: RPN dinâmico (MODO SOMBRA por padrão) ============
# Hoje o RPN vem de match de TEXTO contra a Lista_Prioridades: nome fora do
# padrão => RPN 99 silencioso (é uma das 4 falhas silenciosas da auditoria).
# O dinâmico calcula dos DADOS que a API já entrega:
#   Impacto (40%)     criticidade Fracttal + tipo de equipamento + porte da usina
#   Recorrência (30%) corretivas no MESMO equipamento em 90 dias + nº de vezes programada
#   Urgência (30%)    dias em aberto + solicitação marcada como urgente
# Decisão 24: componente sem dado NÃO zera nem manda pro fim — entra neutro 5,0
# e a linha é marcada "RPN parcial". Nunca o silêncio do 99.
# Decisão 23: congela na geração de sexta (o painel mostra o recalculado ao lado).
# Sombra: só reporta na coluna; a ordenação continua usando o RPN da planilha.
RPN_DINAMICO_ATIVO = os.environ.get('PCM_RPN_DINAMICO', '0').strip() == '1'
PESO_IMPACTO    = float(os.environ.get('PCM_RPN_W_IMPACTO', 40))
PESO_RECORR     = float(os.environ.get('PCM_RPN_W_RECORR', 30))
PESO_URGENCIA   = float(os.environ.get('PCM_RPN_W_URGENCIA', 30))
RPN_JANELA_DIAS = int(os.environ.get('PCM_RPN_JANELA_DIAS', 90))
_NEUTRO = 5.0

_CRIT_NOTA = {'muito alto': 10.0, 'very_high': 10.0, 'alto': 8.0, 'high': 8.0,
              'medio': 5.0, 'médio': 5.0, 'medium': 5.0,
              'baixo': 2.5, 'low': 2.5, 'muito baixo': 1.0, 'very_low': 1.0}
# Equipamento que, parado, derruba geração — peso maior no impacto
_EQUIP_NOTA = [(('inversor', 'skid'), 10.0),
               (('transformador', 'trafo', 'cabine', 'qgbt', 'quadro de juncao',
                 'quadro de junção', 'cabos ca', 'cabos cc', 'cabo ca', 'cabo cc',
                 'medicao', 'medição'), 9.0),
               (('string', 'modulo', 'módulo', 'tracker', 'conjunto de'), 7.0),
               (('estacao solarimetrica', 'estação solarimétrica', 'solarimetric',
                 'piranometro', 'piranômetro', 'monitor'), 5.0),
               (('nobreak', 'spda', 'seguranca', 'segurança', 'camera', 'câmera',
                 'ppci', 'incendio', 'incêndio', 'aterramento'), 4.0),
               (('infraestrutura', 'caixa d', 'agua', 'água', 'container', 'contêiner',
                 'piso', 'cerca', 'abrigo'), 3.0),
               (('zeladoria', 'roçagem', 'rocagem', 'limpeza', 'supressao', 'supressão',
                 'vegetal', 'capina', 'poda', 'aceiro'), 2.0),
               (('relatorio', 'relatório', 'inspecao', 'inspeção', 'acompanhamento',
                 'coleta de dados', 'handover'), 5.0)]


def _chave_usina(s):
    """Chave para casar usina entre API e AUXILIAR.

    As duas fontes divergem de duas formas — as MESMAS já tratadas no
    gerar_mpas_json: a API traz " - UF" no fim ("Athon - Jacundá 1 - PA") e o
    AUXILIAR numera de 100 em 100 ("Jacundá 100" x "Jacundá 1"). Sem isso o
    lookup falhava em 44% das usinas, em silêncio.
    """
    s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode().lower()
    s = re.sub(r'\s*-\s*[a-z]{2}\s*$', '', s.strip())      # tira " - UF"
    s = re.sub(r'[^a-z0-9]+', ' ', s).strip()
    s = re.sub(r'\b(\d)00\b', r'\1', s)                    # 100->1, 200->2
    return s


def _carregar_mwp():
    """Porte da usina (MWp) do AUXILIAR — entra no Impacto."""
    mp = {}
    try:
        df_aux = fonte_bd_api.df_auxiliar()
        cu = next((c for c in df_aux.columns if str(c).strip().upper() == 'UFV'), None)
        cm = next((c for c in df_aux.columns if 'MWP' in str(c).upper()), None)
        if cu is not None and cm is not None:
            for _, _r in df_aux.iterrows():
                if isinstance(_r[cu], str) and isinstance(_r[cm], (int, float)):
                    mp[_chave_usina(_r[cu])] = float(_r[cm])
    except Exception as e:
        print(f'[RPN] porte da usina indisponível ({e}) — Impacto usa neutro nesse componente')
    return mp


USINA_MWP = _carregar_mwp()
_MWP_MAX = max(USINA_MWP.values()) if USINA_MWP else 0.0


def _corretivas_recentes(df_bd_full):
    """{código_equipamento: nº de corretivas nos últimos RPN_JANELA_DIAS}.
    Decisão 22: recorrência por EQUIPAMENTO; recuo para categoria+usina."""
    por_equip, por_usina = {}, {}
    try:
        d = df_bd_full.copy()
        d['_dt'] = pd.to_datetime(d['Data Programada'], errors='coerce')
        corte = pd.Timestamp('today').normalize() - pd.Timedelta(days=RPN_JANELA_DIAS)
        d = d[(d['_dt'] >= corte) & d['Tipo de tarefa'].astype(str).str.lower().str.startswith('corretiva')]
        for _, _r in d.iterrows():
            _c = str(_r.get('Código') or '').strip()
            if _c:
                por_equip[_c] = por_equip.get(_c, 0) + 1
            _u = _norm_usina(_r.get('Ativo Classificação 1') or _r.get('Ativo') or '')
            if _u:
                por_usina[_u] = por_usina.get(_u, 0) + 1
    except Exception as e:
        print(f'[RPN] histórico de corretivas indisponível ({e})')
    return por_equip, por_usina


CORR_EQUIP, CORR_USINA = _corretivas_recentes(df_bd)


def _nota_equipamento(txt):
    t = norm(txt)
    for chaves, nota in _EQUIP_NOTA:
        if any(k in t for k in chaves):
            return nota
    return None


def calcular_rpn_dinamico(row):
    """Devolve (rpn 1..40, detalhe, parcial). Menor = mais prioritário — mesma
    escala do RPN da planilha, para poder comparar lado a lado."""
    faltou = []

    # ---- Impacto (40%) ----
    crit = _CRIT_NOTA.get(str(row.get('Tarefa -> Criticidade') or '').strip().lower())
    if crit is None:
        crit = _NEUTRO; faltou.append('criticidade')
    equip = _nota_equipamento(f"{row.get('Tarefa','')} {row.get('Código','')}")
    if equip is None:
        equip = _NEUTRO; faltou.append('equipamento')
    _k = _chave_usina(row.get('_ativo', ''))
    # recuo: a API costuma numerar a 1ª usina ("Altair 1") onde o AUXILIAR não
    # numera ("Altair"). Sem isto, 35 usinas ficavam sem porte.
    mwp = USINA_MWP.get(_k) or USINA_MWP.get(re.sub(r'\s+\d+$', '', _k))
    if mwp is None or not _MWP_MAX:
        porte = _NEUTRO; faltou.append('porte')
    else:
        porte = 1.0 + 9.0 * (mwp / _MWP_MAX)          # 0,13 MWp -> ~1 ; 7,13 -> 10
    impacto = 0.45 * crit + 0.35 * equip + 0.20 * porte

    # ---- Recorrência (30%) ----
    cod = str(row.get('Código') or '').strip()
    n = CORR_EQUIP.get(cod) if cod else None
    if n is None:
        n = CORR_USINA.get(_norm_usina(row.get('_ativo', '')))
        if n is not None:
            n = n / 4.0                                # recuo: dilui o nº da usina
    if n is None:
        rec_base = _NEUTRO; faltou.append('recorrência')
    else:
        rec_base = min(10.0, 2.0 + 2.0 * float(n))     # 1 corretiva -> 4 ; 4+ -> 10
    vezes = historico.get(row.get('_key'), {}).get('count', 0)
    recorrencia = min(10.0, rec_base + min(3.0, 0.75 * vezes))

    # ---- Urgência (30%) ----
    idade = row.get('idade_dias')
    if idade is None or pd.isna(idade):
        urg_idade = _NEUTRO; faltou.append('idade')
    else:
        urg_idade = min(10.0, max(0.0, float(idade)) / 3.0)   # 30 dias -> 10
    urgente = str(row.get('Número de Solicitação') or '').strip() not in ('', 'nan')
    urgencia = min(10.0, urg_idade + (2.0 if urgente else 0.0))

    nota = (PESO_IMPACTO * impacto + PESO_RECORR * recorrencia
            + PESO_URGENCIA * urgencia) / (PESO_IMPACTO + PESO_RECORR + PESO_URGENCIA)
    # 10 (pior caso) -> RPN 1 ; 0 -> RPN 40. Mesma escala da planilha.
    rpn = int(round(40 - (nota / 10.0) * 39))
    rpn = max(1, min(40, rpn))
    det = (f"I{impacto:.1f} R{recorrencia:.1f} U{urgencia:.1f}"
           + (f" | parcial: {', '.join(faltou)}" if faltou else ""))
    return rpn, det, bool(faltou)


df_tasks['corretiva']   = df_tasks['Tipo de tarefa'].apply(is_corretiva)
df_tasks['termografia'] = df_tasks['Tarefa'].apply(is_termografia)
df_tasks['zeladoria']   = df_tasks['Tarefa'].apply(is_zeladoria)
df_tasks['mp_cat']      = df_tasks['Tarefa'].apply(classify_preventiva)

# ---------- v8: tier, idade e boosts ----------
df_tasks['tier'] = [classify_tier(t, ta, c) for t, ta, c in
                    zip(df_tasks['Tipo de tarefa'], df_tasks['Tarefa'], df_tasks['mp_cat'])]
df_tasks['idade_dias'] = (pd.Timestamp('today').normalize()
                          - df_tasks['Data Programada']).dt.days
df_tasks['aging'] = df_tasks['idade_dias'].fillna(-1) >= AGING_DIAS
if FIM_MES_ATIVO:
    # v8 bloco 2: só recebe boost a preventiva do mês corrente cuja USINA ainda
    # tem déficit no Placar Mensal (usina com o mês fechado não precisa furar fila).
    df_tasks['boost_mes'] = ((df_tasks['tier'] == TIER_PREVENTIVA) &
                             (df_tasks['Data Programada'].dt.month == WEEK_START.month) &
                             (df_tasks['Data Programada'].dt.year  == WEEK_START.year) &
                             (df_tasks['_ativo'].astype(str).isin(DEFICIT_MENSAL.keys())))
else:
    df_tasks['boost_mes'] = False
if df_tasks['aging'].any():
    print(f"[AGING] {int(df_tasks['aging'].sum())} tarefa(s) há >= {AGING_DIAS} dias "
          f"em andamento — topo do próprio tier")
if FIM_MES_ATIVO and df_tasks['boost_mes'].any():
    print(f"[FIM DE MÊS] {int(df_tasks['boost_mes'].sum())} preventiva(s) do mês "
          f"corrente com prioridade elevada")


# v9/M4 — calculado AQUI (e não antes) porque depende de idade_dias/aging,
# criados no bloco acima. Na 1ª versão rodava antes e 100% das linhas saíam
# marcadas 'parcial: idade' — a marca da decisão 24 foi o que denunciou.
_rpn_novo, _rpn_det, _rpn_parcial = [], [], []
for _, r in df_tasks.iterrows():
    a, b, c = calcular_rpn_dinamico(r)
    _rpn_novo.append(a); _rpn_det.append(b); _rpn_parcial.append(c)
df_tasks['rpn_novo'] = _rpn_novo
df_tasks['rpn_detalhe'] = _rpn_det
df_tasks['rpn_parcial'] = _rpn_parcial
if len(df_tasks):
    _n99 = int((df_tasks['rpn'] == 99).sum())
    _npar = int(df_tasks['rpn_parcial'].sum())
    print(f"[RPN DINÂMICO] modo {'ATIVO' if RPN_DINAMICO_ATIVO else 'SOMBRA'} | "
          f"{_n99} tarefa(s) sem match na planilha (RPN 99) receberam nota calculada | "
          f"{_npar} com algum componente ausente (marcadas 'parcial')")
if RPN_DINAMICO_ATIVO:
    df_tasks['rpn'] = df_tasks['rpn_novo']
    print('[RPN DINÂMICO] ATIVO — a ordenação passa a usar a nota calculada')


def estimate_h(row):
    # Durações fixas por tipo de atividade (independente do BD).
    # Calibradas em 31/07/2026 com o tempo de execução real (real_duration) de
    # 26.267 tarefas finalizadas do Fracttal — média após remoção de outliers (IQR).
    # Ver 'Planilha Confiabilidade R01.xlsx', aba "Motor x Real".
    cat = row.get('mp_cat')
    tarefa = str(row.get('Tarefa', '')).lower()
    if cat == 'MPM':                          return 1.10  # real 1.10h (n=4303) — era 0.75
    if cat == 'MPS':                          return 1.30  # real 1.31h (n=1247) — era 1.0
    if cat == 'MPQ':                          return 0.75  # sem dado válido: apontamentos de ~3s (só SEMP)
    if cat == 'MPA':                          return 4.0   # real 3.95h (n=526)  — era 1.5 (subestimava 2.6x)
    if 'mpw' in tarefa:                       return 0.75  # sem amostra no histórico
    if 'handover' in tarefa:                  return 0.65  # real 0.65h (n=2217) — era 0.5
    dur = parse_duration_to_hours(row.get('Tarefa -> Duração estimada'))
    # Para corretivas, dar prioridade à Média (h) da Planilha Confiabilidade
    if (not dur or dur <= 0) and row['corretiva']:
        dur = find_corretiva_mttr(row.get('Tarefa', ''))
    if not dur or dur <= 0: dur = row.get('mttr_h_fallback')
    if not dur or dur <= 0: dur = 1.5
    if not row['corretiva'] and not row['zeladoria'] and dur < MIN_PREVENTIVA_H:
        dur = MIN_PREVENTIVA_H
    return dur


df_tasks['dur_h'] = df_tasks.apply(estimate_h, axis=1)


# ====== v9 — Melhoria 6 nível 1: duração APRENDIDA (MODO SOMBRA por padrão) ======
# O aprender_duracoes.py mede o tempo real (real_duration) das finalizadas e grava
# duracoes_aprendidas.json com a duração típica por categoria × usina → × cluster
# → global (decisão 30). Aqui só LEMOS. Em sombra, a coluna "Duração aprendida (h)"
# mostra o que o motor usaria; a capacidade continua sendo consumida pela estimativa
# atual (armadilha 3.4 — nada que mude capacidade entra sem uma semana de sombra).
DURACAO_APRENDIDA_ATIVA = os.environ.get('PCM_DURACAO_APRENDIDA', '0').strip() == '1'
_APRENDIDO = {}
try:
    _pa = os.path.join(BASE_DIR, 'duracoes_aprendidas.json')
    if os.path.exists(_pa):
        import json as _json
        with open(_pa, encoding='utf-8') as _f:
            _APRENDIDO = _json.load(_f)
        print(f"[DURAÇÃO APRENDIDA] {len(_APRENDIDO.get('porUsina', {}))} categoria×usina, "
              f"{len(_APRENDIDO.get('porCluster', {}))} categoria×cluster, "
              f"{len(_APRENDIDO.get('global', {}))} global "
              f"(de {str(_APRENDIDO.get('geradoEm'))[:10]})")
    else:
        print('[DURAÇÃO APRENDIDA] duracoes_aprendidas.json ausente — '
              'rode aprender_duracoes.py na sexta, antes da geração')
except Exception as _e:
    print(f'[DURAÇÃO APRENDIDA] arquivo ilegível ({_e}) — seguindo com a estimativa atual')


def _cat_aprendizado(row):
    c = row.get('mp_cat')
    if c and c != 'OUTRA':
        return c
    t = norm(row.get('Tarefa', ''))
    return 'HANDOVER' if 'handover' in t else 'OUTRA'


APRENDE_TETO_TAREFA = float(os.environ.get('PCM_APRENDE_TETO_TAREFA', 2.0))


def duracao_aprendida(row):
    """Duração medida p/ esta tarefa, do mais específico ao mais geral
    (decisão 30). None quando não há amostra suficiente em nenhum nível.

    Dois guarda-corpos descobertos NO MODO SOMBRA (12/08/2026):

    1) ZELADORIA fica de fora. É terceirizada: o apontamento no Fracttal
       registra a conferência do técnico (~1 h), não as horas da equipe de
       roçagem (20 h estimadas). Aplicar cortaria 296 h da carga da semana e o
       motor encheria os dias com trabalho que não cabe.

    2) Teto POR TAREFA, além do teto por categoria do aprender_duracoes.py:
       uma Inspeção estimada em 29 h não vira 0,89 h só porque a média global
       da categoria é essa. Divergência acima de 2x mantém a estimativa atual.
    """
    if not _APRENDIDO:
        return None, ''
    if row.get('zeladoria'):
        return None, 'zeladoria (terceirizada — apontamento não mede o esforço)'
    cat = _cat_aprendizado(row)
    k = _chave_usina(row.get('_ativo', ''))
    achado = None
    it = _APRENDIDO.get('porUsina', {}).get(f'{cat}|{k}')
    if it:
        achado = (it['h'], f"usina (n={it['n']})")
    if achado is None:
        eq = str(row.get('Equipe') or '').strip()
        it = _APRENDIDO.get('porCluster', {}).get(f'{cat}|{eq}')
        if it:
            achado = (it['h'], f"cluster (n={it['n']})")
    if achado is None:
        it = _APRENDIDO.get('global', {}).get(cat)
        if it:
            achado = (it['h'], f"global (n={it['n']})")
    if achado is None:
        return None, ''
    h, origem = achado
    atual = float(row.get('dur_h') or 0)
    if atual > 0:
        razao = h / atual
        if razao > APRENDE_TETO_TAREFA or razao < 1.0 / APRENDE_TETO_TAREFA:
            return None, f'fora do teto ({razao:.1f}x vs estimativa) — mantida'
    return h, origem


_dap, _dorig = [], []
for _, _r in df_tasks.iterrows():
    _h, _org = duracao_aprendida(_r)
    _dap.append(_h); _dorig.append(_org)
df_tasks['dur_aprendida'] = _dap
df_tasks['dur_origem'] = _dorig
if _APRENDIDO and len(df_tasks):
    _com = df_tasks['dur_aprendida'].notna()
    if _com.any():
        _dif = (df_tasks.loc[_com, 'dur_aprendida'] - df_tasks.loc[_com, 'dur_h'])
        print(f"[DURAÇÃO APRENDIDA] modo {'ATIVO' if DURACAO_APRENDIDA_ATIVA else 'SOMBRA'} | "
              f"{int(_com.sum())} de {len(df_tasks)} tarefa(s) com medição | "
              f"efeito na carga: {_dif.sum():+.1f} h "
              f"({100 * _dif.sum() / max(df_tasks['dur_h'].sum(), 1):+.1f}%)")
    if DURACAO_APRENDIDA_ATIVA:
        df_tasks.loc[_com, 'dur_h'] = df_tasks.loc[_com, 'dur_aprendida']
        print('[DURAÇÃO APRENDIDA] ATIVO — a capacidade passa a usar a medição')

# Reprogramada = chave já apareceu em semana anterior
def is_reprog(key):
    return key in historico and historico[key].get('count', 0) > 0


df_tasks['reprogramada'] = df_tasks['_key'].apply(is_reprog)

print(f'Corretivas={df_tasks["corretiva"].sum()}, '
      f'Preventivas={(~df_tasks["corretiva"]).sum()}, '
      f'Zeladoria={df_tasks["zeladoria"].sum()}, '
      f'Reprogramadas (com histórico)={df_tasks["reprogramada"].sum()}')


# ====================== DayCapacity ======================
class DayCapacity:
    def __init__(self, day_idx, equipe=None):
        self.day_idx = day_idx
        self.date = DAYS[day_idx]
        self.equipe = equipe
        # Feriado nacional bloqueia todas; estadual bloqueia só equipes da UF
        self.is_holiday = _is_feriado_para_equipe(self.date, equipe)
        self.start_min = FIRST_OS_MIN
        if day_idx == 0:
            self.start_min = max(self.start_min, PCM_END_MIN)
        self.cursor = self.start_min
        self.used = 0
        self.ativo = None    # ativo do dia
        self.coord = None
        self.night_used = False
        self.night_cursor = NIGHT_START_MIN   # próximo horário livre à noite
        self.night_used_min = 0               # minutos já ocupados à noite
        # v8: fim do expediente diurno por dia (dia com MPA é truncado às 15:30)
        self.work_end_min = WORK_END_MIN
        self.useful_min   = USEFUL_DAY_MIN
        self.mpa_night    = False
        self.mpa_rows     = []   # [(idx_em_rows, start_min, end_min)] p/ reajuste final

    def truncar_para_mpa(self):
        """Dia que recebe MPA: expediente diurno acaba 15:30 (MPA começa aí)."""
        self.mpa_night   = True
        self.work_end_min = min(self.work_end_min, MPA_START_MIN)
        self.useful_min   = min(self.useful_min, MPA_DAY_USEFUL_MIN)

    def remaining_min(self):
        return self.useful_min - self.used


def alloc_diurno(dc, dur_h, prefer_morning, termo, coord, allow_partial=False, force=False):
    """Aloca uma tarefa no dia. Se force=True, ignora capacidade (corretivas emergenciais)."""
    MIN_PIECE_MIN = 30
    if dc.is_holiday: return None
    dur_min = int(round(dur_h * 60))
    remaining = dc.useful_min - dc.used        # v8: capacidade do dia (MPA trunca)
    if not force:
        if remaining <= 0: return None
        if dur_min > remaining and not allow_partial: return None
        if allow_partial and remaining < MIN_PIECE_MIN and dur_min > remaining:
            return None
    effective_min = min(dur_min, remaining)
    if dc.coord is None or coord is None or not DESLOC_CALCULA:
        desloc_h = 0.0; chegada = dc.cursor
    else:
        desloc_h = (haversine_km(dc.coord, coord) / TRAVEL_KMH) * TRAVEL_FACTOR
        # Em modo sombra o tempo é medido e reportado, mas não desloca o cursor.
        chegada = dc.cursor + (desloc_h * 60 if DESLOC_CONSOME else 0.0)
    start = chegada
    if termo:
        start = max(start, TERMO_START_MIN)
        if start < LUNCH_START_MIN and start + effective_min > LUNCH_START_MIN:
            start = LUNCH_END_MIN
        end = start + effective_min
        if end > TERMO_END_MIN: return None
        return (start, end, desloc_h, effective_min/60.0, effective_min >= dur_min)
    if dc.day_idx == 0 and start < PCM_END_MIN:
        start = PCM_END_MIN
    if prefer_morning and start + effective_min <= LUNCH_START_MIN:
        return (start, start + effective_min, desloc_h, effective_min/60.0,
                effective_min >= dur_min)
    if start < LUNCH_START_MIN and start + effective_min > LUNCH_START_MIN:
        end = start + effective_min + LUNCH_BREAK_MIN
    elif LUNCH_START_MIN <= start < LUNCH_END_MIN:
        start = LUNCH_END_MIN
        end = start + effective_min
    else:
        end = start + effective_min
    if end > dc.work_end_min:
        if not allow_partial: return None
        if start < LUNCH_START_MIN:
            available = (LUNCH_START_MIN - start) + (dc.work_end_min - LUNCH_END_MIN)
        elif start >= LUNCH_END_MIN:
            available = dc.work_end_min - start
        else:
            available = dc.work_end_min - LUNCH_END_MIN
        if available <= 0 or available < MIN_PIECE_MIN: return None
        effective_min = available
        if start < LUNCH_START_MIN and start + effective_min > LUNCH_START_MIN:
            end = start + effective_min + LUNCH_BREAK_MIN
        else:
            end = start + effective_min
        end = min(end, dc.work_end_min)
        return (start, end, desloc_h, effective_min/60.0, False)
    return (start, end, desloc_h, effective_min/60.0, True)


def commit_diurno(dc, start, end, dur_min_used, coord, ativo):
    dc.cursor = end + GAP_BETWEEN_OS
    dc.used += dur_min_used
    dc.coord = coord
    if dc.ativo is None:
        dc.ativo = ativo


def alloc_noturno(dc, dur_h):
    """Aloca uma tarefa no turno da NOITE (18h→03h) do dia. None se sem capacidade."""
    if dc.is_holiday:
        return None
    dur_min = int(round(dur_h * 60))
    if dc.night_used_min + dur_min > NIGHT_CAPACITY_MIN:
        return None
    start = dc.night_cursor
    return (start, start + dur_min, dur_min)


def agendar_mpa_janela(equipe, dias, day_caps, tks, rows, ativo=None,
                       day_assignments=None):
    """v8: agenda tarefas de MPA na janela estendida 15:30 → 01:30 (D+1).

    A MPA entra mesmo estourando a janela (a tarefa que cruza 01:30 é
    programada e marcada EXCEDE JANELA MPA), e TRANSBORDA para as noites
    seguintes da lista. O que não couber nas noites da semana NÃO é cramado:
    volta como sobra para virar pendência — a MPA é reprogramada para a
    semana seguinte, com a data definida nas observações.

    dias: índices de dias candidatos, em ordem de preferência.
    Retorna a lista de tarefas que não couberam (sobra).
    """
    dias = [d for d in dias if 0 <= d <= 4 and not day_caps[d].is_holiday]
    if not dias:
        return list(tks)
    fila = list(tks)
    agendadas = 0
    tocados = set()

    def _por(di, tk):
        """Coloca uma tarefa na janela do dia di. Devolve os minutos usados."""
        nonlocal agendadas
        dc = day_caps[di]
        cur = dc.night_cursor if dc.night_used else max(MPA_START_MIN, dc.cursor)
        dur_min = int(round(tk['dur_h'] * 60))
        fim = cur + dur_min
        rows.append(_row(equipe, di, tk, cur, fim, 0.0,
                         noturno=True, over_capacity=fim > MPA_END_MIN))
        dc.mpa_rows.append((len(rows) - 1, cur, fim))
        dc.night_used = True
        dc.night_cursor = fim + GAP_BETWEEN_OS
        dc.night_used_min += dur_min
        agendadas += 1
        tocados.add(di)
        return dur_min

    # Passe 1 — enche cada noite até 01:30, na ordem de preferência
    for di in dias:
        dc = day_caps[di]
        while fila:
            cur = dc.night_cursor if dc.night_used else max(MPA_START_MIN, dc.cursor)
            dur_min = int(round(fila[0]['dur_h'] * 60))
            if cur + dur_min > MPA_END_MIN and dc.night_used:
                break                      # noite cheia → próxima
            _por(di, fila.pop(0))
        if not fila:
            break

    for di in tocados:
        dc = day_caps[di]
        dc.truncar_para_mpa()          # expediente diurno desse dia acaba 15:30
        if (day_assignments is not None and ativo is not None
                and day_assignments[di] is None):
            day_assignments[di] = ativo
            dc.ativo = ativo
        if di + 1 < 5:                 # vira a noite → começa mais tarde no dia seguinte
            nxt = day_caps[di + 1]
            nxt.start_min = max(nxt.start_min, POST_NIGHT_START_MIN)
            nxt.cursor = max(nxt.cursor, nxt.start_min)
    return fila                        # sobra → pendência (próxima semana)


def commit_noturno(dc, end, dur_min, day_idx, day_caps):
    dc.night_cursor = end + GAP_BETWEEN_OS
    dc.night_used_min += dur_min
    if not dc.night_used:
        dc.night_used = True
        if day_idx + 1 < 5:   # quem trabalha à noite começa mais tarde no dia seguinte
            nxt = day_caps[day_idx + 1]
            nxt.start_min = max(nxt.start_min, POST_NIGHT_START_MIN)
            nxt.cursor = max(nxt.cursor, nxt.start_min)


# ====================== Distribuição proporcional de dias ======================
def distribute_days(usinas_carga, days_disp):
    """usinas_carga: dict {usina: carga_h}. days_disp: int.
    Retorna dict {usina: nº_dias} somando até days_disp, com mínimo 1 por usina (se houver carga).
    Se houver mais usinas que dias, prioriza maior carga.
    Cada usina recebe no máximo ceil(carga / USEFUL_DAY_HOURS) dias —
    evita dar mais dias do que a usina consegue ocupar."""
    USEFUL_DAY_HOURS = USEFUL_DAY_MIN / 60.0
    usinas_validas = [(u, c) for u, c in usinas_carga.items() if c > 0]
    n = len(usinas_validas)
    if n == 0:
        return {}
    # Teto: dias necessários para esgotar a carga da usina
    max_days = {u: max(1, math.ceil(c / USEFUL_DAY_HOURS))
                 for u, c in usinas_validas}
    if n >= days_disp:
        usinas_validas.sort(key=lambda x: -x[1])
        chosen = usinas_validas[:days_disp]
        return {u: 1 for u, _ in chosen}
    total = sum(c for _, c in usinas_validas)
    dist = {u: 1 for u, _ in usinas_validas}
    extras = days_disp - n
    if extras > 0 and total > 0:
        shares = [(u, c / total * extras) for u, c in usinas_validas]
        floor_shares = [(u, int(s), s - int(s)) for u, s in shares]
        for u, f, _ in floor_shares:
            dist[u] += f
        used = sum(f for _, f, _ in floor_shares)
        rem = extras - used
        floor_shares.sort(key=lambda x: -x[2])
        for u, _, _ in floor_shares[:rem]:
            dist[u] += 1
    # Aplicar teto: redistribuir excessos para usinas com folga
    for _ in range(10):  # algumas iterações para estabilizar
        changed = False
        for u in list(dist.keys()):
            if dist[u] > max_days[u]:
                excess = dist[u] - max_days[u]
                dist[u] = max_days[u]
                # Distribuir o excesso para usinas que ainda comportam mais dias
                # priorizando as mais carregadas
                cand = sorted([u2 for u2 in dist if dist[u2] < max_days[u2]],
                              key=lambda x: -usinas_carga[x])
                for u2 in cand:
                    if excess <= 0: break
                    pode = max_days[u2] - dist[u2]
                    add = min(pode, excess)
                    dist[u2] += add
                    excess -= add
                changed = True
        if not changed: break
    return dist


def assign_days_to_usinas(usinas_carga_h, day_caps, mpa_done_idx=None):
    """Atribui cada dia útil disponível a uma usina.
    Retorna lista de tamanho 5: day_assignments[i] = usina (ou None se feriado/dia indisponível).
    
    Estratégia:
     - Calcula nº de dias por usina (distribute_days)
     - Distribui no calendário intercalando usinas (regra de intercalar)
    """
    available_idx = [i for i in range(5) if not day_caps[i].is_holiday]
    days_disp = len(available_idx)
    if not days_disp:
        return [None] * 5

    def _dias_permitidos(u):
        # Regra "@usina X = dias": só os dias permitidos (∩ disponíveis). Sem regra: todos.
        base = None
        for regra, dias in USINA_DIAS.items():
            if _usina_match(u, regra):
                base = [i for i in available_idx if i in dias]
                break
        if base is None:
            base = list(available_idx)
        # v9 (0.5): feriado MUNICIPAL tira o dia da cidade daquela usina
        bloq = _mun_dias_bloqueados(u)
        return [i for i in base if i not in bloq] if bloq else base

    days_per_usina = distribute_days(usinas_carga_h, days_disp)
    # Usina com regra não recebe mais dias do que os permitidos disponíveis
    for u in list(days_per_usina):
        perm = _dias_permitidos(u)
        if len(perm) < len(available_idx):   # tem regra ativa
            days_per_usina[u] = min(days_per_usina[u], len(perm)) if perm else 0

    # Sequência intercalada round-robin (usina mais carregada primeiro)
    usinas_ord = sorted(days_per_usina.keys(), key=lambda u: -usinas_carga_h[u])
    sequence = []
    remaining = dict(days_per_usina)
    while sum(remaining.values()) > 0:
        progrediu = False
        for u in usinas_ord:
            if remaining.get(u, 0) > 0:
                sequence.append(u); remaining[u] -= 1; progrediu = True
        if not progrediu:
            break
    sequence = sequence[:days_disp]

    # Aloca cada dia à 1ª usina (em ordem de prioridade) PERMITIDA naquele dia
    assignments = [None] * 5
    seq = list(sequence)
    for di in available_idx:
        for k, u in enumerate(seq):
            if u is None:
                continue
            if di in _dias_permitidos(u):
                assignments[di] = u
                seq[k] = None
                break
    return assignments


# ====================== Estrutura de tarefa atômica ======================
def build_tarefas_atomicas(sub):
    """Cada linha do BD vira uma tarefa atômica."""
    tarefas = []
    for _, r in sub.iterrows():
        ativo = r['_ativo']
        # 0.1: chave = usina (chavear por cidade daria distância zero entre
        # duas usinas da mesma cidade). extract_city_from_usina segue viva
        # para r['_city'] — não remover.
        coord = resolve_coord(str(ativo))
        cidade = r['_city']
        _oid = int(float(str(r['OSs ID']).strip()))   # robusto p/ '7975' e '7975.0'
        tarefas.append({
            'os_id': _oid,
            'codigo': str(r.get('Código','')).strip(),
            'task_key': r['_key'],
            'ativo': ativo,
            'cidade': cidade,
            'coord': coord,
            'tarefa_txt': r.get('Tarefa',''),
            'estado_antes': r.get('Estado da Tarefa',''),
            'tipo_bd': r.get('Tipo de tarefa',''),
            'corretiva': bool(r['corretiva']),
            'zeladoria': bool(r['zeladoria']),
            'termo': bool(r['termografia']),
            'mp_cat': r['mp_cat'],
            'mp_order': PREVENTIVA_ORDER.get(r['mp_cat'], 5) if not r['corretiva'] else -1,
            'rpn': int(r['rpn']),
            'rpn_novo': int(r.get('rpn_novo', 0) or 0),          # v9/M4 (sombra)
            'rpn_detalhe': str(r.get('rpn_detalhe', '') or ''),
            'dur_aprendida': (float(r['dur_aprendida'])
                              if pd.notna(r.get('dur_aprendida')) else None),   # v9/M6
            'dur_origem': str(r.get('dur_origem', '') or ''),
            'reprog': bool(r['reprogramada']),
            'tier': int(r['tier']),                                   # v8
            'aging': bool(r['aging']),                                # v8: >= AGING_DIAS em andamento
            'idade_dias': int(r['idade_dias']) if pd.notna(r['idade_dias']) else None,  # v8
            'boost_mes': bool(r['boost_mes']),                        # v8: preventiva do mês (fim de mês)
            'dur_h': float(r['dur_h']),
            'etiquetas': str(r.get('Etiquetas','') or '').strip(),
            'pin': OS_PINS.get(_oid),  # None ou lista de pins {'dia','tarefas','start_min'}
        })
    return tarefas


# ====================== Escalonamento por equipe ======================
def schedule_team(equipe, tarefas_all):
    day_caps = [DayCapacity(di, equipe=equipe) for di in range(5)]
    rows, pendentes = [], []

    # 1) Zeladoria: paralela na segunda
    zel = [t_ for t_ in tarefas_all if t_['zeladoria']]
    zel_cursor = FIRST_OS_MIN
    for tk in zel:
        dur_min = int(round(tk['dur_h']*60))
        rows.append(_row(equipe, 0, tk, zel_cursor, zel_cursor + dur_min, 0.0,
                         zeladoria=True))
        zel_cursor += dur_min + GAP_BETWEEN_OS

    # 2) MPA noturno — OS com pin nas observações saem daqui e vão pro step 3
    mpa = [t_ for t_ in tarefas_all
           if t_['mp_cat'] == 'MPA' and not t_['zeladoria'] and not t_.get('pin')]
    # Agrupa por OS para fazer todas as tarefas da MPA na mesma noite
    by_os = {}
    for tk in mpa:
        by_os.setdefault(tk['os_id'], []).append(tk)
    # v8: a MPA é SEMPRE inserida (janela 15:30 → 01:30). Se estourar a janela,
    # as linhas excedentes são marcadas mas continuam programadas — só vira
    # pendência se a equipe não tiver nenhum dia útil na semana.
    for os_id, tks in by_os.items():
        total_min = sum(int(round(tk['dur_h']*60)) for tk in tks) + GAP_BETWEEN_OS*(len(tks)-1)
        uteis = [(di, dc) for di, dc in enumerate(day_caps) if not dc.is_holiday]
        if not uteis:
            for tk in tks:
                pendentes.append(_pend(equipe, tk, 'Sem dia útil na semana (feriados)'))
            continue
        # 1º dia ainda sem MPA; se todos já têm, empilha no de menor ocupação noturna
        livres = [di for di, dc in uteis if not dc.night_used]
        # 1º dia sem MPA (ou o de menor ocupação) e, na sequência, os demais
        # dias úteis — para o bloco transbordar em vez de virar 40h seguidas.
        if livres:
            ordem = livres + [di for di, _ in uteis if di not in livres]
        else:
            ordem = [di for di, _ in sorted(uteis, key=lambda x: x[1].night_used_min)]
        sobra = agendar_mpa_janela(equipe, ordem, day_caps, tks, rows)
        if total_min > MPA_CAPACITY_MIN:
            print(f'  [MPA] OS {os_id} ({equipe}): {total_min/60:.1f}h > janela de '
                  f'{MPA_CAPACITY_MIN/60:.0f}h — distribuída entre as noites da semana'
                  + (f' ({len(sobra)} tarefa(s) p/ próxima semana)' if sobra else ''))
        for tk in sobra:
            pendentes.append(_pend(equipe, tk,
                'MPA excede a capacidade noturna da semana — reprogramar '
                '(defina o dia nas observações)'))

    # 3) Resto: corretivas + preventivas (não MPA, não zeladoria)
    # MPA pinadas entram aqui também (saíram do step 2 para obedecer o pin)
    resto = [t_ for t_ in tarefas_all
             if not t_['zeladoria']
             and (t_['mp_cat'] != 'MPA' or t_.get('pin'))]
    if not resto:
        return rows, pendentes

    # Carga por usina (em h)
    usinas_carga = {}
    for tk in resto:
        usinas_carga[tk['ativo']] = usinas_carga.get(tk['ativo'], 0.0) + tk['dur_h']

    # Atribuir dias a usinas (distribuição proporcional)
    day_assignments = assign_days_to_usinas(usinas_carga, day_caps)

    # Marcar ativo do dia
    for di, ativo in enumerate(day_assignments):
        if ativo is not None:
            day_caps[di].ativo = ativo

    # Agrupar tarefas por usina e ordenar dentro de cada usina (v8):
    # TIER (religamento > emergencial > corretiva > inspeção > preventivas >
    # admin/handover) > aging (>=10d fura pro topo do próprio tier) >
    # mp_order (só preventivas: MPA>MPS>MPT>MPM>MPQ) > lista de 38 (rpn) >
    # reprogramada primeiro.
    # Fim de mês: preventiva do mês corrente sobe p/ 2.5 (acima da Inspeção,
    # nunca acima de corretiva).
    def task_sort(tk):
        et = tk['tier']
        if tk.get('boost_mes'):
            et = TIER_CORRETIVA + 0.5
        return (et,
                0 if tk.get('aging') else 1,
                tk['mp_order'] if tk['tier'] == TIER_PREVENTIVA else 0,
                tk['rpn'],
                0 if tk['reprog'] else 1)

    by_usina = {}
    for tk in resto:
        by_usina.setdefault(tk['ativo'], []).append(tk)
    for u in by_usina:
        by_usina[u].sort(key=task_sort)

    # --- Pinning: alocar OS fixadas antes do fluxo normal ---
    def _tarefa_matches_pin(tk, pin):
        """Verifica se a tarefa bate com o filtro de subtarefas do pin."""
        if not pin.get('tarefas'):
            return True  # sem filtro = todas as tarefas da OS
        txt = tk['tarefa_txt'].lower()
        return any(f.lower() in txt for f in pin['tarefas'])

    for ativo, lista in by_usina.items():
        # Agrupa por OS para decidir MODO A (sigla → distribui bloco pelos dias)
        # vs MODO B (OS sem sigla → fixa tarefa específica no dia).
        tarefas_por_os = {}
        for tk in lista:
            tarefas_por_os.setdefault(tk['os_id'], []).append(tk)

        pinned_A = []        # [(tk, allowed_days, turno_map)] — distribui pelos dias
        pinned_com_dia = []  # MODO B — tarefa fixada num dia exato
        pinned_sem_dia = []  # pin sem dia → alta prioridade na fila normal
        normais = []

        for os_id, tks in tarefas_por_os.items():
            pins_os = OS_PINS.get(os_id)
            if not pins_os:
                normais.extend(tks); continue
            if isinstance(pins_os, dict):
                pins_os = [pins_os]   # compat: pin único antigo

            if _os_eh_multi(tks):
                # MODO A: bloco (Handover/MPx). "sem:" remove tarefas da semana.
                tks = [tk for tk in tks if not _tarefa_excluida(tk)]
                if not tks:
                    continue
                # Agrupa as linhas da OS em DIRETRIZES por (filtro "só:", turno).
                # Cada diretriz junta os dias das suas linhas.
                _dirs = {}
                for p in pins_os:
                    if p.get('dia') is None:
                        continue
                    _inc = list(p.get('incluir') or [])
                    _key = (tuple(sorted(k.lower() for k in _inc)), p.get('start_min'))
                    _d = _dirs.setdefault(_key, {'dias': set(), 'inc': _inc, 'start': p.get('start_min')})
                    _d['dias'].add(p['dia'])
                if not _dirs:
                    pinned_sem_dia.extend(tks)   # sigla sem dia → só prioridade
                else:
                    _com = [d for d in _dirs.values() if d['inc']]   # linhas com "só:"
                    _sem_so = [d for d in _dirs.values() if not d['inc']]  # linha que pega o resto
                    for tk in tks:
                        _txt = (tk['tarefa_txt'] or '').lower()
                        alvo = next((d for d in _com if any(k.lower() in _txt for k in d['inc'])), None)
                        if alvo is None and _sem_so:
                            alvo = _sem_so[0]
                        if alvo is None:
                            continue   # todas as linhas têm "só:" e esta tarefa não bateu → fora
                        _ds = sorted(alvo['dias'])
                        _tmap = {dd: alvo['start'] for dd in _ds} if alvo['start'] is not None else {}
                        pinned_A.append((tk, _ds, _tmap))
            else:
                # MODO B: cada tarefa casa com o pin do seu nome (mais específico 1º)
                for tk in tks:
                    pin = None
                    for p in sorted(pins_os, key=lambda x: 0 if x.get('tarefas') else 1):
                        if _tarefa_matches_pin(tk, p):
                            pin = p; break
                    if pin and pin.get('dia') is not None:
                        tk['_pin'] = pin; pinned_com_dia.append(tk)
                    elif pin:
                        tk['_pin'] = pin; pinned_sem_dia.append(tk)
                    else:
                        normais.append(tk)

        # Sempre reconstrói a fila normal desta usina a partir de `normais`:
        # garante que tarefas de OS pinada (Modo A/B) NÃO caiam no passe normal,
        # mesmo quando TODAS foram filtradas por "só:" (senão vazariam pro dia errado).
        by_usina[ativo] = pinned_sem_dia + sorted(normais, key=task_sort)

        # Tarefas pinadas com dia: alocar agora no dia exato
        for tk in pinned_com_dia:
            pin = tk['_pin']
            di = pin['dia']
            if di < 0 or di > 4:
                pendentes.append(_pend(equipe, tk, f'Pin: dia inválido ({di})'))
                continue
            dc = day_caps[di]
            if dc.is_holiday:
                pendentes.append(_pend(equipe, tk, f'Pin: dia {di} é feriado'))
                continue
            # Respeita turno; se não definido usa o cursor atual do dia
            if pin.get('start_min') is not None:
                # Avança o cursor para o início do turno se ainda não chegou lá
                dc.cursor = max(dc.cursor, pin['start_min'])
            res = alloc_diurno(dc, tk['dur_h'],
                               prefer_morning=True,
                               termo=tk['termo'],
                               coord=tk['coord'],
                               allow_partial=False)
            if res is None:
                if tk['mp_cat'] == 'MPA':
                    # v8: MPA pinada não cabe de dia → janela estendida 15:30→01:30
                    if agendar_mpa_janela(equipe, [di], day_caps, [tk], rows,
                                          ativo=ativo, day_assignments=day_assignments):
                        pendentes.append(_pend(equipe, tk,
                            f'MPA: noite do dia {di} sem espaço — reprogramar '
                            '(defina o dia nas observações)'))
                    else:
                        print(f'[MPA] OS {tk["os_id"]} | {tk["tarefa_txt"][:44]} '
                              f'→ dia {di} na janela estendida (dia cheio)')
                    continue
                pendentes.append(_pend(equipe, tk,
                    f'Pin: OS {tk["os_id"]} não coube no dia {di} (sem capacidade)'))
                continue
            start, end, desloc_h, used_h, completed = res
            used_min = int(round(used_h * 60))
            is_first = (dc.coord != tk['coord'])
            rows.append(_row(equipe, di, tk, start, end,
                             desloc_h if is_first else 0.0))
            commit_diurno(dc, start, end, used_min, tk['coord'], ativo)
            # Garante que o dia fique atribuído à usina correta
            if day_assignments[di] is None:
                day_assignments[di] = ativo
                dc.ativo = ativo
            print(f'[PIN] OS {tk["os_id"]} | {tk["tarefa_txt"][:50]} '
                  f'→ dia {di} {t(start)}-{t(end)}')

        # MODO A: distribui as tarefas da OS pelos dias permitidos.
        # Preenche o 1º dia até esgotar a capacidade; overflow vai pro próximo dia.
        pinned_A.sort(key=lambda x: task_sort(x[0]))
        _DN = {0: 'seg', 1: 'ter', 2: 'qua', 3: 'qui', 4: 'sex'}
        for tk, allowed, turno_map in pinned_A:
            placed = False
            for di in allowed:
                if di < 0 or di > 4:
                    continue
                dc = day_caps[di]
                if dc.is_holiday:
                    continue
                _turno = turno_map.get(di)
                if _turno == NIGHT_START_MIN:
                    # turno da NOITE (18h→03h) → slot noturno
                    res = alloc_noturno(dc, tk['dur_h'])
                    if res is None:
                        continue
                    start, end, used_min = res
                    rows.append(_row(equipe, di, tk, start, end, 0.0, noturno=True))
                    commit_noturno(dc, end, used_min, di, day_caps)
                    if day_assignments[di] is None:
                        day_assignments[di] = ativo; dc.ativo = ativo
                    placed = True
                    break
                # turno de DIA (manhã/tarde) ou sem turno
                if _turno is not None:
                    dc.cursor = max(dc.cursor, _turno)
                res = alloc_diurno(dc, tk['dur_h'],
                                   prefer_morning=(_turno is None or _turno < LUNCH_START_MIN),
                                   termo=tk['termo'], coord=tk['coord'], allow_partial=False)
                if res is None:
                    continue  # dia cheio → tenta o próximo dia permitido
                start, end, desloc_h, used_h, completed = res
                used_min = int(round(used_h * 60))
                is_first = (dc.coord != tk['coord'])
                rows.append(_row(equipe, di, tk, start, end,
                                 desloc_h if is_first else 0.0))
                commit_diurno(dc, start, end, used_min, tk['coord'], ativo)
                if day_assignments[di] is None:
                    day_assignments[di] = ativo
                    dc.ativo = ativo
                placed = True
                break
            if not placed and tk['mp_cat'] == 'MPA':
                # v8: MPA não coube de dia → janela estendida 15:30→01:30
                # dos dias permitidos do pin. Se nem lá couber, vai p/ pendência.
                if not agendar_mpa_janela(equipe, list(allowed), day_caps, [tk], rows,
                                          ativo=ativo, day_assignments=day_assignments):
                    print(f'[MPA] OS {tk["os_id"]} | {tk["tarefa_txt"][:44]} '
                          f'→ janela estendida (dia cheio)')
                    placed = True
            if not placed:
                pendentes.append(_pend(equipe, tk,
                    f'Modo A: OS {tk["os_id"]} não coube nos dias '
                    f'{[_DN[d] for d in allowed]}'))

    # Alocar tarefas por usina nos dias atribuídos
    # Estratégia em 2 passes por dia:
    #   Passe 1 — corretivas (sempre entram; se sem capacidade → force + [EXCEDE HH])
    #             Na última semana do mês: MPM também entra no passe 1
    #   Passe 2 — preventivas restantes (só entram se houver capacidade)
    def _deve_forcar(tk):
        """True se a tarefa deve entrar mesmo sem capacidade disponível."""
        if tk['corretiva']: return True
        if ULTIMA_SEMANA_MES and tk['mp_cat'] == 'MPM': return True
        return False

    for di in range(5):
        ativo = day_assignments[di]
        if ativo is None: continue
        dc = day_caps[di]
        if dc.is_holiday: continue
        if ativo not in by_usina: continue
        tarefas_da_usina = by_usina[ativo]

        # --- Passe 1: prioritárias (corretivas + MPM na última semana) ---
        prioritarias = [tk for tk in tarefas_da_usina if _deve_forcar(tk)]
        for tk in prioritarias:
            res = alloc_diurno(dc, tk['dur_h'],
                               prefer_morning=True,
                               termo=tk['termo'],
                               coord=tk['coord'],
                               allow_partial=False)
            excede = False
            if res is None:
                # Força entrada mesmo sem capacidade
                dur_min = int(round(tk['dur_h'] * 60))
                if dc.coord is None or tk['coord'] is None or not DESLOC_CALCULA:
                    desloc_h_val = 0.0; chegada = dc.cursor
                else:
                    desloc_h_val = (haversine_km(dc.coord, tk['coord']) / TRAVEL_KMH) * TRAVEL_FACTOR
                    chegada = dc.cursor + (desloc_h_val * 60 if DESLOC_CONSOME else 0.0)
                start = chegada
                end = start + dur_min
                res = (start, end, desloc_h_val, tk['dur_h'], True)
                excede = True
            start, end, desloc_h_val, used_h, completed = res
            used_min = int(round(used_h * 60))
            is_first = (dc.coord != tk['coord'])
            rows.append(_row(equipe, di, tk, start, end,
                             desloc_h_val if is_first else 0.0,
                             over_capacity=excede))
            commit_diurno(dc, start, end, used_min, tk['coord'], ativo)
            tarefas_da_usina.remove(tk)
            if excede:
                print(f'[EXCEDE HH] OS {tk["os_id"]} | {tk["tarefa_txt"][:50]} '
                      f'→ dia {di} (corretiva/MPM forçada)')

        # --- Passe 2: preventivas (só entram se houver capacidade restante) ---
        i = 0
        while i < len(tarefas_da_usina):
            tk = tarefas_da_usina[i]
            res = alloc_diurno(dc, tk['dur_h'],
                               prefer_morning=False,
                               termo=tk['termo'],
                               coord=tk['coord'],
                               allow_partial=True)
            if res is None:
                i += 1
                continue
            start, end, desloc_h_val, used_h, completed = res
            used_min = int(round(used_h*60))
            is_first = (dc.coord != tk['coord'])
            rows.append(_row(equipe, di, tk, start, end,
                             desloc_h_val if is_first else 0.0,
                             partial=not completed,
                             used_h_override=used_h if not completed else None))
            commit_diurno(dc, start, end, used_min, tk['coord'], ativo)
            if completed:
                tarefas_da_usina.pop(i)
            else:
                tk['dur_h'] = tk['dur_h'] - used_h
                break



    # 4) Passe de preenchimento: dias atribuídos mas vazios (usina já esgotada
    # OU usina não tinha tarefas pendentes) recebem trabalho de outras usinas.
    for di in range(5):
        dc = day_caps[di]
        if dc.is_holiday: continue
        if dc.used > 0: continue  # dia já em uso de fato
        # Escolher usina com mais carga pendente
        pendentes_carga = {u: sum(tk['dur_h'] for tk in tks)
                            for u, tks in by_usina.items() if tks}
        if not pendentes_carga: break
        ativo = max(pendentes_carga, key=pendentes_carga.get)
        # Sobrescreve ativo do dia
        dc.ativo = ativo
        dc.coord = None  # reset para evitar deslocamento fantasma da atribuição original
        dc.cursor = dc.start_min
        tarefas_da_usina = by_usina[ativo]
        i = 0
        while i < len(tarefas_da_usina):
            tk = tarefas_da_usina[i]
            res = alloc_diurno(dc, tk['dur_h'],
                               prefer_morning=tk['corretiva'],
                               termo=tk['termo'],
                               coord=tk['coord'],
                               allow_partial=True)
            if res is None:
                i += 1
                continue
            start, end, desloc_h, used_h, completed = res
            used_min = int(round(used_h*60))
            is_first = (dc.coord != tk['coord'])
            rows.append(_row(equipe, di, tk, start, end,
                             desloc_h if is_first else 0.0,
                             partial=not completed,
                             used_h_override=used_h if not completed else None))
            commit_diurno(dc, start, end, used_min, tk['coord'], ativo)
            if completed:
                tarefas_da_usina.pop(i)
            else:
                tk['dur_h'] = tk['dur_h'] - used_h
                break

    # 4) Sobras: tentar reaproveitar capacidade em dias com sobra significativa
    # (se uma usina terminou rápido, podemos colocar tarefas de outra usina pendente
    # com mesma cidade? Não — regra de 1 usina/dia. Então sobras viram pendência.)
    for ativo, tarefas_restantes in by_usina.items():
        for tk in tarefas_restantes:
            # Verifica se essa usina foi atribuída a algum dia
            tem_dia = ativo in day_assignments
            motivo = ('Sem capacidade no(s) dia(s) atribuído(s) à usina'
                      if tem_dia else
                      f'Usina "{ativo}" não recebeu dia na semana (cluster com {len(usinas_carga)} usinas)')
            pendentes.append(_pend(equipe, tk, motivo))

    # v8 — Ajuste final da MPA: corretivas entram forçadas (EXCEDE HH) e podem
    # empurrar o expediente diurno para além das 15:30. Nesse caso o bloco de
    # MPA daquele dia é deslocado para começar depois, evitando duas atividades
    # no mesmo horário para a mesma equipe.
    for dc in day_caps:
        if not dc.mpa_rows:
            continue
        ini_mpa = min(s for _, s, _ in dc.mpa_rows)
        if dc.cursor <= ini_mpa:
            continue
        desl = int(dc.cursor - ini_mpa)
        for idx, s, e in dc.mpa_rows:
            ns, ne = s + desl, e + desl
            rows[idx]['Hora Início'] = t(ns)
            rows[idx]['Hora Fim']    = t(ne)
            rows[idx]['_ord']        = int(ns)
            if ne > MPA_END_MIN and 'EXCEDE JANELA MPA' not in rows[idx]['Dia']:
                rows[idx]['Dia'] = (rows[idx]['Dia'][:-1] + '/EXCEDE JANELA MPA]'
                                    if rows[idx]['Dia'].endswith(']')
                                    else rows[idx]['Dia'] + ' [EXCEDE JANELA MPA]')

    return rows, pendentes


def _row(equipe, day_idx, tk, start_min, end_min, desloc_h,
         noturno=False, zeladoria=False, partial=False, used_h_override=None,
         over_capacity=False):
    d = DAYS[day_idx]
    cat = tk['mp_cat']
    if zeladoria:
        tipo_lbl = 'Zeladoria'
    elif tk['corretiva']:
        tipo_lbl = 'Corretiva'
    elif cat == 'MPA':
        tipo_lbl = 'MPA'
    elif cat == 'MPM':
        # diferenciar inversor/mod-tracker no rótulo só para visualização
        txt = norm(tk['tarefa_txt'])
        if 'inversor' in txt: tipo_lbl = 'MPM-Inversor'
        elif any(k in txt for k in ['modulo','string e conjunto','tracker','mod fv']):
            tipo_lbl = 'MPM-Mod/Tracker'
        else: tipo_lbl = 'MPM'
    elif tk.get('tier') == TIER_INSPECAO:
        tipo_lbl = 'Inspeção'
    elif tk.get('tier') == TIER_ADMIN:
        tipo_lbl = 'Handover' if 'handover' in norm(tk['tarefa_txt']) else 'Administrativa'
    else:
        tipo_lbl = cat if (cat and cat != 'OUTRA') else 'Preventiva'

    flags = []
    if noturno: flags.append('NOTURNO')
    if zeladoria: flags.append('ZELADORIA/PARALELO')
    if partial: flags.append('PARCIAL')
    if over_capacity:
        flags.append('EXCEDE JANELA MPA' if noturno else 'EXCEDE HH')
    if tk.get('aging'):
        flags.append(f"+{tk.get('idade_dias') or AGING_DIAS}D EM ANDAMENTO")
    flag_txt = ' [' + '/'.join(flags) + ']' if flags else ''

    dur_h = used_h_override if used_h_override is not None else tk['dur_h']

    # Histórico
    h = historico.get(tk['task_key'], {})
    n_prev = h.get('count', 0)
    n_total = n_prev + 1

    return {
        'Equipe': equipe,
        'Dia': f"{DAY_NAMES[day_idx]} ({d.strftime('%d/%m')}){flag_txt}",
        'DiaIdx': day_idx,
        'OSs ID': tk['os_id'],
        'Ativo (Usina)': tk['ativo'],
        'Cidade': str(tk['cidade']).title() if tk['cidade'] else '',
        'Código Equipamento': tk['codigo'],
        'Tipo': tipo_lbl,
        'Tarefa': tk['tarefa_txt'],
        'Estado Tarefa (antes)': tk['estado_antes'],
        'RPN/Prioridade': tk['rpn'],
        'RPN (novo)': tk.get('rpn_novo'),                 # v9/M4 — sombra
        'RPN (como calculou)': tk.get('rpn_detalhe'),
        'Duração aprendida (h)': tk.get('dur_aprendida'),  # v9/M6 — sombra
        'Duração (base)': tk.get('dur_origem'),
        'Etiquetas': tk.get('etiquetas',''),
        'Reprogramada': 'Sim' if tk['reprog'] else 'Não',
        'Idade (dias)': tk.get('idade_dias'),
        'Nº vezes programada': n_total,
        'Termografia': 'Sim' if tk['termo'] else 'Não',
        'Hora Início': t(start_min),
        'Hora Fim': t(end_min),
        'Duração (h)': round(dur_h, 2),
        'Desloc (h)': round(desloc_h, 2),
        'Paralelo (terceirizada)': 'Sim' if zeladoria else 'Não',
        '_task_key': tk['task_key'],
        # v8: ordenação cronológica real. 'Hora Início' é string, então
        # "00:00 (D+1)" viria ANTES de "15:30" num sort textual.
        '_ord': int(start_min),
    }


def _pend(equipe, tk, motivo):
    if tk['zeladoria']:
        _tp = 'Zeladoria'
    elif tk['corretiva']:
        _tp = 'Corretiva'
    elif tk.get('tier') == TIER_INSPECAO:
        _tp = 'Inspeção'
    elif tk.get('tier') == TIER_ADMIN:
        _tp = 'Handover' if 'handover' in norm(tk['tarefa_txt']) else 'Administrativa'
    else:
        _tp = tk['mp_cat'] if tk['mp_cat'] != 'OUTRA' else 'Preventiva'
    if tk.get('aging'):
        motivo = f"[+{tk.get('idade_dias') or AGING_DIAS}d em andamento] {motivo}"
    return {
        'Equipe': equipe, 'OSs ID': tk['os_id'], 'Ativo': tk['ativo'],
        'Tarefa': tk['tarefa_txt'], 'Código': tk['codigo'],
        'Tipo': _tp,
        'Duração (h)': round(tk['dur_h'], 2), 'Motivo': motivo,
    }


# ====================== Execução ======================
all_rows, all_pend, resumo = [], [], []
equipes_com_tarefa = sorted(df_tasks['Equipe'].dropna().unique())

# === Equipes a SEMPRE incluir como aba (mesmo sem tarefa) ===
# Une: equipes com tarefa nesta semana + todas as equipes conhecidas no BD/AUXILIAR
equipes_todas = set(equipes_com_tarefa)
try:
    if 'Ativo Classificação 2' in df_bd.columns:
        for e in df_bd['Ativo Classificação 2'].dropna().unique():
            if isinstance(e, str) and e.strip():
                equipes_todas.add(norm_equipe(e))
except Exception:
    pass
try:
    df_aux_eq = fonte_bd_api.df_auxiliar()   # AUXILIAR - FABRICIO.xlsx (era aba AUXILIAR do BD)
    if 'Equipe Cluster' in df_aux_eq.columns:
        for e in df_aux_eq['Equipe Cluster'].dropna().unique():
            if isinstance(e, str) and e.strip():
                equipes_todas.add(e.strip())
except Exception:
    pass
equipes = sorted(e for e in equipes_todas if e)
print(f'Equipes com tarefa: {len(equipes_com_tarefa)} | Equipes-aba total: {len(equipes)}')

for equipe in equipes_com_tarefa:
    sub = df_tasks[df_tasks['Equipe'] == equipe].copy()
    tarefas = build_tarefas_atomicas(sub)
    rows, pend = schedule_team(equipe, tarefas)
    all_rows.extend(rows); all_pend.extend(pend)
    resumo.append({
        'Equipe': equipe,
        'Tarefas total': len(sub),
        'Usinas no cluster': sub['_ativo'].nunique(),
        'Usinas visitadas': len(set(r['Ativo (Usina)'] for r in rows)),
        'Tarefas agendadas': len(rows),
        'Tarefas pendentes': len(pend),
        'Corretivas': int(sub['corretiva'].sum()),
        'Zeladoria (paralela)': int(sub['zeladoria'].sum()),
        'MPA noturnas': int((sub['mp_cat']=='MPA').sum()),
    })

df_out = pd.DataFrame(all_rows)
df_pend = pd.DataFrame(all_pend)
df_resumo = pd.DataFrame(resumo)

# ── Melhoria 0.1 + decisão 2 ANTECIPADA EM SOMBRA (13/08/2026) ───────────────
# A validação da W34 mostrou que o deslocamento ENTRE tarefas do mesmo dia é
# estruturalmente ~0 (o motor raramente mistura usinas no dia). O custo real é
# a PRIMEIRA viagem do dia: base da equipe -> usina. Aqui ela é medida e
# escrita na coluna 'Desloc (h)' da primeira tarefa de cada (equipe, dia).
#
# É pós-processamento PURO da coluna de relatório: roda depois de toda a
# distribuição, portanto não move cursor nem capacidade — sombra por construção.
# Origem: 'Base Equipe' do AUXILIAR ("Cidade/UF"); coordenada da base vem por
# âncora = média das usinas cadastradas na mesma cidade (sem geocodificador).
BASES_SEM_ANCORA = {}
_DESLOC_BASE_TOT = {}
if DESLOC_CALCULA and len(df_out):
    def _norm_txt_b(s):
        import unicodedata as _ud
        s = _ud.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode().lower()
        return re.sub(r'[^a-z0-9]+', ' ', s).strip()
    try:
        _cid_coord = {}
        _cids = fonte_bd_api.mapa_cidades()
        for _u, _c in _cids.items():
            _xy = MAPA_COORD.get(_u)
            if _xy:
                _cid_coord.setdefault(_norm_txt_b(_c), []).append(_xy)
        _cid_coord = {k: (sum(x[0] for x in v)/len(v), sum(x[1] for x in v)/len(v))
                      for k, v in _cid_coord.items()}
        _df_aux_b = fonte_bd_api.df_auxiliar()
        _base_eq = {}
        for _, _r in _df_aux_b.iterrows():
            _eq = str(_r.get('Equipe Cluster') or '').strip()
            _b = str(_r.get('Base Equipe') or '').strip()
            if _eq and _b and _b.lower() != 'nan' and _eq not in _base_eq:
                _xy = _cid_coord.get(_norm_txt_b(_b.split('/')[0]))
                if _xy:
                    _base_eq[_eq] = _xy
                else:
                    BASES_SEM_ANCORA[_eq] = _b
        _n_base = 0
        for (_eq, _dia), _g in df_out.groupby(['Equipe', 'Dia'], sort=False):
            _bxy = _base_eq.get(str(_eq))
            if not _bxy:
                continue
            _gd = _g[_g.get('Paralelo (terceirizada)', '') != 'Sim'] if 'Paralelo (terceirizada)' in _g else _g
            if not len(_gd):
                continue
            _i0 = _gd['Hora Início'].astype(str).idxmin()
            _uxy = resolve_coord(str(df_out.at[_i0, 'Ativo (Usina)']))
            if not _uxy:
                continue
            _h = (haversine_km(_bxy, _uxy) / TRAVEL_KMH) * TRAVEL_FACTOR
            df_out.at[_i0, 'Desloc (h)'] = round(_h, 2)
            _DESLOC_BASE_TOT[_eq] = _DESLOC_BASE_TOT.get(_eq, 0.0) + _h
            _n_base += 1
        print(f"[DESLOC] viagem base→usina medida em {_n_base} dia(s)-equipe "
              f"({len(_base_eq)} base(s) com âncora, {len(BASES_SEM_ANCORA)} sem)")
    except Exception as _e:
        print(f"[DESLOC] AVISO: medição base→usina falhou ({type(_e).__name__}: {_e}) "
              "— coluna fica como estava. Nada mais é afetado.")

# Output Excel
with pd.ExcelWriter(OUTPUT, engine='openpyxl') as writer:
    df_resumo.to_excel(writer, sheet_name='_Resumo', index=False)
    # Aba _Observacoes (sempre, mesmo vazia, pra documentar)
    obs_rows = []
    if OBSERVACOES_SEMANA:
        for i, (_ln, obs) in enumerate(OBSERVACOES_SEMANA, start=1):
            obs_rows.append({"#": i, "Observação": obs})
    else:
        obs_rows.append({"#": 1, "Observação": "(Nenhuma observação para esta semana — edite Observacoes_Semana.txt na pasta base)"})
    pd.DataFrame(obs_rows).to_excel(writer, sheet_name='_Observacoes', index=False)
    if len(df_pend):
        df_pend.to_excel(writer, sheet_name='_Pendentes', index=False)
    # v9 — Portão de qualidade (0.7): sempre escreve a aba, mesmo sem achados,
    # para documentar que o portão RODOU (ausência de aba = portão não rodou).
    _qrows = QUALIDADE or [{'Tipo': 'OK', 'Item': '—',
                            'Detalhe': 'Nenhum achado de cadastro nesta geração',
                            'Ação': '—'}]
    pd.DataFrame(_qrows, columns=['Tipo', 'Item', 'Detalhe', 'Ação']).to_excel(
        writer, sheet_name='_Qualidade', index=False)
    # v8 bloco 2 — Plano do Mês: meta/feitas/faltam de preventivas por usina
    if len(PLACAR_MENSAL):
        _pm = PLACAR_MENSAL.copy()
        # quantas dessas o motor conseguiu colocar NESTA semana
        _prog = {}
        if len(df_out):
            _sem = df_out[~df_out['Tipo'].isin(['Corretiva', 'Zeladoria',
                                                'Inspeção', 'Administrativa', 'Handover'])]
            _prog = _sem.groupby('Ativo (Usina)').size().to_dict()
        _pm['Programadas nesta semana'] = _pm['_ativo'].map(lambda a: _prog.get(a, 0))
        _pm['Saldo após a semana'] = (_pm['Faltam'] - _pm['Programadas nesta semana']).clip(lower=0)
        _pm = _pm.rename(columns={'_ativo': 'Usina'})[
            ['Equipe', 'Usina', 'Meta', 'Feitas', 'Faltam',
             'Programadas nesta semana', 'Saldo após a semana', '% do mês']]
        _pm = _pm.sort_values(['Equipe', 'Saldo após a semana'], ascending=[True, False])
        _pm.to_excel(writer, sheet_name='_Plano do Mês', index=False)
    # Colunas padrão de uma aba de equipe (mesmo header das que têm tarefas)
    COLS_PADRAO_EQUIPE = [
        'Equipe','Dia','OSs ID','Ativo (Usina)','Responsável','Cidade',
        'Código Equipamento','Tipo','Tarefa','Estado Tarefa (antes)',
        'RPN/Prioridade','RPN (novo)','RPN (como calculou)',
        'Duração aprendida (h)','Duração (base)',
        'Etiquetas','Reprogramada','Idade (dias)','Nº vezes programada','Termografia',
        'Hora Início','Hora Fim','Duração (h)','Desloc (h)','Paralelo (terceirizada)'
    ]
    for equipe in equipes:
        sheet_nm = re.sub(r'[\\/*?:\[\]]', '', equipe)[:31]
        if not len(df_out):
            # Cria aba vazia só com cabeçalho
            pd.DataFrame(columns=COLS_PADRAO_EQUIPE).to_excel(writer, sheet_name=sheet_nm, index=False)
            continue
        sub = df_out[df_out['Equipe'] == equipe]
        if len(sub) == 0:
            # Equipe sem tarefa nesta semana — aba vazia com cabeçalho
            pd.DataFrame(columns=COLS_PADRAO_EQUIPE).to_excel(writer, sheet_name=sheet_nm, index=False)
            continue
        _ordcols = ['DiaIdx', '_ord'] if '_ord' in sub.columns else ['DiaIdx', 'Hora Início']
        sub = (sub.sort_values(_ordcols)
                  .drop(columns=[c for c in ['DiaIdx','_task_key','_ord'] if c in sub.columns]))
        # Inclui coluna Responsavel via lookup UFV -> RESPONSAVEL O&M (AUXILIAR)
        col_ativo = 'Ativo (Usina)' if 'Ativo (Usina)' in sub.columns else ('Ativo' if 'Ativo' in sub.columns else None)
        if col_ativo:
            sub.insert(min(len(sub.columns), 4), 'Responsável',
                       sub[col_ativo].apply(_resolver_responsavel))
        sub.to_excel(writer, sheet_name=sheet_nm, index=False)

# ====================== Atualização do histórico ======================
new_keys_this_week = set()
for r in all_rows:
    k = r['_task_key']
    new_keys_this_week.add(k)
    if k in historico:
        weeks = historico[k]['weeks'] or ''
        weeks_set = set(weeks.split(',')) if weeks else set()
        if CURRENT_WEEK not in weeks_set:
            weeks_set.add(CURRENT_WEEK)
            historico[k]['weeks'] = ','.join(sorted(weeks_set))
            historico[k]['count'] = len(weeks_set)
            historico[k]['last_week'] = CURRENT_WEEK
    else:
        historico[k] = {
            'first_week': CURRENT_WEEK,
            'last_week':  CURRENT_WEEK,
            'count':      1,
            'weeks':      CURRENT_WEEK,
        }

hist_rows = [{'task_key': k, **v} for k, v in historico.items()]
pd.DataFrame(hist_rows).to_excel(HISTORICO, index=False)

# Melhoria 0.1 — o resolve_coord antigo falhou calado por dois anos; este
# bloco existe para isso nunca se repetir.
print(f'\n[DESLOC] modo={DESLOC_MODO} · {len(MAPA_COORD)} usinas com coordenada no Fracttal')
if _DESLOC_BASE_TOT:
    _tot_sem = sum(_DESLOC_BASE_TOT.values())
    print(f'[DESLOC] viagem base→usina na semana: {_tot_sem:.1f} h somadas '
          f'({len(_DESLOC_BASE_TOT)} equipes) — capacidade hoje contada como zero')
    for _eq, _h in sorted(_DESLOC_BASE_TOT.items(), key=lambda x: -x[1])[:10]:
        print(f'         {_eq:<16} {_h:5.1f} h/semana')
if BASES_SEM_ANCORA:
    print(f'[DESLOC] {len(BASES_SEM_ANCORA)} base(s) sem âncora de coordenada '
          '(cidade da base não tem usina cadastrada com lat/lon):')
    for _eq, _b in sorted(BASES_SEM_ANCORA.items()):
        print(f'         - {_eq}: {_b}')
if USINAS_SEM_COORD:
    print(f'[DESLOC] {len(USINAS_SEM_COORD)} usina(s) SEM coordenada — trecho contado como 0 h:')
    for _u in sorted(USINAS_SEM_COORD):
        print(f'         - {_u}')
    print('         Cadastre latitude/longitud no Fracttal para que entrem no cálculo.')

print(f'\nOK -> {OUTPUT}')
print(f'Histórico atualizado -> {HISTORICO}')

# v9 (eco): por último, para ser a primeira coisa que se vê ao olhar o fim do log.
relatorio_observacoes()
print(df_resumo.to_string(index=False))
print(f'\nTotal: {len(df_out)} agendadas / {len(df_pend)} pendentes')
