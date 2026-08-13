# -*- coding: utf-8 -*-
"""
teste_semana_dourada.py — regressão do motor (Melhoria 0.3)
-----------------------------------------------------------
Congela os INSUMOS numa caixa de areia e roda o `programacao_v7.py` duas vezes
— antes e depois de uma edição — comparando as planilhas geradas célula a
célula. Se a edição não devia mudar a distribuição, a comparação tem que dar
ZERO diferenças ("em sombra, a distribuição tem que ser idêntica").

Uso:
    py -3 teste_semana_dourada.py gravar [--semana 34]   # baseline (antes de editar)
    py -3 teste_semana_dourada.py comparar               # depois de editar
    py -3 teste_semana_dourada.py limpar                 # descarta a caixa de areia

Como congela:
  - Copia p/ a caixa de areia (_dourada/sandbox): caches da API (.cache_semanal_api.pkl,
    _ativos_classificacao_cache.json), AUXILIAR, Confiabilidade, Lista de Prioridades,
    Feriados/, Observacoes_Semana.txt, Historico_Programacoes.xlsx e .env.
  - Roda o motor com PCM_PROG_DIR=sandbox e TTLs enormes: a API NUNCA é chamada,
    e o Historico de produção NUNCA é tocado (o motor escreve no da caixa de areia).
  - PCM_WEEK_FORCE fixa a semana; 'comparar' reusa a MESMA caixa de areia.

Limite honesto: baseline e comparação devem rodar NO MESMO DIA (a coluna
"Idade (dias)" usa a data de hoje). O script avisa se o baseline for antigo.
"""
import argparse
import datetime as dt
import os
import shutil
import subprocess
import sys

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

REPO = os.path.dirname(os.path.abspath(__file__))
DOURADA = os.path.join(REPO, "_dourada")
SANDBOX = os.path.join(DOURADA, "sandbox")
META = os.path.join(DOURADA, "meta.txt")
OUT_BASE = os.path.join(DOURADA, "baseline.xlsx")
OUT_ATUAL = os.path.join(DOURADA, "atual.xlsx")

# Pasta de dados de produção (OneDrive)
DADOS = os.path.join(os.path.expanduser("~"), "GRID CO", "GRID CO. - Gridco",
                     "4. O&M", "11.Pré-Operação", "6. PCM", "09. Programação Semanal")

INSUMOS = [".cache_semanal_api.pkl", "_ativos_classificacao_cache.json",
           "_usinas_coordenadas_cache.json",   # Melhoria 0.1 (13/08/2026)
           "AUXILIAR - FABRICIO.xlsx", "Planilha Confiabilidade R00.xlsx",
           "Lista_Prioridades_GridCo.xlsx", "Observacoes_Semana.txt",
           "Historico_Programacoes.xlsx", ".env"]


def log(m):
    print(f"[dourada] {m}", flush=True)


def montar_sandbox():
    os.makedirs(SANDBOX, exist_ok=True)
    faltou = []
    for f in INSUMOS:
        src = os.path.join(DADOS, f)
        if os.path.exists(src):
            shutil.copy2(src, os.path.join(SANDBOX, f))
        else:
            faltou.append(f)
    fer_src = os.path.join(DADOS, "Feriados")
    fer_dst = os.path.join(SANDBOX, "Feriados")
    if os.path.isdir(fer_src):
        shutil.copytree(fer_src, fer_dst, dirs_exist_ok=True)
    if ".cache_semanal_api.pkl" in faltou:
        sys.exit("ERRO: sem .cache_semanal_api.pkl na pasta de dados — rode uma "
                 "geração normal primeiro para aquecer o cache da API.")
    if faltou:
        log(f"aviso: insumos ausentes (seguem de fora): {', '.join(faltou)}")
    # congela: TTL gigante + mtime agora (o motor jamais chama a API daqui)
    agora = None
    for f in (".cache_semanal_api.pkl", "_ativos_classificacao_cache.json",
              "_usinas_coordenadas_cache.json"):
        p = os.path.join(SANDBOX, f)
        if os.path.exists(p):
            os.utime(p, agora)
    log(f"caixa de areia pronta: {SANDBOX}")


def rodar_motor(saida, semana):
    # O motor ESCREVE no Historico ao fim de cada rodada — sem restaurar, a 2ª
    # rodada veria tudo como "Reprogramada = Sim" (foi exatamente o que o
    # autoteste pegou). Congela: restaura o snapshot antes de CADA rodada.
    freeze = os.path.join(DOURADA, "Historico_freeze.xlsx")
    vivo = os.path.join(SANDBOX, "Historico_Programacoes.xlsx")
    if not os.path.exists(freeze) and os.path.exists(vivo):
        shutil.copy2(vivo, freeze)
    if os.path.exists(freeze):
        shutil.copy2(freeze, vivo)
    env = os.environ.copy()
    env["PCM_PROG_DIR"] = SANDBOX
    env["PCM_OUTPUT"] = saida
    env["PCM_WEEK_FORCE"] = str(semana)
    env["PROG_API_TTL_MIN"] = "999999"        # nunca expira dentro do teste
    env["GESTAO_ATIVOS_TTL_H"] = "999999"
    alvo = os.path.join(REPO, "programacao_v7.py")
    log(f"rodando o motor (semana {semana}) -> {os.path.basename(saida)}")
    r = subprocess.run([sys.executable, "-X", "utf8", alvo],
                       env=env, cwd=SANDBOX, capture_output=True,
                       text=True, encoding="utf-8", errors="replace", timeout=1800)
    if r.returncode != 0:
        print(r.stdout[-3000:])
        print(r.stderr[-3000:])
        sys.exit(f"ERRO: motor saiu com código {r.returncode}")
    # ecoa só o resumo
    for ln in r.stdout.splitlines():
        if ln.startswith(("Tarefas a programar", "Total:", "[FIM DE MÊS]", "[AGING]", "[PLANO")):
            log("  " + ln.strip())
    if not os.path.exists(saida):
        sys.exit("ERRO: o motor não gravou a planilha de saída.")


def carregar(xlsx):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    dados = {}
    for s in wb.sheetnames:
        dados[s] = [tuple(r) for r in wb[s].iter_rows(values_only=True)]
    wb.close()
    return dados


def comparar_planilhas(a_path, b_path):
    A, B = carregar(a_path), carregar(b_path)
    difs = []
    for s in sorted(set(A) | set(B)):
        if s not in A:
            difs.append(f"aba '{s}' só existe no ATUAL"); continue
        if s not in B:
            difs.append(f"aba '{s}' só existe no BASELINE"); continue
        la, lb = A[s], B[s]
        if len(la) != len(lb):
            difs.append(f"aba '{s}': {len(la)} vs {len(lb)} linhas")
        for i, (ra, rb) in enumerate(zip(la, lb)):
            if ra != rb:
                # aponta a primeira célula diferente da linha
                for j, (ca, cb) in enumerate(zip(ra, rb)):
                    if ca != cb:
                        difs.append(f"aba '{s}' linha {i+1} col {j+1}: {ca!r} -> {cb!r}")
                        break
                if len(difs) > 40:
                    difs.append("... (interrompido em 40 diferenças)")
                    return difs
    return difs


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("acao", choices=["gravar", "comparar", "limpar"])
    ap.add_argument("--semana", default=None,
                    help="semana ISO do teste (default: a próxima, gravada no meta)")
    a = ap.parse_args()

    if a.acao == "limpar":
        shutil.rmtree(DOURADA, ignore_errors=True)
        log("caixa de areia descartada.")
        return 0

    if a.acao == "gravar":
        shutil.rmtree(DOURADA, ignore_errors=True)
        os.makedirs(DOURADA, exist_ok=True)
        semana = a.semana or str((dt.date.today() + dt.timedelta(days=7)).isocalendar()[1])
        montar_sandbox()
        rodar_motor(OUT_BASE, semana)
        open(META, "w", encoding="utf-8").write(
            f"semana={semana}\ndata={dt.date.today().isoformat()}\n")
        log(f"BASELINE gravado (semana {semana}). Edite o motor e rode 'comparar'.")
        return 0

    # comparar
    if not os.path.exists(OUT_BASE):
        sys.exit("ERRO: sem baseline — rode 'gravar' antes de editar o motor.")
    meta = dict(ln.split("=", 1) for ln in open(META, encoding="utf-8").read().split())
    if meta.get("data") != dt.date.today().isoformat():
        log(f"AVISO: baseline é de {meta.get('data')} — a coluna 'Idade (dias)' "
            "vai diferir. Regrave o baseline hoje para uma comparação limpa.")
    rodar_motor(OUT_ATUAL, a.semana or meta["semana"])
    difs = comparar_planilhas(OUT_BASE, OUT_ATUAL)
    if not difs:
        log("IDÊNTICO — zero diferenças. Edição segura para a distribuição.")
        return 0
    log(f"{len(difs)} DIFERENÇA(S):")
    for d in difs:
        print("   " + d)
    return 1


if __name__ == "__main__":
    sys.exit(main())
